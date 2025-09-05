import logging
from pathlib import Path
from functools import wraps
from collections import Counter
from io import BytesIO
import json
import re
import math

import pandas as pd
import numpy as np
from urllib.parse import quote

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.forms import modelformset_factory
from django.http import (
    HttpResponse,
    HttpResponseForbidden,
    HttpResponseNotFound,
    JsonResponse,
)
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils.text import slugify

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

from ..filters import CardImagesFilter, PResponseFilter, SearchReferenceFilter
from ..forms import BulkResponseUploadForm, ResponseCodeForm
from ..models import (
    CardImages,
    Client,
    PopularResponse,
    ResponseCode,
    SearchReference,
    StructuralSummary,
)

from ._base import (
    group_min_required,
    normalize_card_to_num,
    to_roman,
)


TOTAL_CAP = 100
DEFAULT_EXTRA = 40

def _normalize_text_value(v):
    if v is None:
        return ''
    if not isinstance(v, str):
        return v
    s = v
    s = s.replace('\u00A0', ' ').replace('\u200b', '').replace('\u200c', '')
    s = s.strip()
    trans = {
        '“':'"', '”':'"', '‘':"'", '’':"'", '′':"'", '´':"'", '｀':"'",
        '·':'.', 'ㆍ':'.', '‧':'.', '•':'.',
        '：':':', ' ': ' ', '，':',', '．':'.', '／':'/', '－':'-',
    }
    s = ''.join(trans.get(ch, ch) for ch in s)
    return s

def _compact(s: str) -> str:
    s = (s or '').strip()
    s = re.sub(r'[:：]\s*$', '', s)
    return s.replace(' ', '').lower()

HEADER_MAP_COMPACT = {
    'id': None,
    '카드':'card','card':'card',
    'n':'response_num','응답수':'response_num','response_num':'response_num',
    '시간':'time','time':'time',
    '반응':'response','response':'response',
    '질문':'inquiry','inquiry':'inquiry',
    '회전':'rotation','v':'rotation','rotation':'rotation',
    '반응영역':'location','위치':'location','location':'location',
    '발달질':'dev_qual','dq':'dev_qual','dev_qual':'dev_qual','devqual':'dev_qual',
    '영역번호':'loc_num','loc_num':'loc_num','locnum':'loc_num',
    '결정인':'determinants','determinants':'determinants',
    '형태질':'form_qual','형태 질':'form_qual','form_qual':'form_qual','formquality':'form_qual',
    '(2)':'pair','2':'pair','pair':'pair',
    '내용인':'content','내용':'content','content':'content',
    'p':'popular','popular':'popular',
    'z':'Z','Z':'Z',
    '특수점수':'special','special':'special',
    '코멘트':'comment','메모':'comment','comment':'comment',
}

REQUIRED_FIELDS = [
    'card','response_num','time','response','inquiry','rotation','location',
    'dev_qual','loc_num','determinants','form_qual','pair','content','popular','Z','special','comment'
]

def normalize_header(h):
    if h is None:
        return None
    key = _compact(str(h))
    return HEADER_MAP_COMPACT.get(key, None)

def _pick_input_sheet(wb):
    for name in ('입력', 'responses'):
        if name in wb.sheetnames:
            return wb[name]
    for name in wb.sheetnames:
        ws_try = wb[name]
        raw = [(c.value or '') for c in ws_try[1]]
        mapped = [normalize_header(h) for h in raw]
        idx_map = {f: i for i, f in enumerate(mapped) if f}
        if all(f in idx_map for f in REQUIRED_FIELDS):
            return ws_try
    return wb.active

def _row_is_blank(row_tuple):
    return all((v is None) or (str(v).strip() == '') for v in row_tuple)

_TOKEN_SEP = r"[,\s;+/]+"

def _normalize_special_tokens(s: str) -> str:
    if not s:
        return s
    toks = re.split(_TOKEN_SEP, str(s).strip())
    toks = [t.strip().upper() for t in toks if t and t.strip()]
    return ", ".join(dict.fromkeys(toks))

def _detect_token(s: str, token: str) -> bool:
    if not s:
        return False
    return re.search(rf"(^|{_TOKEN_SEP}){re.escape(token)}($|{_TOKEN_SEP})", s, flags=re.IGNORECASE) is not None


@group_min_required('advanced')
def advanced_entry(request):
    cid = request.GET.get('client_id') or request.GET.get('client')
    if cid:
        return redirect('scoring:advanced_upload', client_id=cid)
    return redirect(f"{reverse('scoring:client_list')}?next=advanced")


@group_min_required('advanced')
def advanced_upload(request, client_id):
    client = get_object_or_404(Client, id=client_id)
    if client.tester != request.user:
        return HttpResponseForbidden("본인 수검자에게만 업로드할 수 있습니다.")

    existing_count = ResponseCode.objects.filter(client=client).count()
    has_existing = existing_count > 0

    form = BulkResponseUploadForm(request.user, request.POST or None, request.FILES or None)
    form.fields['client'].queryset = Client.objects.filter(id=client.id, tester=request.user)
    if request.method == 'GET' and not form.is_bound:
        form.initial['client'] = client

    if request.method == 'POST' and form.is_valid():
        posted_client = form.cleaned_data['client']
        if posted_client.id != client.id:
            messages.error(request, "요청 경로의 수검자와 폼의 수검자가 일치하지 않습니다.")
            return render(request, 'advanced_upload.html', {
                'client': client, 'form': form,
                'has_existing': has_existing, 'existing_count': existing_count,
            })

        xfile = form.cleaned_data['file']
        replace = form.cleaned_data['replace_existing']

        try:
            wb = load_workbook(filename=xfile, data_only=True)
            ws = _pick_input_sheet(wb)
        except Exception:
            messages.error(request, "엑셀 파일을 열 수 없습니다. (.xlsx 형식 확인)")
            return render(request, 'advanced_upload.html', {
                'client': client, 'form': form,
                'has_existing': has_existing, 'existing_count': existing_count,
            })

        raw_headers = [(c.value or '') for c in ws[1]]
        mapped = [normalize_header(h) for h in raw_headers]
        index_by_field = {f: idx for idx, f in enumerate(mapped) if f}
        missing = [f for f in REQUIRED_FIELDS if f not in index_by_field]
        if missing:
            messages.error(request, "필수 열이 누락되었습니다: " + ", ".join(missing))
            return render(request, 'advanced_upload.html', {
                'client': client, 'form': form,
                'has_existing': has_existing, 'existing_count': existing_count,
            })

        rows_all = list(ws.iter_rows(min_row=2, values_only=True))
        rows = [r for r in rows_all if not _row_is_blank(r)]
        if not rows:
            messages.warning(request, "업로드 가능한 데이터 행이 없습니다.")
            return render(request, 'advanced_upload.html', {
                'client': client, 'form': form,
                'has_existing': has_existing, 'existing_count': existing_count,
            })

        allow = TOTAL_CAP if replace else max(0, TOTAL_CAP - existing_count)
        if allow <= 0:
            messages.error(request, f"이미 {TOTAL_CAP}행이 저장되어 있어 더 추가할 수 없습니다.")
            return redirect('scoring:client_detail', client_id=client.id)

        trimmed = False
        if len(rows) > allow:
            rows = rows[:allow]
            trimmed = True

        created, errors = 0, []
        with transaction.atomic():
            if replace:
                ResponseCode.objects.filter(client=client).delete()

            for ridx, row in enumerate(rows, start=2):
                data = {}
                for f in REQUIRED_FIELDS:
                    idx = index_by_field[f]
                    v = row[idx] if idx < len(row) else ''
                    v = '' if v is None else v
                    v = _normalize_text_value(v)
                    data[f] = v

                data['card'] = to_roman(data.get('card', ''))
                for int_key in ('response_num', 'loc_num'):
                    txt = str(data.get(int_key, '')).strip()
                    if txt == '':
                        data[int_key] = ''
                    else:
                        try:
                            data[int_key] = int(float(txt))
                        except Exception:
                            errors.append(f"{ridx}행: {int_key} 정수 변환 실패")

                def _std_symbols(s):
                    if not s:
                        return s
                    toks = [t.strip() for t in re.split(r'[;,]', str(s)) if t.strip()]
                    toks = [t.replace(' .', '.').replace('. ', '.') for t in toks]
                    seen, out = set(), []
                    for t in toks:
                        if t not in seen:
                            seen.add(t); out.append(t)
                    return ', '.join(out)

                for sym_key in ('determinants','form_qual','content','special'):
                    data[sym_key] = _std_symbols(data.get(sym_key, ''))

                form_row = ResponseCodeForm(data)
                if not form_row.is_valid():
                    errs = '; '.join([f"{fld}: {', '.join(e)}" for fld, e in form_row.errors.items()])
                    errors.append(f"{ridx}행 유효성 오류 → {errs}")
                    continue

                obj = form_row.save(commit=False)
                obj.client = client
                obj.card = to_roman(obj.card)
                obj.save()
                created += 1

        if trimmed:
            messages.warning(request, f"총 {TOTAL_CAP}행 제한으로 앞 {allow}행만 처리했습니다.")
        if errors:
            messages.warning(request, f"성공 {created}건, 오류 {len(errors)}건")
        else:
            messages.success(request, f"성공 {created}건 업로드 완료")

        return redirect('scoring:client_detail', client_id=client.id)

    return render(request, 'advanced_upload.html', {
        'client': client, 'form': form,
        'has_existing': has_existing, 'existing_count': existing_count,
    })


RESOURCE_DIR = Path(getattr(settings, 'SCORING_RESOURCE_DIR',
                            Path(__file__).resolve().parent.parent / 'resources')).resolve()

DEFAULT_RESOURCE_FILENAMES = {
    'symbol_score':   'symbol_score_mapping.json',
    'response_score': 'response_token_scores_mapping.json',
    'inquiry_score':  'inquiry_token_scores_mapping.json',
    'score_stats':    'scoring_area_card_stats.json',
    'index_stats':    'projection_index_card_stats.json',
}
RESOURCE_FILENAMES = getattr(settings, 'SCORING_RESOURCE_FILENAMES', DEFAULT_RESOURCE_FILENAMES)

def _read_json_df(path: Path, required_cols=None) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"필요한 JSON 데이터 파일이 없습니다: {path}")

    def _ensure_required(df: pd.DataFrame) -> pd.DataFrame:
        if not required_cols:
            return df
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            raise ValueError(
                f"{path.name} 에 필요한 열이 없습니다: {', '.join(missing)} "
                f"(사용 가능한 열: {', '.join(map(str, getattr(df, 'columns', [])))} )"
            )
        return df

    for kwargs in ({}, {'lines': True}):
        try:
            df_try = pd.read_json(path, dtype=False, **kwargs)
            if isinstance(df_try, pd.DataFrame) and not df_try.empty:
                rename = {
                    'card':'카드','Card':'카드','card_id':'카드',
                    'token':'토큰','Token':'토큰',
                    'pos':'품사','POS':'품사',
                    'score':'점수','Score':'점수'
                }
                df_try = df_try.rename(columns=rename)
                return _ensure_required(df_try)
        except Exception:
            pass

    raw = json.loads(path.read_text(encoding='utf-8'))

    if isinstance(raw, dict):
        rows = []

        # 카드 → {mean,std}
        if all(isinstance(v, dict) and set(v.keys()) & {'mean', 'std'} for v in raw.values()):
            for card, obj in raw.items():
                rows.append({'카드': str(card), 'mean': obj.get('mean'), 'std': obj.get('std')})
            return _ensure_required(pd.DataFrame(rows))

        # 카드 → 영역 → {mean,std}
        is_area_stats = any(
            isinstance(areas, dict) and any(isinstance(v, dict) and {'mean','std'} <= set(v.keys())
                                            for v in areas.values())
            for areas in raw.values()
        )
        if is_area_stats:
            for card, areas in raw.items():
                if not isinstance(areas, dict):
                    continue
                for area_name, stat in areas.items():
                    if isinstance(stat, dict) and {'mean','std'} <= set(stat.keys()):
                        rows.append({'카드': str(card), '채점영역': str(area_name),
                                     'mean': stat.get('mean'), 'std': stat.get('std')})
            return _ensure_required(pd.DataFrame(rows))

        # 카드 → 영역 → {기호: 점수}
        is_symbol_scores = any(isinstance(areas, dict) and any(isinstance(v, dict) for v in areas.values())
                               for areas in raw.values())
        if is_symbol_scores:
            for card, areas in raw.items():
                if not isinstance(areas, dict):
                    continue
                for area_name, symbols in areas.items():
                    if isinstance(symbols, dict):
                        for sym, sc in symbols.items():
                            rows.append({'카드': str(card), '채점영역': str(area_name),
                                         '기호': str(sym), '점수': sc})
            return _ensure_required(pd.DataFrame(rows))

        for key in ('records','data','rows'):
            if isinstance(raw.get(key), list):
                df = pd.DataFrame(raw[key])
                return _ensure_required(df)

        raise ValueError(f"{path.name} 을(를) DataFrame으로 변환할 수 없습니다.")

    if isinstance(raw, list):
        df = pd.DataFrame(raw)
        rename = {
            'card':'카드','Card':'카드','card_id':'카드',
            'token':'토큰','Token':'토큰',
            'pos':'품사','POS':'품사',
            'score':'점수','Score':'점수'
        }
        df = df.rename(columns=rename)
        return _ensure_required(df)

    raise ValueError(f"{path.name} 을(를) DataFrame으로 변환할 수 없습니다.")


STOPWORDS = set([
    '이','그','저','나','너','그것','이것','저것','들','\n','때','것','그리고','하지만','또는','즉','그렇지','그래서','그러므로',
    '대해','대하여','위해','때문에','그런데','근데','이런','저런','그런','같은','처럼','듯','도','만','또','조차','까지',
    '네','예','수검자','검사자','있다','보이다','Q','A','반응반복','같다','하다','반응','반복','부분','여기','이렇게','거','그렇다','어떻다','얘','보다'
])

try:
    from konlpy.tag import Okt
    _OKT = Okt()

    def _preprocess_text(text: str) -> str:
        text = re.sub(r'[^\w\sㄱ-ㅎㅏ-ㅣ가-힣]', ' ', str(text))
        text = re.sub(r'\s+', ' ', text).strip()
        return text

    def tokenize_with_pos(text: str):
        text = _preprocess_text(text)
        target_pos = {'Noun','Verb','Adjective','Adverb'}
        return [(w, p) for (w, p) in _OKT.pos(text, stem=True) if (p in target_pos and w not in STOPWORDS)]
except Exception:
    def _preprocess_text(text: str) -> str:
        text = re.sub(r'[^\w\sㄱ-ㅎㅏ-ㅣ가-힣]', ' ', str(text))
        text = re.sub(r'\s+', ' ', text).strip()
        return text
    def tokenize_with_pos(text: str):
        text = _preprocess_text(text)
        toks = [t for t in re.split(r'\s+', text) if t and t not in STOPWORDS]
        toks = [t for t in toks if len(t) >= 2]
        return [(t, 'Noun') for t in toks]


def _apply_pair_into_determinants(row):
    det = str(row.get('결정인') or '').strip()
    val2 = row.get('(2)')
    val2 = (str(int(val2)).strip() if (val2 not in (None, '', ' ') and not pd.isna(val2)) else '')
    det_tokens = [x.strip() for x in re.split(r'[,.]', det) if x.strip()]
    if val2 and val2 not in det_tokens:
        det_tokens.append(val2)
    row['결정인'] = ', '.join(det_tokens) if det_tokens else np.nan
    return row

def _calculate_token_score(df, token_freq_df, token_column_name, score_column_name):
    tf = token_freq_df.copy()
    tf['토큰_튜플'] = list(zip(tf['토큰'], tf['품사']))
    token_score_map = tf.set_index(['카드', '토큰_튜플'])['점수'].to_dict()

    def compute_row(row):
        card = str(row['카드'])
        tokens = row.get(token_column_name) or []
        uniq = set(tokens)
        total = 0
        for tok in uniq:
            total += token_score_map.get((card, tok), 2)
        return total

    df[score_column_name] = df.apply(compute_row, axis=1)
    return df

def _apply_symbol_score(df, score_table_df):
    areas = ['결정인', '내용인', '특수점수']
    score_dict = {col: [] for col in areas}
    lookup = score_table_df.set_index(['카드','채점영역','기호'])['점수'].to_dict()

    for _, row in df.iterrows():
        card = str(row['카드'])
        for area in areas:
            score = 0
            cell = row.get(area)
            if pd.isna(cell) or cell in (None, ''):
                score_dict[area].append(0)
                continue
            symbols = [s.strip() for s in str(cell).split(',') if s.strip()]
            for sym in symbols:
                score += lookup.get((card, area, sym), 2)
            score_dict[area].append(score)

    for area in areas:
        df[f'{area}_점수'] = score_dict[area]
    return df

def _count_col_shd_blends(codes):
    color = {'FC', 'CF', 'C'}
    shading = {"FC'", "C'F", "C'", 'FT', 'TF', 'T', 'FV', 'VF', 'V', 'FY', 'YF', 'Y'}
    cnt = 0
    for rc in codes:
        det = _normalize_text_value(getattr(rc, 'determinants', '') or '')
        if not det:
            continue
        has_color = any(_detect_token(det, t) for t in color)
        has_shading = any(_detect_token(det, t) for t in shading)
        if has_color and has_shading:
            cnt += 1
    return cnt

@group_min_required('advanced')
def export_structural_summary_xlsx_advanced(request, client_id):
    try:
        client = Client.objects.get(id=client_id)
        if client.tester != request.user:
            return HttpResponse("액세스 거부: 해당 정보를 볼 수 있는 권한이 없습니다.", status=403)

        response_codes = ResponseCode.objects.filter(client_id=client_id)

        numbers_found = {normalize_card_to_num(rc.card) for rc in response_codes if rc.card}
        required = {str(i) for i in range(1, 11)}
        missing = sorted(required - numbers_found, key=int)
        if missing:
            missing_roman = [to_roman(n) for n in missing]
            return HttpResponse("다음 카드의 반응이 없습니다: " + ", ".join(missing_roman))

        structural_summary, _ = StructuralSummary.objects.get_or_create(client_id=client_id)
        try:
            structural_summary.calculate_values()
        except Exception:
            structural_summary.save()

    except Client.DoesNotExist:
        logging.error("해당 ID의 클라이언트를 찾을 수 없음")
        return HttpResponseNotFound("클라이언트 정보를 찾을 수 없습니다.")
    except Exception as e:
        logging.error(f"예기치 못한 오류 발생: {e}")
        return JsonResponse({'error': f"{type(e).__name__}: {str(e)}"}, status=500)

    PASTEL_FILL = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
    LINE_COLOR  = "FFB7B7B7" 
    THIN_EDGE   = Side(border_style='thin', color=LINE_COLOR)
    HDR_FONT    = Font(bold=True)

    def box_border(ws, cell_range, line_style="thin", color="FFB7B7B7"):
        rows = list(ws[cell_range])
        if not rows:
            return
        edge = Side(style=line_style, color=color)
        max_y = len(rows) - 1
        for y, row in enumerate(rows):
            max_x = len(row) - 1
            for x, cell in enumerate(row):
                b = cell.border
                cell.border = Border(
                    left=edge if x == 0 else b.left,
                    right=edge if x == max_x else b.right,
                    top=edge if y == 0 else b.top,
                    bottom=edge if y == max_y else b.bottom,
                )

    def header_cell(ws, addr: str, value: str):
        ws[addr] = value
        ws[addr].fill = PASTEL_FILL
        ws[addr].font = HDR_FONT
        ws[addr].alignment = Alignment(horizontal='center', vertical='center')

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    ws   = wb.create_sheet(title='상단부')
    wsd  = wb.create_sheet(title='하단부')
    wsi  = wb.create_sheet(title='특수지표')
    ws_raw = wb.create_sheet(title='반응별 정보')

    ws.sheet_view.showGridLines = False
    ws.merge_cells('A4:B4'); ws.merge_cells('A14:B14'); ws.merge_cells('A20:D20')
    ws.merge_cells('F4:H4'); ws.merge_cells('G5:H5'); ws.merge_cells('J4:K4')
    ws.merge_cells('M4:N4'); ws.merge_cells('M16:P16')

    header_cell(ws, 'A4',  'Location Features')
    header_cell(ws, 'A14', 'Developmental Quality')
    header_cell(ws, 'A20', 'Form Quality')
    header_cell(ws, 'F4',  'Determinants')
    header_cell(ws, 'F5',  'Blends')
    header_cell(ws, 'G5',  'Single')
    header_cell(ws, 'J4',  'Contents')
    header_cell(ws, 'M4',  'approach')
    header_cell(ws, 'M16', 'Special Scores')

    for pos in [
        'A4:B4','A5:B12','A14:B14','A15:B18','A20:D20','A21:D26',
        'F4:H4','F5:F5','G5:H5','F6:F29','G6:H29','J4:K4','J5:K31',
        'M4:N4','M5:N14','M16:P16','M17:P30'
    ]:
        box_border(ws, pos)
    ws['A5']='Zf'; ws['A6']='Zsum'; ws['A7']='Zest'
    ws['A9']='W'; ws['A10']='D'; ws['A11']='Dd'; ws['A12']='S'
    ws['A15']='+'; ws['A16']='o'; ws['A17']='v/+'; ws['A18']='v'
    ws['B21']='FQx'; ws['C21']='MQual'; ws['D21']='W+D'
    ws['A22']='+'; ws['A23']='o'; ws['A24']='u'; ws['A25']='-'; ws['A26']='none'

    ws['B5']  = structural_summary.Zf
    ws['B6']  = structural_summary.Zsum; ws['B6'].number_format = '0.0'
    ws['B7']  = structural_summary.Zest; ws['B7'].number_format = '0.0'
    ws['B9']  = structural_summary.W;   ws['B10'] = structural_summary.D
    ws['B11'] = structural_summary.Dd;  ws['B12'] = structural_summary.S

    ws['B15'] = structural_summary.dev_plus
    ws['B16'] = structural_summary.dev_o
    ws['B17'] = structural_summary.dev_vplus
    ws['B18'] = structural_summary.dev_v

    ws['B22'] = structural_summary.fqx_plus
    ws['B23'] = structural_summary.fqx_o
    ws['B24'] = structural_summary.fqx_u
    ws['B25'] = structural_summary.fqx_minus
    ws['B26'] = structural_summary.fqx_none

    ws['C22'] = structural_summary.mq_plus
    ws['C23'] = structural_summary.mq_o
    ws['C24'] = structural_summary.mq_u
    ws['C25'] = structural_summary.mq_minus
    ws['C26'] = structural_summary.mq_none

    ws['D22'] = structural_summary.wd_plus
    ws['D23'] = structural_summary.wd_o
    ws['D24'] = structural_summary.wd_u
    ws['D25'] = structural_summary.wd_minus
    ws['D26'] = structural_summary.wd_none

    blends_str = getattr(structural_summary, 'blends', '') or ''
    blends_list = [b.strip() for b in str(blends_str).split(',') if b.strip()]

    row = 6 
    for b in blends_list:
        ws.cell(row=row, column=6, value=b)  # column=6 → 'F'
        row += 1
    
    fields = ['M','FM','m',"FC","CF","C","Cn","FC'","C'F","C'",
              'FT','TF','T','FV','VF','V','FY','YF','Y','Fr','rF','FD','F','(2)']
    real_field = ['M','FM','m_l','FC','CF','C','Cn','FCa','CaF','Ca',
                  'FT','TF','T','FV','VF','V','FY','YF','Y','Fr','rF','FD','F','pair']
    r = 6
    for name in fields:
        ws.cell(row=r, column=7, value=name); r += 1
    r = 6
    for fname in real_field:
        ws.cell(row=r, column=8, value=getattr(structural_summary, fname)); r += 1

    cont_names = ['H','(H)','Hd','(Hd)','Hx','A','(A)','Ad','(Ad)','An',
                  'Art','Ay','Bl','Bt','Cg','Cl','Ex','Fd','Fi','Ge','Hh','Ls',
                  'Na','Sc','Sx','Xy','Id']
    cont_real  = ['H','H_paren','Hd','Hd_paren','Hx','A','A_paren','Ad','Ad_paren','An',
                  'Art','Ay','Bl','Bt','Cg','Cl','Ex','Fd_l','Fi','Ge','Hh','Ls',
                  'Na','Sc','Sx','Xy','Idio']
    r = 5
    for name in cont_names:
        ws.cell(row=r, column=10, value=name); r += 1
    r = 5
    for fname in cont_real:
        ws.cell(row=r, column=11, value=getattr(structural_summary, fname)); r += 1

    header_cell(ws, 'M4', 'approach')
    for i, label in enumerate(['I','II','III','IV','V','VI','VII','VIII','IX','X'], start=5):
        ws.cell(row=i, column=13, value=label)
    ws['N5']=structural_summary.app_I;   ws['N6']=structural_summary.app_II
    ws['N7']=structural_summary.app_III; ws['N8']=structural_summary.app_IV
    ws['N9']=structural_summary.app_V;   ws['N10']=structural_summary.app_VI
    ws['N11']=structural_summary.app_VII;ws['N12']=structural_summary.app_VIII
    ws['N13']=structural_summary.app_IX; ws['N14']=structural_summary.app_X

    header_cell(ws, 'M16', 'Special Scores')
    ws['N17']='Lvl-1' 
    ws['O17']='Lvl-2'
    for i, name in enumerate(['DV','INC','DR','FAB','ALOG','CON'], start=18):
        ws.cell(row=i, column=13, value=name)  # M열 라벨

    for i, fn in enumerate(['sp_dv','sp_inc','sp_dr','sp_fab','sp_alog','sp_con'], start=18):
        c = ws.cell(row=i, column=14, value=getattr(structural_summary, fn))  # N열 값
        c.number_format = '0'

    for i, fn in enumerate(['sp_dv2','sp_inc2','sp_dr2','sp_fab2'], start=18):
        c = ws.cell(row=i, column=15, value=getattr(structural_summary, fn))  # O열 값
        c.number_format = '0'
    ws['M24'] = 'Raw Sum6';      ws['N24'] = structural_summary.sum6
    ws['M25'] = 'Weighted Sum6'; ws['N25'] = structural_summary.wsum6
    ws['M26'] = 'AB';            ws['N26'] = structural_summary.sp_ab
    ws['M27'] = 'AG';            ws['N27'] = structural_summary.sp_ag
    ws['M28'] = 'COP';           ws['N28'] = structural_summary.sp_cop
    ws['M29'] = 'CP';            ws['N29'] = structural_summary.sp_cp
    for r in range(24, 30):
        ws.cell(row=r, column=14).number_format = '0'  # N열 값 포맷

    ws['O26'] = 'GHR'; ws['P26'] = structural_summary.sp_ghr
    ws['O27'] = 'PHR'; ws['P27'] = structural_summary.sp_phr
    ws['O28'] = 'MOR'; ws['P28'] = structural_summary.sp_mor
    ws['O29'] = 'PER'; ws['P29'] = structural_summary.sp_per
    ws['O30'] = 'PSV'; ws['P30'] = structural_summary.sp_psv
    for r in range(26, 31):
        ws.cell(row=r, column=16).number_format = '0'  # P열 값 포맷

    for col_cells in ws.columns:
        length = max(len(str(c.value)) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = max(8, min(32, int(length*1.2)))

    wsd.sheet_view.showGridLines = False
    wsd.merge_cells('A3:F3'); wsd.merge_cells('H3:I3'); wsd.merge_cells('K3:N3')
    wsd.merge_cells('A14:D14'); wsd.merge_cells('F14:G14'); wsd.merge_cells('I14:J14'); wsd.merge_cells('L14:M14')

    header_cell(wsd, 'A3',  'Core')
    header_cell(wsd, 'H3',  'Affect')
    header_cell(wsd, 'K3',  'Interpersonal')
    header_cell(wsd, 'A14', 'Ideation')
    header_cell(wsd, 'F14', 'Mediation')
    header_cell(wsd, 'I14', 'Processing')
    header_cell(wsd, 'L14', 'Self')

    for pos in ['A3:F3','A4:F4','A5:F7','A8:F9','H3:I3','H4:I10','K3:N3','K4:N12',
                'A14:D14','A15:D19','F14:G14','F15:G21','I14:J14','I15:J21','L14:M14','L15:M21']:
        box_border(wsd, pos)

    wsd['A4']='R'; wsd['C4']='L'
    wsd['A5']='EB'; wsd['A6']='eb'
    wsd['C5']='EA'; wsd['C6']='es'; wsd['C7']='Adj es'
    wsd['E5']='EBper'; wsd['E6']='D'; wsd['E7']='Adj D'
    wsd['A8']='FM'; wsd['A9']='m'; wsd['C8']="SumC'"; wsd['C9']='SumV'
    wsd['E8']='SumT'; wsd['E9']='SumY'

    wsd['B4']=structural_summary.R;         wsd['D4']=structural_summary.L
    wsd['B5']=structural_summary.ErleBnistypus
    wsd['B6']=structural_summary.eb
    wsd['D5']=structural_summary.EA;        wsd['D6']=structural_summary.es; wsd['D7']=structural_summary.adj_es
    wsd['F5']='NA' if structural_summary.EBper == 0 else structural_summary.EBper
    wsd['F6']=structural_summary.D_score;   wsd['F7']=structural_summary.adj_D
    wsd['B8']=structural_summary.sum_FM;    wsd['B9']=structural_summary.sum_m
    wsd['D8']=structural_summary.sum_Ca;    wsd['D9']=structural_summary.sum_V
    wsd['F8']=structural_summary.sum_T;     wsd['F9']=structural_summary.sum_Y

    wsd['H4']='FC:CF+C'; wsd['H5']='Pure C'; wsd['H6']="SumC':WsumC"
    wsd['H7']='Afr';     wsd['H8']='S';       wsd['H9']='Blends:R'; wsd['H10']='CP'
    wsd['I4']=structural_summary.f_c_prop;    wsd['I5']=structural_summary.pure_c
    wsd['I6']=structural_summary.ca_c_prop;   wsd['I7']=structural_summary.afr; wsd['I7'].number_format="0.##;-0.##;0"
    wsd['I8']=structural_summary.S;          wsd['I9']=structural_summary.blends_r
    wsd['I10']=structural_summary.sp_cp

    wsd['K4']='COP'; wsd['M4']='AG'
    wsd['K5']='GHR:PHR'; wsd['K6']='a:p'; wsd['K7']='Food'
    wsd['K8']='SumT';    wsd['K9']='Human Content'; wsd['K10']='Pure H'
    wsd['K11']='PER';    wsd['K12']='Isolation Index'
    wsd['L4']=structural_summary.sp_cop;   wsd['N4']=structural_summary.sp_ag
    wsd['M5']=structural_summary.GHR_PHR;  wsd['M6']=structural_summary.a_p
    wsd['M7']=structural_summary.Fd_l;     wsd['M8']=structural_summary.sum_T
    wsd['M9']=structural_summary.human_cont; wsd['M10']=structural_summary.H
    wsd['M11']=structural_summary.sp_per;    wsd['M12']=structural_summary.Isol

    wsd['A15']='a:p'; wsd['A16']='Ma:Mp'; wsd['A17']='Intel(2AB+Art+Ay)'; wsd['A18']='MOR'
    wsd['C15']='Sum6'; wsd['C16']='Lvl-2'; wsd['C17']='Wsum6'; wsd['C18']='M-'; wsd['C19']='M none'
    wsd['B15']=structural_summary.a_p;   wsd['B16']=structural_summary.Ma_Mp
    wsd['B17']=structural_summary.intel; wsd['B18']=structural_summary.sp_mor
    wsd['D15']=structural_summary.sum6;  wsd['D16']=structural_summary.Lvl_2
    wsd['D17']=structural_summary.wsum6; wsd['D18']=structural_summary.mq_minus; wsd['D19']=structural_summary.mq_none

    for i, (lbl, attr) in enumerate(
        [('XA%','xa_per'),('WDA%','wda_per'),('X-%','x_minus_per'),('S-','s_minus'),
         ('P','popular'),('X+%','x_plus_per'),('Xu%','xu_per')], start=15):
        v = getattr(structural_summary, attr)
        c = wsd.cell(row=i, column=7, value=v)
        c.number_format = "0" if isinstance(v, int) else "0.00"
    for i, attr in enumerate(['Zf','W_D_Dd','W_M','Zd','sp_psv','dev_plus','dev_v'], start=15):
        wsd.cell(row=i, column=10, value=getattr(structural_summary, attr))
    for i, attr in enumerate(['ego','fr_rf','sum_V','fdn','an_xy','sp_mor','h_prop'], start=15):
        wsd.cell(row=i, column=13, value=getattr(structural_summary, attr))

    row0 = 22
    def cb(txt, pos): return f"☑ {txt}" if pos else txt
    pti_pos  = (structural_summary.sumPTI >= 3)
    depi_pos = (structural_summary.sumDEPI >= 5)
    cdi_pos  = (structural_summary.sumCDI  >= 4)
    scon_pos = (structural_summary.sumSCON >= 8)
    hvi_pos  = (structural_summary.sumHVI  >= 4) and bool(structural_summary.HVI_premise)
    obs_pos  = bool(structural_summary.OBS_posi)
    obs_score = sum(1 for ch in (structural_summary.OBS or '') if ch == 'o')

    wsd.cell(row=row0, column=1,  value=cb(f"PTI={structural_summary.sumPTI}", pti_pos)).font = HDR_FONT
    # wsd.cell(row=row0, column=5,  value=cb("HVI", hvi_pos)).font = HDR_FONT
    hvi_cell = wsd.cell(row=row0, column=3, value=cb(f"HVI={structural_summary.sumHVI}", hvi_pos))
    hvi_cell.font = HDR_FONT
    wsd.cell(row=row0, column=6,  value=cb(f"DEPI={structural_summary.sumDEPI}", depi_pos)).font = HDR_FONT
    wsd.cell(row=row0, column=9,  value=cb(f"OBS={obs_score}", obs_pos)).font = HDR_FONT
    wsd.cell(row=row0, column=12, value=cb(f"CDI={structural_summary.sumCDI}", cdi_pos)).font = HDR_FONT
    wsd.cell(row=row0, column=15, value=cb(f"S-CON={structural_summary.sumSCON}", scon_pos)).font = HDR_FONT
    for col in (1,3,6,9,12,15):
        wsd.cell(row=row0, column=col).alignment = Alignment(horizontal='center')

    try:
        df_raw = pd.DataFrame([{
            'ID': client.id,
            '카드': normalize_card_to_num(rc.card),
            'N': rc.response_num,
            '반응': rc.response,
            '질문': rc.inquiry,
            '결정인': rc.determinants,
            '(2)': rc.pair,
            '내용인': rc.content,
            '특수점수': (special := _normalize_special_tokens(rc.special or '')),
            'Card': to_roman(rc.card),
            'time': rc.time, 'V': rc.rotation, 'Location': rc.location, 'loc_num': rc.loc_num,
            'Dev Qual': rc.dev_qual, 'Form Quality': rc.form_qual,
            'P': rc.popular, 'Z': rc.Z,
        } for rc in response_codes])

        df_proc = df_raw[['ID','카드','N','반응','질문','결정인','(2)','내용인','특수점수']].copy()
        df_proc = df_proc.apply(_apply_pair_into_determinants, axis=1).drop(columns=['(2)'])

        # 리소스 읽기
        rsp_score = _read_json_df(RESOURCE_DIR / RESOURCE_FILENAMES['response_score'],
                                  required_cols=['카드','토큰','품사','점수'])
        inq_score = _read_json_df(RESOURCE_DIR / RESOURCE_FILENAMES['inquiry_score'],
                                  required_cols=['카드','토큰','품사','점수'])
        sym_score = _read_json_df(RESOURCE_DIR / RESOURCE_FILENAMES['symbol_score'],
                                  required_cols=['카드','채점영역','기호','점수'])
        sc_stats  = _read_json_df(RESOURCE_DIR / RESOURCE_FILENAMES['score_stats'],
                                  required_cols=['카드','채점영역','mean','std'])
        idx_stats = _read_json_df(RESOURCE_DIR / RESOURCE_FILENAMES['index_stats'],
                                  required_cols=['카드','mean','std'])
        for _df in (rsp_score, inq_score, sym_score, sc_stats, idx_stats):
            _df['카드'] = _df['카드'].astype(str)

        # 토큰화
        df_proc['RESPONSE_토큰'] = df_raw['반응'].apply(lambda x: list(set(tokenize_with_pos(x))))
        df_proc['INQUIRY_토큰']  = df_raw['질문'].apply(lambda x: list(set(tokenize_with_pos(x))))

        # 점수화
        df_sc = _calculate_token_score(df_proc.copy(), rsp_score, 'RESPONSE_토큰', 'RESPONSE_점수')
        df_sc = _calculate_token_score(df_sc,          inq_score,  'INQUIRY_토큰',  'INQUIRY_점수')
        df_sc  = _apply_symbol_score(df_sc, sym_score)

        mean_df = sc_stats.pivot(index='카드', columns='채점영역', values='mean')
        std_df  = sc_stats.pivot(index='카드', columns='채점영역', values='std')
        for a in ['RESPONSE_점수','INQUIRY_점수','결정인_점수','내용인_점수','특수점수_점수']:
            colz = f'{a}_z'
            if a in df_sc.columns and a in mean_df.columns:
                def _z(r, aa=a):
                    c = str(r['카드'])
                    try:
                        m = float(mean_df.loc[c, aa]); s = float(std_df.loc[c, aa])
                        return (r[aa]-m)/s if s else 0.0
                    except Exception:
                        return 0.0
                df_sc[colz] = df_sc.apply(_z, axis=1)
            else:
                df_sc[colz] = 0.0
        z_cols = ['결정인_점수_z','내용인_점수_z','특수점수_점수_z','INQUIRY_점수_z','RESPONSE_점수_z']
        df_sc['투사지수_final'] = df_sc[z_cols].sum(axis=1)

        # T변환
        df_sc = df_sc.merge(idx_stats, on='카드', how='left')
        df_sc['std']  = df_sc['std'].replace(0, np.nan).fillna(1.0)
        df_sc['mean'] = df_sc['mean'].fillna(0.0)
        df_sc['투사지수_T'] = 50 + 10*(df_sc['투사지수_final'] - df_sc['mean'])/df_sc['std']
        df_sc = df_sc.drop(columns=['mean','std'])

        # 카드별/전체 평균
        df_card_avg = (df_sc.groupby(['ID','카드'], as_index=False)
                          .agg(투사지수_T평균=('투사지수_T','mean')))
        df_card_avg['카드'] = pd.to_numeric(df_card_avg['카드'], errors='coerce').fillna(0).astype(int)
        df_card_avg = df_card_avg.sort_values(['ID','카드'])
        overall_t = float(df_card_avg['투사지수_T평균'].mean()) if not df_card_avg.empty else float(df_sc['투사지수_T'].mean())

        row1 = row0 + 2
        cell_tp = wsd.cell(row=row1, column=1, value='투사지표')
        cell_tp.font = HDR_FONT
        cell_tp.fill = PASTEL_FILL   # 헤더 색 채우기
        c_avg = wsd.cell(row=row1, column=2, value=round(overall_t, 2)); c_avg.number_format = "0.00"
        box_border(wsd, f"A{row1}:B{row1}")
        
        row2 = row1 + 2
        # 헤더
        wsd.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=2)
        tcell = wsd.cell(row=row2, column=1, value='카드별 투사점수'); tcell.fill = PASTEL_FILL; tcell.font = HDR_FONT
        tcell.alignment = Alignment(horizontal='center')
        # 목록
        r = row2 + 1
        t_map = {int(k): float(v) for k, v in zip(df_card_avg['카드'], df_card_avg['투사지수_T평균'])}
        for n in range(1, 11):
            wsd.cell(row=r, column=1, value=to_roman(str(n)))
            c = wsd.cell(row=r, column=2, value=None if t_map.get(n) is None else round(t_map.get(n), 2))
            c.number_format = "0.00"; r += 1
        box_border(wsd, f"A{row2}:B{r-1}")
        df_sc = df_sc.drop(columns=['반응', '질문', '결정인', '내용인', '특수점수'], errors='ignore')

        df_out = (df_sc.merge(
                    df_raw[['카드','N','Card','time','반응','질문','V','Location','Dev Qual',
                            'loc_num','결정인','Form Quality','(2)','내용인','P','Z','특수점수']],
                    on=['카드','N'], how='left')
                 )
        df_out['카드'] = pd.to_numeric(df_out['카드'], errors='coerce').fillna(0).astype(int)
        df_out = df_out.sort_values(['카드','N'], kind='mergesort')

        cols = ['카드','Card','N','time','반응','질문','V','Location','Dev Qual','loc_num',
                '결정인','Form Quality','(2)','내용인','P','Z','특수점수','투사지수_T']
        for i, name in enumerate(cols, start=1):
            c = ws_raw.cell(row=1, column=i, value=name)
            c.font = HDR_FONT; c.fill = PASTEL_FILL
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = Border(top=THIN_EDGE, bottom=THIN_EDGE, left=THIN_EDGE, right=THIN_EDGE)

        for rowv in df_out[cols].values.tolist():
            ws_raw.append(rowv)

        ws_raw.freeze_panes = "A2"
        ws_raw.auto_filter.ref = f"A1:{get_column_letter(ws_raw.max_column)}1"

        for col in range(1, ws_raw.max_column + 1):
            max_len = 0
            for r_ in range(1, ws_raw.max_row + 1):
                v = ws_raw.cell(row=r_, column=col).value
                max_len = max(max_len, len(str(v)) if v is not None else 0)
            ws_raw.column_dimensions[get_column_letter(col)].width = max(8, min(80, int(max_len * 1.1)))

    except Exception as e:
        logging.exception("고급 산출 실패")
        ws_err = wb.create_sheet(title='advanced_error')
        ws_err['A1'] = "고급 산출 과정에서 오류가 발생했습니다."
        ws_err['A2'] = f"{type(e).__name__}: {str(e)}"
        ws_err.column_dimensions['A'].width = 120

    header_cell(wsi, 'A1', "PTI"); header_cell(wsi, 'A9', "DEPI"); header_cell(wsi, 'A20', "CDI")
    header_cell(wsi, 'D1', "S-CON"); header_cell(wsi, 'D17', "HVI"); header_cell(wsi, 'A29', "OBS")

    wsi['A2']="XA%<.70 AND WDA%<.75"; wsi['A3']="X-%>0.29"; wsi['A4']="LVL2>2 AND FAB2>0"
    wsi['A5']="R<17 AND Wsum6>12 OR R>16 AND Wsum6>17*"; wsi['A6']="M- > 1 OR X-% > 0.40"; wsi['A7']="TOTAL"
    r = 2
    for i in range(0,5):
        wsi.cell(row=r, column=2, value=("✔" if structural_summary.PTI[i] == "o" else "")); r += 1
    wsi['B7'] = structural_summary.sumPTI

    wsi['A10']="SumV>0 OR FD>2"; wsi['A11']="Col-shd blends>0 OR S>2"
    wsi['A12']="ego sup AND Fr+rF=0 OR ego inf"; wsi['A13']="Afr<0.46 OR Blends<4"
    wsi['A14']="SumShd>FM+m OR SumC'>2"; wsi['A15']="MOR>2 OR INTELL>3"; wsi['A16']="COP<2 OR ISOL>0.24"
    wsi['A17']="TOTAL"; wsi['A18']="POSITIVE?"
    r = 10
    for i in range(0,7):
        wsi.cell(row=r, column=2, value=("✔" if structural_summary.DEPI[i] == "o" else "")); r += 1
    wsi['B17']=structural_summary.sumDEPI; wsi['B18']=structural_summary.sumDEPI >= 5

    wsi['A21']="EA<6 OR Daj<0"; wsi['A22']="COP<2 AND AG<2"; wsi['A23']="WSumC<2.5 OR Afr<0.46"
    wsi['A24']="p > a+1 OR pure H<2"; wsi['A25']="SumT>1 OR ISOL>0.24 OR Fd>0"
    wsi['A26']="TOTAL"; wsi['A27']="POSITIVE?"
    r = 21
    for i in range(0,5):
        wsi.cell(row=r, column=2, value=("✔" if structural_summary.CDI[i] == "o" else "")); r += 1
    wsi['B26']=structural_summary.sumCDI; wsi['B27']=structural_summary.sumCDI >= 4

    labels = ["SumV+FD>2","col-shd blends>0","ego <0.31 ou >0.44","mor>3","Zd>3.5 ou <-3.5",
              "es>EA","CF+C>FC","X+%<0.70","S>3","P<3 OU P>8","PURE H<2","R<17"]
    for i, txt in enumerate(labels, start=2):
        wsi.cell(row=i, column=4, value=txt)
        wsi.cell(row=i, column=5, value=("✔" if structural_summary.SCON[i-2] == "o" else ""))
    wsi['E14']=structural_summary.sumSCON; wsi['E15']=structural_summary.sumSCON >= 8

    hvi_txt = ['SumT = 0','Zf>12','Zd>3.5','S>3','H+(H)+Hd+(Hd)>6','(H)+(A)+(Hd)+(Ad)>3','H+A : 4:1','Cg>3']
    wsi['E18']=structural_summary.HVI_premise
    for i, txt in enumerate(hvi_txt[1:], start=19):  # 19~25
        wsi.cell(row=i, column=4, value=txt)
        wsi.cell(row=i, column=5, value=("✔" if structural_summary.HVI[i-19] == "o" else ""))
    wsi['E26']=structural_summary.sumHVI
    wsi['E27']=(structural_summary.sumHVI >= 4) and bool(structural_summary.HVI_premise)

    obs_l = [(1,"Dd>3"),(2,"Zf>12"),(3,"Zd>3.0"),(4,"P>7"),(5,"FQ+>1")]
    for i,(n,txt) in enumerate(obs_l, start=30):
        wsi.cell(row=i, column=1, value=n); wsi.cell(row=i, column=2, value=txt)
        wsi.cell(row=i, column=3, value=("✔" if structural_summary.OBS[i-30] == "o" else ""))
    obs2 = ["1-5 are true","FQ+>3 AND 2 items 1-4","X+%>0,89 et 3 items","FQ+>3 et X+%>0,89"]
    for i, txt in enumerate(obs2, start=30):
        wsi.cell(row=i, column=4, value=txt)
        wsi.cell(row=i, column=5, value=("✔" if structural_summary.OBS[i+5-30] == "o" else ""))
    wsi['E34']=structural_summary.OBS_posi

    for col_cells in wsi.columns:
        length = max(len(str(c.value)) for c in col_cells)
        wsi.column_dimensions[col_cells[0].column_letter].width = max(8, min(40, int(length*1.2)))

    col_shd_blends_total = _count_col_shd_blends(response_codes)
    
    wsdev = wb.create_sheet(title='이탈정도')
    wsdev.sheet_view.showGridLines = False

    TITLE = "규준자료 대비 지표별 백분위 이탈정도 계산파일(국제규준)"
    wsdev.merge_cells('A1:H1')
    title_cell = wsdev['A1']
    title_cell.value = TITLE
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    wsdev.row_dimensions[1].height = 40

    headers = ['No', '변인', '점수', '평균(국제)', '표준편차', 'Z', '%', '구분(국제)']
    HDR_FILL = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
    for i, h in enumerate(headers, start=1):
        c = wsdev.cell(row=2, column=i, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = Border(top=THIN_EDGE, bottom=THIN_EDGE, left=THIN_EDGE, right=THIN_EDGE)

    def _safe_get(obj, name, default=0.0):
        try:
            v = getattr(obj, name)
        except Exception:
            return default
        return default if v is None else v

    def _norm_cdf(z: float) -> float:
        return 0.5 * (1.0 + math.erf(z / math.sqrt(2.0)))

    def _parse_ratio_pair(text):
        s = str(text or '').strip()
        try:
            a, b = s.split(':', 1)
            a = float(a.strip() or 0)
            b = float(b.strip() or 0)
            return a, b
        except Exception:
            try:
                x = float(s)
                return x, 1.0
            except Exception:
                return 0.0, 0.0

    CAT_COLOR = {
        '매우낮음': '92CDDC',
        '낮음'   : 'B7DEE8',
        '평균하' : 'DAEEF3',
        '평균'   : 'E4DFEC',
        '평균상' : 'F2DCDB',
        '높음'   : 'E6B8B7',
        '매우높음': 'DA9694',
    }
    def _grade_label(p: float):
        if p < 0.05:  return '매우낮음'
        if p < 0.12:  return '낮음'
        if p < 0.25:  return '평균하'
        if p < 0.75:  return '평균'
        if p < 0.88:  return '평균상'
        if p < 0.95:  return '높음'
        return '매우높음'

    def _get_active(ss):
        a, p = _parse_ratio_pair(_safe_get(ss, 'a_p', '0:0'))
        return a
    def _get_passive(ss):
        a, p = _parse_ratio_pair(_safe_get(ss, 'a_p', '0:0'))
        return p
    def _get_ma(ss):
        a, b = _parse_ratio_pair(_safe_get(ss, 'Ma_Mp', '0:0'))
        return a
    def _get_mp(ss):
        a, b = _parse_ratio_pair(_safe_get(ss, 'Ma_Mp', '0:0'))
        return b
    def _get_sumc(ss):
        return float(_safe_get(ss, 'FC', 0)) + float(_safe_get(ss, 'CF', 0)) + float(_safe_get(ss, 'C', 0)) + float(_safe_get(ss, 'Cn', 0))
    def _get_wsumc(ss):
        v = getattr(ss, 'WsumC', None)
        if v is None:
            v = _safe_get(ss, 'sum_Ca', 0)
        return float(v)
    def _get_sumc_prime(ss):
        return float(_safe_get(ss, 'sum_Ca', 0))
    def _get_sumsh(ss):
        return float(_get_sumc_prime(ss)) + float(_safe_get(ss, 'sum_T', 0)) + float(_safe_get(ss, 'sum_V', 0)) + float(_safe_get(ss, 'sum_Y', 0))
    def _get_blends_count(ss):
        v = getattr(ss, 'blends_r', None)
        if v is None:
            return float(getattr(ss, 'blends', 0) or 0)
        a, b = _parse_ratio_pair(v)
        return a
    def _get_blends_ratio(ss):
        v = getattr(ss, 'blends_r', None)
        a, b = _parse_ratio_pair(v)
        return (a / b) if b else 0.0
    def _get_human_all(ss):
        return float(_safe_get(ss, 'human_cont', 0))

    GET = {
        'R':        lambda ss: _safe_get(ss, 'R', 0),
        'W':        lambda ss: _safe_get(ss, 'W', 0),
        'D':        lambda ss: _safe_get(ss, 'D', 0),
        'Dd':       lambda ss: _safe_get(ss, 'Dd', 0),
        'S':        lambda ss: _safe_get(ss, 'S', 0),

        'DQ+':      lambda ss: _safe_get(ss, 'dev_plus', 0),
        'DQo':      lambda ss: _safe_get(ss, 'dev_o', 0),
        'DQv':      lambda ss: _safe_get(ss, 'dev_v', 0),
        'DQv/+':    lambda ss: _safe_get(ss, 'dev_vplus', 0),

        'FQ+':      lambda ss: _safe_get(ss, 'fqx_plus', 0),
        'FQo':      lambda ss: _safe_get(ss, 'fqx_o', 0),
        'FQu':      lambda ss: _safe_get(ss, 'fqx_u', 0),
        'FQ-':      lambda ss: _safe_get(ss, 'fqx_minus', 0),
        'FQnone':   lambda ss: _safe_get(ss, 'fqx_none', 0),

        'MQ+':      lambda ss: _safe_get(ss, 'mq_plus', 0),
        'Mqo':      lambda ss: _safe_get(ss, 'mq_o', 0),
        'Mqu':      lambda ss: _safe_get(ss, 'mq_u', 0),
        'MQ-':      lambda ss: _safe_get(ss, 'mq_minus', 0),
        'Mqnone':   lambda ss: _safe_get(ss, 'mq_none', 0),

        'S-':       lambda ss: _safe_get(ss, 's_minus', 0),

        'M':        lambda ss: _safe_get(ss, 'M', 0),
        'FM':       lambda ss: _safe_get(ss, 'sum_FM', _safe_get(ss, 'FM', 0)),
        'm':        lambda ss: _safe_get(ss, 'sum_m', 0),
        'FM+m':     lambda ss: _safe_get(ss, 'sum_FM', _safe_get(ss, 'FM', 0)) + _safe_get(ss, 'sum_m', 0),

        'FC':       lambda ss: _safe_get(ss, 'FC', 0),
        'CF':       lambda ss: _safe_get(ss, 'CF', 0),
        'C':        lambda ss: _safe_get(ss, 'C', 0),
        'Cn':       lambda ss: _safe_get(ss, 'Cn', 0),

        'SumC':     _get_sumc,
        'WSumC':    _get_wsumc,
        "SumC'":    _get_sumc_prime,
        'SumT':     lambda ss: _safe_get(ss, 'sum_T', 0),
        'SumV':     lambda ss: _safe_get(ss, 'sum_V', 0),
        'SumY':     lambda ss: _safe_get(ss, 'sum_Y', 0),
        'SumSh':    _get_sumsh,

        'Fr+rF':    lambda ss: _safe_get(ss, 'fr_rf', 0),
        'FD':       lambda ss: _safe_get(ss, 'FD', 0),

        'F':        lambda ss: _safe_get(ss, 'F', 0),
        '2':        lambda ss: _safe_get(ss, 'pair', 0),

        '3r+2/R':   lambda ss: _safe_get(ss, 'ego', 0),
        'Lambda':   lambda ss: _safe_get(ss, 'Lambda', 0),
        'EA':       lambda ss: _safe_get(ss, 'EA', 0),
        'es':       lambda ss: _safe_get(ss, 'es', 0),
        'D score':  lambda ss: _safe_get(ss, 'D_score', 0),
        'Adj D':    lambda ss: _safe_get(ss, 'adj_D', 0),

        'active':   _get_active,
        'passive':  _get_passive,
        'Ma':       _get_ma,
        'Mp':       _get_mp,
        'Intellect':lambda ss: _safe_get(ss, 'intel', 0),

        'Zf':       lambda ss: _safe_get(ss, 'Zf', 0),
        'Zd':       lambda ss: _safe_get(ss, 'Zd', 0),

        'Blends':   _get_blends_count,
        'Blends/R': _get_blends_ratio,
        'Col-Shd Blends': lambda ss, v=col_shd_blends_total: v,

        'Afr':      lambda ss: _safe_get(ss, 'afr', 0),

        'Popular':  lambda ss: _safe_get(ss, 'popular', 0),
        'XA%':      lambda ss: _safe_get(ss, 'xa_per', 0.0),
        'WDA%':     lambda ss: _safe_get(ss, 'wda_per', 0.0),
        'X+%':      lambda ss: _safe_get(ss, 'x_plus_per', 0.0),
        'X-%':      lambda ss: _safe_get(ss, 'x_minus_per', 0.0),
        'Xu%':      lambda ss: _safe_get(ss, 'xu_per', 0.0),

        'Isolate/R':lambda ss: _safe_get(ss, 'Isol', 0),

        'H':        lambda ss: _safe_get(ss, 'H', 0),
        '(H)':      lambda ss: _safe_get(ss, 'H_paren', 0),
        'Hd':       lambda ss: _safe_get(ss, 'Hd', 0),
        '(Hd)':     lambda ss: _safe_get(ss, 'Hd_paren', 0),
        'Hx':       lambda ss: _safe_get(ss, 'Hx', 0),
        'All H cont': _get_human_all,
        'A':        lambda ss: _safe_get(ss, 'A', 0),
        '(A)':      lambda ss: _safe_get(ss, 'A_paren', 0),
        'Ad':       lambda ss: _safe_get(ss, 'Ad', 0),
        '(Ad)':     lambda ss: _safe_get(ss, 'Ad_paren', 0),
        'An':       lambda ss: _safe_get(ss, 'An', 0),
        'Art':      lambda ss: _safe_get(ss, 'Art', 0),
        'Ay':       lambda ss: _safe_get(ss, 'Ay', 0),
        'Bl':       lambda ss: _safe_get(ss, 'Bl', 0),
        'Bt':       lambda ss: _safe_get(ss, 'Bt', 0),
        'Cg':       lambda ss: _safe_get(ss, 'Cg', 0),
        'Cl':       lambda ss: _safe_get(ss, 'Cl', 0),
        'Ex':       lambda ss: _safe_get(ss, 'Ex', 0),
        'Fi':       lambda ss: _safe_get(ss, 'Fi', 0),
        'Fd':       lambda ss: _safe_get(ss, 'Fd_l', 0),
        'Ge':       lambda ss: _safe_get(ss, 'Ge', 0),
        'Hh':       lambda ss: _safe_get(ss, 'Hh', 0),
        'Ls':       lambda ss: _safe_get(ss, 'Ls', 0),
        'Na':       lambda ss: _safe_get(ss, 'Na', 0),
        'Sc':       lambda ss: _safe_get(ss, 'Sc', 0),
        'Sx':       lambda ss: _safe_get(ss, 'Sx', 0),
        'Xy':       lambda ss: _safe_get(ss, 'Xy', 0),
        'Id':       lambda ss: _safe_get(ss, 'Idio', 0),

        'DV':       lambda ss: _safe_get(ss, 'sp_dv', 0),
        'INC':      lambda ss: _safe_get(ss, 'sp_inc', 0),
        'DR':       lambda ss: _safe_get(ss, 'sp_dr', 0),
        'FAB':      lambda ss: _safe_get(ss, 'sp_fab', 0),
        'DV2':      lambda ss: _safe_get(ss, 'sp_dv2', 0),
        'INC2':     lambda ss: _safe_get(ss, 'sp_inc2', 0),
        'DR2':      lambda ss: _safe_get(ss, 'sp_dr2', 0),
        'FAB2':     lambda ss: _safe_get(ss, 'sp_fab2', 0),
        'ALOG':     lambda ss: _safe_get(ss, 'sp_alog', 0),
        'CONTAM':   lambda ss: _safe_get(ss, 'sp_con', 0),

        'Sum6':     lambda ss: _safe_get(ss, 'sum6', 0),
        'Lvl 2 Sp Sc': lambda ss: _safe_get(ss, 'Lvl_2', 0),
        'Wsum6':    lambda ss: _safe_get(ss, 'wsum6', 0),
        'AB':       lambda ss: _safe_get(ss, 'sp_ab', 0),
        'AG':       lambda ss: _safe_get(ss, 'sp_ag', 0),
        'COP':      lambda ss: _safe_get(ss, 'sp_cop', 0),
        'CP':       lambda ss: _safe_get(ss, 'sp_cp', 0),
        'GHR':      lambda ss: _safe_get(ss, 'sp_ghr', 0),
        'PHR':      lambda ss: _safe_get(ss, 'sp_phr', 0),
        'MOR':      lambda ss: _safe_get(ss, 'sp_mor', 0),
        'PER':      lambda ss: _safe_get(ss, 'sp_per', 0),
        'PSV':      lambda ss: _safe_get(ss, 'sp_psv', 0),
    }

    NORM_SPECS = [
        ('R', 22.31, 7.90), ('W', 9.08, 4.54), ('D', 9.89, 5.81), ('Dd', 3.33, 3.37),
        ('S', 2.49, 2.15),
        ('DQ+', 6.24, 3.54), ('DQo', 14.68, 6.74), ('DQv', 1.09, 1.50), ('DQv/+', 0.29, 0.67),
        ('FQ+', 0.21, 0.68), ('FQo', 11.11, 3.74), ('FQu', 6.20, 3.93), ('FQ-', 4.43, 3.23), ('FQnone', 0.33, 0.71),
        ('MQ+', 0.12, 0.43), ('Mqo', 2.26, 1.66), ('Mqu', 0.69, 0.99), ('MQ-', 0.63, 1.05), ('Mqnone', 0.03, 0.20),
        ('S-', 0.87, 1.15),

        ('M', 3.73, 2.66), ('FM', 3.37, 2.18), ('m', 1.50, 1.54), ('FM+m', 4.87, 2.89),
        ('FC', 1.91, 1.70), ('CF', 1.65, 1.55), ('C', 0.34, 0.66), ('Cn', 0.02, 0.14),
        ('SumC', 3.91, 2.53), ('WSumC', 3.11, 2.17), ("SumC'", 1.75, 1.71),
        ('SumT', 0.65, 0.91), ('SumV', 0.52, 0.92), ('SumY', 1.34, 1.63), ('SumSh', 4.29, 3.48),

        ('Fr+rF', 0.41, 0.88), ('FD', 1.02, 1.19), ('F', 8.92, 5.34), ('2', 7.04, 3.83),
        ('3r+2/R', 0.38, 0.16), ('Lambda', 0.86, 0.95), ('EA', 6.84, 3.76), ('es', 9.09, 5.04),
        ('D score', -0.68, 1.48), ('Adj D', -0.20, 1.23),

        ('active', 4.96, 3.08), ('passive', 3.73, 2.65), ('Ma', 2.09, 1.83), ('Mp', 1.67, 1.61),
        ('Intellect', 2.35, 2.57),

        ('Zf', 12.50, 4.92), ('Zd', -0.67, 4.72),

        ('Blends', 4.01, 2.97), ('Blends/R', 0.18, 0.13), ('Col-Shd Blends', 0.60, 0.92),
        ('Afr', 0.53, 0.20),

        ('Popular', 5.36, 1.84), ('XA%', 0.79, 0.11), ('WDA%', 0.82, 0.11), ('X+%', 0.52, 0.13), ('X-%', 0.19, 0.11), ('Xu%', 0.27, 0.11),
        ('Isolate/R', 0.20, 0.14),

        ('H', 2.43, 1.89), ('(H)', 1.22, 1.24), ('Hd', 1.52, 1.71), ('(Hd)', 0.64, 0.92), ('Hx', 0.41, 0.98),
        ('All H cont', 5.83, 3.51), ('A', 7.71, 3.18), ('(A)', 0.42, 0.73), ('Ad', 2.41, 1.97), ('(Ad)', 0.16, 0.45),
        ('An', 1.16, 1.42), ('Art', 1.22, 1.45), ('Ay', 0.52, 0.87), ('Bl', 0.25, 0.55), ('Bt', 1.41, 1.44),
        ('Cg', 1.89, 1.77), ('Cl', 0.18, 0.46), ('Ex', 0.19, 0.48), ('Fi', 0.50, 0.80), ('Fd', 1.02, 1.19),
        ('Ge', 0.26, 0.62), ('Hh', 0.84, 1.03), ('Ls', 0.87, 1.12), ('Na', 0.75, 1.11), ('Sc', 1.11, 1.35),
        ('Sx', 0.47, 0.94), ('Xy', 0.19, 0.52), ('Id', 0.89, 1.21),

        ('DV', 0.65, 0.99), ('INC', 0.73, 0.97), ('DR', 0.49, 0.96), ('FAB', 0.45, 0.76),
        ('DV2', 0.01, 0.14), ('INC2', 0.10, 0.33), ('DR2', 0.06, 0.31), ('FAB2', 0.08, 0.31),
        ('ALOG', 0.16, 0.46), ('CONTAM', 0.02, 0.13),

        ('Sum6', 2.75, 2.39), ('Lvl 2 Sp Sc', 0.25, 0.62), ('Wsum6', 7.63, 7.75),
        ('AB', 0.32, 0.82), ('AG', 0.54, 0.86), ('COP', 1.07, 1.18), ('CP', 0.02, 0.15),
        ('GHR', 3.70, 2.18), ('PHR', 2.86, 2.52), ('MOR', 1.26, 1.43), ('PER', 0.75, 1.12), ('PSV', 0.23, 0.56),
    ]

    r = 3
    for idx, (name, mean, std) in enumerate(NORM_SPECS, start=1):
        getter = GET.get(name, lambda ss: None)
        score = getter(structural_summary)
        try:
            score = float(score)
        except Exception:
            score = None

        if (score is None) or (std is None) or (float(std) == 0):
            z = None
            p = None
        else:
            z = (score - float(mean)) / float(std)
            p = _norm_cdf(z)

        wsdev.cell(row=r, column=1, value=idx)             # No
        wsdev.cell(row=r, column=2, value=name)            # 변인
        wsdev.cell(row=r, column=3, value=score)           # 점수
        wsdev.cell(row=r, column=4, value=float(mean))     # 평균(국제)
        wsdev.cell(row=r, column=5, value=float(std))      # 표준편차
        wsdev.cell(row=r, column=6, value=(0.0 if z is None else z))  # Z
        pct_cell = wsdev.cell(row=r, column=7, value=(0.0 if p is None else p))  # %
        grade_cell = wsdev.cell(row=r, column=8)

        DEC2_KEYS = {"Blends/R", "Afr", "XA%", "WDA%", "X+%", "X-%", "Xu%", "Isolate/R"}
        if score is None:
            wsdev.cell(row=r, column=3).number_format = '@'
        else:
            if name in DEC2_KEYS:
                wsdev.cell(row=r, column=3, value=round(score, 2)).number_format = '0.00'
            else:
                wsdev.cell(row=r, column=3, value=int(round(score))).number_format = '0'

        wsdev.cell(row=r, column=4).number_format = '0.00'
        wsdev.cell(row=r, column=5).number_format = '0.00'
        wsdev.cell(row=r, column=6).number_format = '0.00'
        pct_cell.number_format = '0.0%'

        if p is not None:
            label = _grade_label(p)
            grade_cell.value = label
            color = CAT_COLOR.get(label)
            if color:
                grade_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        r += 1

    last_row = wsdev.max_row

    IDX_FILL  = PatternFill(start_color="EEECE1", end_color="EEECE1", fill_type="solid")  # A열
    META_FILL = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")  # B,D,E,F,G
    POINT_FILL= PatternFill(start_color="FABF8F", end_color="FABF8F", fill_type="solid")  # C열

    for rr in range(3, last_row + 1):
        wsdev.cell(row=rr, column=1).fill = IDX_FILL
        for cc in (2, 4, 5, 6, 7):
            wsdev.cell(row=rr, column=cc).fill = META_FILL
        wsdev.cell(row=rr, column=3).fill = POINT_FILL

    for rr in range(2, last_row + 1):
        for cc in range(1, 9):
            cell = wsdev.cell(row=rr, column=cc)
            cell.border = Border(left=THIN_EDGE, right=THIN_EDGE, top=THIN_EDGE, bottom=THIN_EDGE)
            if rr >= 3 and cc in (1, 3, 4, 5, 6, 7):
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif rr >= 3 and cc == 2:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    widths = [6, 14, 9, 12, 10, 7, 8, 12]  # A~H
    for i, w in enumerate(widths, start=1):
        wsdev.column_dimensions[get_column_letter(i)].width = w

    wsdev.auto_filter.ref = f"A2:H{last_row}"
    wsdev.freeze_panes = "A3"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    safe_name = f"{client.name}_{client.testDate:%Y-%m-%d}.xlsx"
    fallback  = f"{slugify(client.name)}_{client.testDate:%Y-%m-%d}.xlsx"

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{fallback}"; filename*=UTF-8\'\'{quote(safe_name)}'
    )
    return response


@group_min_required('advanced')
def advanced_edit_responses(request, client_id):
    client = get_object_or_404(Client, id=client_id)
    if client.tester != request.user:
        return HttpResponse("액세스 거부: 작성 권한이 없습니다.", status=403)

    qs = ResponseCode.objects.filter(client=client).order_by('id')
    current = qs.count()
    extra = max(0, min(DEFAULT_EXTRA, TOTAL_CAP - current))
    FormSet = modelformset_factory(ResponseCode, form=ResponseCodeForm, extra=extra, max_num=TOTAL_CAP)

    if request.method == 'POST':
        formset = FormSet(request.POST, queryset=qs)
        if formset.is_valid():
            saved = 0
            for form in formset:
                if form.cleaned_data.get('card') and form.cleaned_data.get('response'):
                    inst = form.save(commit=False)
                    inst.client = client
                    inst.card = to_roman(inst.card)
                    inst.save()
                    saved += 1

            structural_summary, _ = StructuralSummary.objects.get_or_create(client=client)
            try:
                structural_summary.calculate_values()
            except Exception:
                structural_summary.save()

            messages.success(request, f"{saved}건 저장되었습니다.")
            return redirect('scoring:client_detail', client_id=client.id)
        else:
            errors = []
            for i, f in enumerate(formset.forms, start=1):
                if f.errors:
                    for k, errs in f.errors.items():
                        errors.append(f"{i}행 {k}: {', '.join(errs)}")
            if errors:
                messages.error(request, "입력 오류: " + " / ".join(errors))
    else:
        formset = FormSet(queryset=qs)

    return render(request, 'update_response_codes.html', {
        'client': client,
        'formset': formset,
    })
