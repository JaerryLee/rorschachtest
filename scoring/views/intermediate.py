import logging
import json
import re
from io import BytesIO
from pathlib import Path

import pandas as pd
import numpy as np

from django.conf import settings
from django.contrib import messages
from django.forms import formset_factory, modelformset_factory
from django.http import (
    HttpResponse,
    HttpResponseForbidden,
    HttpResponseNotFound,
    JsonResponse,
)
from django.shortcuts import get_object_or_404, redirect, render
from django.db.models import IntegerField, Value
from django.db.models.functions import Cast, Coalesce

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

from ..filters import CardImagesFilter, PResponseFilter, SearchReferenceFilter
from ..forms import ClientForm, ResponseCodeForm
from ..models import (
    CardImages,
    Client,
    PopularResponse,
    ResponseCode,
    SearchReference,
    StructuralSummary,
)
from ._base import (
    group_min_required, GROUP_LEVEL, GROUP_LABEL,
    normalize_card_to_num, to_roman,
)

TOTAL_CAP = 100
DEFAULT_EXTRA = 40

HEADER_MAP_COMPACT = {
    'id': None,
    '카드': 'card', 'card': 'card',
    'n': 'response_num', '응답수': 'response_num', 'response_num': 'response_num',
    '시간': 'time', 'time': 'time',
    '반응': 'response', 'response': 'response',
    '질문': 'inquiry', 'inquiry': 'inquiry',
    '회전': 'rotation', 'v': 'rotation', 'rotation': 'rotation',
    '반응영역': 'location', '위치': 'location', 'location': 'location',
    '발달질': 'dev_qual', 'dq': 'dev_qual', 'devqual': 'dev_qual', 'dev_qual': 'dev_qual',
    '영역번호': 'loc_num', 'locnum': 'loc_num', 'loc_num': 'loc_num',
    '결정인': 'determinants', 'determinants': 'determinants',
    '형태질': 'form_qual', '형태 질': 'form_qual', 'formquality': 'form_qual', 'form_qual': 'form_qual',
    '(2)': 'pair', '2': 'pair', 'pair': 'pair',
    '내용인': 'content', '내용': 'content', 'content': 'content',
    'p': 'popular', 'popular': 'popular',
    'z': 'Z', 'Z': 'Z',
    '특수점수': 'special', 'special': 'special',
    '코멘트': 'comment', '메모': 'comment', 'comment': 'comment',
}

REQUIRED_FIELDS = [
    'card','response_num','time','response','inquiry','rotation','location',
    'dev_qual','loc_num','determinants','form_qual','pair','content',
    'popular','Z','special','comment'
]

_TOKEN_SEP = r"[,\s;+/]+"

def _compact(s: str) -> str:
    s = (s or '').strip()
    s = re.sub(r'[:：]\s*$', '', s)
    return s.replace(' ', '').lower()

def normalize_header(h):
    if h is None:
        return None
    return HEADER_MAP_COMPACT.get(_compact(str(h)), None)

def _normalize_text_value(v):
    if v is None:
        return ''
    if not isinstance(v, str):
        return v
    s = v.replace('\u00A0', ' ').replace('\u200b', '').replace('\u200c', '').strip()
    trans = {
        '“':'"', '”':'"', '‘':"'", '’':"'", '′':"'", '´':"'", '｀':"'",
        '·':'.', 'ㆍ':'.', '‧':'.', '•':'.',
        '：':':', '，':',', '．':'.', '／':'/', '－':'-',
    }
    return ''.join(trans.get(ch, ch) for ch in s)

def _fix_determinant_typos(s: str) -> tuple[str, list[str]]:
    if not s:
        return s, []
    notes = []
    fixed = re.sub(r"\bm['’`]?p\b", "mp", s, flags=re.IGNORECASE)
    if fixed != s:
        notes.append("결정인 m'p→mp")
    return fixed, notes

def _normalize_special_tokens(s: str) -> str:
    if not s:
        return s
    toks = re.split(_TOKEN_SEP, s.strip())
    toks = [t.strip().upper() for t in toks if t and t.strip()]
    return ", ".join(dict.fromkeys(toks))

def _detect_token(s: str, token: str) -> bool:
    if not s:
        return False
    return re.search(rf"(^|{_TOKEN_SEP}){re.escape(token)}($|{_TOKEN_SEP})", s, flags=re.IGNORECASE) is not None

def _apply_row_postprocess(data: dict) -> dict:
    data['card'] = to_roman(data.get('card', ''))
    for int_key in ('response_num', 'loc_num'):
        txt = str(data.get(int_key, '')).strip()
        if txt == '':
            data[int_key] = ''
        else:
            try:
                data[int_key] = int(float(txt))
            except Exception:
                pass
    for key in ('dev_qual','determinants','form_qual','pair','content','popular','Z','special','location','rotation','time'):
        if key in data:
            data[key] = _normalize_text_value(data[key])
    return data

def _pick_input_sheet(wb):
    for name in ("입력", "input", "Input", "INPUT"):
        if name in wb.sheetnames:
            return wb[name]
    return wb.active

def _make_formset_factory(extra: int = DEFAULT_EXTRA):
    safe_extra = max(0, min(extra, TOTAL_CAP))
    return formset_factory(ResponseCodeForm, extra=safe_extra, max_num=TOTAL_CAP)

def _make_dynamic_formset_for_initial(initial_len: int):
    remain = max(0, TOTAL_CAP - int(initial_len))
    desired = min(DEFAULT_EXTRA, remain)
    return _make_formset_factory(extra=desired)

RESOURCE_DIR = Path(getattr(
    settings,
    'SCORING_RESOURCE_DIR',
    Path(__file__).resolve().parent.parent / 'resources'
)).resolve()

STOPWORDS = set([
    '이','그','저','나','너','그것','이것','저것','들','\n','때','것','그리고','하지만','또는','즉','그렇지','그래서','그러므로',
    '대해','대하여','위해','때문에','그런데','근데','이런','저런','그런','같은','처럼','듯','도','만','또','조차','까지',
    '네','예','수검자','검사자','있다','보이다','Q','A','반응반복','같다','하다','반응','반복','부분','여기','이렇게','거','그렇다','어떻다','얘','보다'
])

@group_min_required('intermediate')
def download_response_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "입력"

    headers = [
        'ID','카드','N','시간','반응','질문','회전','반응영역','발달질','영역번호',
        '결정인','형태질','(2)','내용인','P','Z','특수점수','코멘트'
    ]
    ws.append(headers)
    ws.append([
        '', 'I', 1, "12s", "박쥐", "윗부분이 날개 같아요", '', 'W', '+', '',
        "M.F", '+', '', 'H', 'P', 'ZW', '', ''
    ])

    head_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
    bold = Font(bold=True)
    for col in range(1, len(headers)+1):
        c = ws.cell(row=1, column=col)
        c.font = bold
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"

    widths = [6, 6, 6, 8, 28, 28, 8, 10, 8, 10, 12, 10, 6, 10, 6, 6, 10, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    help_ws = wb.create_sheet("도움말")
    help_ws.append(["열 이름", "설명(업로드 시 매핑되는 내부 필드)"])
    mapping = [
        ("ID", "서버에서 사용하지 않음(무시됨)"),
        ("카드", "card (I~X / 1~10 모두 허용)"),
        ("N", "response_num(정수)"),
        ("시간", "time"),
        ("반응", "response"),
        ("질문", "inquiry"),
        ("회전", "rotation"),
        ("반응영역", "location"),
        ("발달질", "dev_qual"),
        ("영역번호", "loc_num(정수)"),
        ("결정인", "determinants"),
        ("형태질", "form_qual"),
        ("(2)", "pair"),
        ("내용인", "content"),
        ("P", "popular"),
        ("Z", "Z"),
        ("특수점수", "special"),
        ("코멘트", "comment"),
    ]
    for r in mapping:
        help_ws.append(list(r))
    for col in ("A","B"):
        help_ws.column_dimensions[col].width = 36
    for col in range(1,3):
        help_ws.cell(row=1, column=col).font = Font(bold=True)
        help_ws.cell(row=1, column=col).fill = head_fill
        help_ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    resp = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="response_template_ko.xlsx"'
    return resp

@group_min_required('intermediate')
def search(request, client_id=None):
    client_id = client_id or request.GET.get('client_id')
    if not client_id:
        return HttpResponseNotFound("client_id가 필요합니다.")
    client = get_object_or_404(Client, id=client_id)
    if client.tester != request.user:
        return HttpResponse("액세스 거부: 작성 권한이 없습니다.", status=403)

    ResponseCodeFormSet = _make_formset_factory(extra=DEFAULT_EXTRA)

    if request.method == 'POST' and request.POST.get('mode') == 'upload_preview':
        xfile = request.FILES.get('xlsx_file')
        if not xfile:
            messages.error(request, "엑셀 파일을 선택해 주세요.")
            formset = ResponseCodeFormSet()
        else:
            try:
                wb = load_workbook(filename=xfile, data_only=True)
                ws = _pick_input_sheet(wb)
            except Exception:
                messages.error(request, "엑셀 파일을 열 수 없습니다. (.xlsx 형식 확인)")
                formset = ResponseCodeFormSet()
            else:
                raw_headers = [(c.value or '') for c in ws[1]]
                mapped = [normalize_header(h) for h in raw_headers]
                index_by_field = {f: idx for idx, f in enumerate(mapped) if f}
                missing = [f for f in REQUIRED_FIELDS if f not in index_by_field]
                if missing:
                    messages.error(request, "필수 열 누락: " + ", ".join(missing) + "  (샘플 템플릿을 사용하세요)")
                    formset = ResponseCodeFormSet()
                else:
                    raw_rows = list(ws.iter_rows(min_row=2, values_only=True))
                    filtered_rows = []
                    for r in raw_rows:
                        vals = [(r[index_by_field[f]] if index_by_field[f] < len(r) else None) for f in REQUIRED_FIELDS]
                        if all(v is None or str(v).strip() == "" for v in vals):
                            continue
                        filtered_rows.append(r)

                    initial, fix_notes, error_notes = [], [], []
                    for row_idx, row in enumerate(filtered_rows, start=2):
                        data = {}
                        for f in REQUIRED_FIELDS:
                            idx = index_by_field[f]
                            v = row[idx] if idx < len(row) else ''
                            v = '' if v is None else v
                            data[f] = _normalize_text_value(v)
                        data = _apply_row_postprocess(data)

                        det_before = data.get('determinants', '')
                        det_after, notes = _fix_determinant_typos(det_before)
                        if det_after != det_before:
                            data['determinants'] = det_after
                        if data.get('special'):
                            data['special'] = _normalize_special_tokens(data['special'])
                        if notes:
                            fix_notes.append(f"{row_idx}행: " + ", ".join(notes))

                        form_probe = ResponseCodeForm(data)
                        if not form_probe.is_valid():
                            errs = "; ".join([f"{fld}: {', '.join(errs)}" for fld, errs in form_probe.errors.items()])
                            error_notes.append(f"{row_idx}행 → {errs}")

                        initial.append(data)

                    if fix_notes:
                        shown = " / ".join(fix_notes[:10])
                        more = f"  …외 {len(fix_notes)-10}건" if len(fix_notes) > 10 else ""
                        messages.info(request, f"자동수정 적용: {shown}{more}")
                    if error_notes:
                        shown = " / ".join(error_notes[:10])
                        more = f"  …외 {len(error_notes)-10}건" if len(error_notes) > 10 else ""
                        messages.warning(request, f"유효성 확인 필요: {shown}{more}")
                    if len(initial) > TOTAL_CAP:
                        initial = initial[:TOTAL_CAP]
                        messages.warning(request, f"총 {TOTAL_CAP}행까지만 불러옵니다.")

                    PreviewFormSet = _make_dynamic_formset_for_initial(len(initial))
                    formset = PreviewFormSet(initial=initial)
                    messages.success(request, f"엑셀에서 {len(initial)}건을 불러왔습니다. 확인 후 저장하세요.")

        reference_list = SearchReference.objects.all()
        card_image_list = CardImages.objects.all()
        Presponse_list = PopularResponse.objects.all()
        search_filter = SearchReferenceFilter(request.GET, queryset=reference_list)
        card_image_filter = CardImagesFilter(request.GET, queryset=card_image_list)
        p_response_filter = PResponseFilter(request.GET, queryset=Presponse_list)

        return render(
            request,
            'intermediate.html',
            {
                'client': client,
                'formset': formset,
                'filter': search_filter,
                'image_filter': card_image_filter,
                'p_response_filter': p_response_filter,
            },
        )

    if request.method == 'POST':
        BoundFormSet = _make_formset_factory(extra=0)
        if request.POST.get('additems') == 'true':
            formset_dictionary_copy = request.POST.copy()
            total_now = int(formset_dictionary_copy.get('form-TOTAL_FORMS', 0) or 0)
            formset_dictionary_copy['form-TOTAL_FORMS'] = str(min(TOTAL_CAP, total_now + 1))
            formset = BoundFormSet(formset_dictionary_copy)
        else:
            formset = BoundFormSet(request.POST)
            if formset.is_valid():
                warnings, created, fix_count = [], 0, 0
                for idx, form in enumerate(formset.forms, start=1):
                    cd = getattr(form, 'cleaned_data', {}) or {}
                    if not (cd.get('card') and cd.get('response')):
                        continue
                    inst = form.instance
                    inst.card = to_roman(cd.get('card', ''))

                    det_before = (inst.determinants or '')
                    inst.determinants, notes = _fix_determinant_typos(det_before)
                    if notes:
                        fix_count += 1
                    inst.special = _normalize_special_tokens(inst.special or '')

                    card = normalize_card_to_num(inst.card)
                    loc  = (cd.get('location') or '')
                    dq   = (cd.get('dev_qual') or '')
                    zval = str(cd.get('Z') or '').strip().upper()
                    if card == '1' and ('W' in loc) and ('+' in dq) and (zval == 'ZW'):
                        warnings.append(idx)

                    inst.client = client
                    inst.save()
                    created += 1

                if warnings:
                    messages.warning(
                        request,
                        f"Z점수 재검토 권고 {len(warnings)}건(행: {', '.join(map(str, warnings))}). "
                        "Z가 결과(예: Zf/Zsum/Zd)에 영향을 줄 수 있으니 검토하세요."
                    )
                if fix_count:
                    messages.info(request, f"자동 보정 적용 {fix_count}행 (m'p→mp, 특수점수 정규화 등)")
                messages.success(request, f"{created}건 저장되었습니다.")
                return redirect('scoring:client_detail', client_id=client.id)
            else:
                details = []
                for i, f in enumerate(formset.forms, start=1):
                    if f.errors:
                        for k, errs in f.errors.items():
                            details.append(f"{i}행 {k}: {', '.join(errs)}")
                if details:
                    messages.error(request, "입력 오류: " + " / ".join(details))
    else:
        ResponseCodeFormSet = _make_formset_factory(extra=DEFAULT_EXTRA)
        formset = ResponseCodeFormSet()

    reference_list = SearchReference.objects.all()
    card_image_list = CardImages.objects.all()
    Presponse_list = PopularResponse.objects.all()
    search_filter = SearchReferenceFilter(request.GET, queryset=reference_list)
    card_image_filter = CardImagesFilter(request.GET, queryset=card_image_list)
    p_response_filter = PResponseFilter(request.GET, queryset=Presponse_list)

    if request.GET.get('Card'):
        card_number = request.GET['Card']
        card_image_filter = CardImagesFilter(request.GET, queryset=CardImages.objects.filter(card_number=card_number))
        p_response_filter = PResponseFilter(request.GET, queryset=PopularResponse.objects.filter(card_number=card_number))

    return render(
        request,
        'intermediate.html',
        {
            'client': client,
            'formset': formset,
            'filter': search_filter,
            'image_filter': card_image_filter,
            'p_response_filter': p_response_filter,
        },
    )

@group_min_required('intermediate')
def update_response_codes(request, client_id):
    client = get_object_or_404(Client, id=client_id)
    if client.tester != request.user:
        return HttpResponse("액세스 거부: 작성 권한이 없습니다.", status=403)

    response_codes = (
        ResponseCode.objects
        .filter(client=client)
        .annotate(n_int=Cast(Coalesce('response_num', Value(0)), IntegerField()))
        .order_by('n_int', 'card', 'id')
    )

    current = response_codes.count()
    extra = max(0, min(DEFAULT_EXTRA, TOTAL_CAP - current))
    ResponseCodeFormSet = modelformset_factory(
        ResponseCode, form=ResponseCodeForm, extra=extra, max_num=TOTAL_CAP
    )

    if request.method == 'POST':
        if request.POST.get('additems') == 'true':
            formset_dictionary_copy = request.POST.copy()
            total_now = int(formset_dictionary_copy.get('form-TOTAL_FORMS', 0) or 0)
            formset_dictionary_copy['form-TOTAL_FORMS'] = str(min(TOTAL_CAP, total_now + 1))
            formset = ResponseCodeFormSet(formset_dictionary_copy, queryset=response_codes)
        else:
            formset = ResponseCodeFormSet(request.POST, queryset=response_codes)
            if formset.is_valid():
                fix_count, saved = 0, 0
                for form in formset:
                    if form.cleaned_data.get('card') and form.cleaned_data.get('response'):
                        form.instance.client = client
                        form.instance.card = to_roman(form.cleaned_data.get('card', ''))
                        det_before = (form.instance.determinants or '')
                        form.instance.determinants, notes = _fix_determinant_typos(det_before)
                        if notes:
                            fix_count += 1
                        form.instance.special = _normalize_special_tokens(form.instance.special or '')
                        form.save()
                        saved += 1

                structural_summary, _ = StructuralSummary.objects.get_or_create(client=client)
                try:
                    structural_summary.calculate_values()
                except Exception:
                    structural_summary.save()

                if fix_count:
                    messages.info(request, f"자동 보정 적용 {fix_count}행 (m'p→mp, 특수점수 정규화 등)")
                messages.success(request, f"{saved}건 저장되었습니다.")
                return redirect('scoring:client_list')
            else:
                details = []
                for i, f in enumerate(formset.forms, start=1):
                    if f.errors:
                        for k, errs in f.errors.items():
                            details.append(f"{i}행 {k}: {', '.join(errs)}")
                if details:
                    shown = " / ".join(details[:10])
                    more = f" …외 {len(details)-10}건" if len(details) > 10 else ""
                    messages.error(request, "입력 오류: " + shown + more)
    else:
        formset = ResponseCodeFormSet(queryset=response_codes)

    reference_list = SearchReference.objects.all()
    card_image_list = CardImages.objects.all()
    Presponse_list = PopularResponse.objects.all()
    search_filter = SearchReferenceFilter(request.GET, queryset=reference_list)
    card_image_filter = CardImagesFilter(request.GET, queryset=card_image_list)
    p_response_filter = PResponseFilter(request.GET, queryset=Presponse_list)

    if request.GET.get('Card'):
        card_number = request.GET['Card']
        card_image_filter = CardImagesFilter(request.GET, queryset=CardImages.objects.filter(card_number=card_number))
        p_response_filter = PResponseFilter(request.GET, queryset=PopularResponse.objects.filter(card_number=card_number))

    return render(
        request,
        'update_response_codes.html',
        {
            'client': client,
            'formset': formset,
            'filter': search_filter,
            'image_filter': card_image_filter,
            'p_response_filter': p_response_filter,
        },
    )

def search_results(request):
    reference_list = SearchReference.objects.all()
    card_image_list = CardImages.objects.all()
    Presponse_list = PopularResponse.objects.all()

    search_filter = SearchReferenceFilter(request.GET, queryset=reference_list)
    card_image_filter = CardImagesFilter(request.GET, queryset=card_image_list)
    p_response_filter = PResponseFilter(request.GET, queryset=Presponse_list)

    if request.GET.get('Card'):
        card_number = request.GET['Card']
        card_image_filter = CardImagesFilter(request.GET, queryset=CardImages.objects.filter(card_number=card_number))
        p_response_filter = PResponseFilter(request.GET, queryset=PopularResponse.objects.filter(card_number=card_number))

    return render(
        request,
        'search_results.html',
        {'filter': search_filter, 'image_filter': card_image_filter, 'p_response_filter': p_response_filter},
    )

@group_min_required('intermediate')
def add_client(request):
    next_target = (request.POST.get('next') or request.GET.get('next') or 'intermediate').strip().lower()
    if next_target not in ('advanced', 'intermediate'):
        next_target = 'intermediate'

    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            client_instance = form.save(commit=False)
            client_instance.tester = request.user
            client_instance.save()

            if next_target == 'advanced':
                return redirect('scoring:advanced_upload', client_id=client_instance.id)
            return redirect('scoring:search', client_id=client_instance.id)
    else:
        initial_examiner = request.user.get_full_name() or request.user.username
        form = ClientForm(initial={
            'name': request.GET.get('name', ''),
            'gender': request.GET.get('gender', ''),
            'birthdate': request.GET.get('birthdate', ''),
            'testDate': request.GET.get('testDate', ''),
            'evaluation_purpose': request.GET.get('evaluation_purpose', ''),
            'rorschach_history': request.GET.get('rorschach_history', '0'),
            'current_psych_treatment': request.GET.get('current_psych_treatment', 'none'),
            'past_psych_treatment': request.GET.get('past_psych_treatment', 'none'),
            'notes': request.GET.get('notes', ''),
            'consent': request.GET.get('consent', ''),
        })

    return render(request, 'add_client.html', {
        'form': form,
        'next': request.GET.get('next', 'intermediate'),
    })


@group_min_required('intermediate')
def client_list(request):
    user = request.user
    clients = Client.objects.filter(tester=user)
    return render(request, 'client_list.html', {'clients': clients})


@group_min_required('intermediate')
def client_detail(request, client_id):
    user = request.user
    client_obj = get_object_or_404(Client, id=client_id, tester=user)
    response_codes = (
        ResponseCode.objects
        .filter(client=client_obj)
        .annotate(n_int=Cast(Coalesce('response_num', Value(0)), IntegerField()))
        .order_by('n_int', 'card', 'id')
    )
    return render(request, 'client_detail.html', {'client': client_obj, 'response_codes': response_codes})

@group_min_required('intermediate')
def export_structural_summary_xlsx(request, client_id):
    try:
        client = Client.objects.get(id=client_id)
        if client.tester != request.user:
            return HttpResponse("액세스 거부: 해당 정보를 볼 수 있는 권한이 없습니다.", status=403)

        response_codes = ResponseCode.objects.filter(client_id=client_id)

        numbers_found = {normalize_card_to_num(rc.card) for rc in response_codes if rc.card}
        required = {str(i) for i in range(1, 11)}
        missing = sorted(required - numbers_found, key=int)
        if missing:
            missing_roman = [to_roman(n) for n in missing]
            return HttpResponse("다음 카드의 반응이 없습니다: " + ", ".join(missing_roman))

        structural_summary, _ = StructuralSummary.objects.get_or_create(client_id=client_id)
        try:
            structural_summary.calculate_values()
        except Exception:
            structural_summary.save()
    except Client.DoesNotExist:
        logging.error("해당 ID의 클라이언트를 찾을 수 없음")
        return HttpResponseNotFound("클라이언트 정보를 찾을 수 없습니다.")
    except Exception as e:
        logging.error(f"예기치 못한 오류 발생: {e}")
        error_message = f"예기치 못한 오류 발생: {type(e).__name__}, {str(e)}"
        return JsonResponse({'error': error_message}, status=500)

    PASTEL_FILL = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
    LINE_COLOR  = "FFB7B7B7"
    THIN_EDGE   = Side(border_style='thin', color=LINE_COLOR)
    HDR_FONT    = Font(bold=True)

    def box_border(ws, cell_range, line_style="thin", color="FF000000"):
        rows = list(ws[cell_range])
        if not rows:
            return

        edge = Side(style=line_style, color=color)
        max_y = len(rows) - 1

        for y, row in enumerate(rows):
            max_x = len(row) - 1
            for x, cell in enumerate(row):
                b = cell.border
                cell.border = Border(
                    left=edge if x == 0 else b.left,
                    right=edge if x == max_x else b.right,
                    top=edge if y == 0 else b.top,
                    bottom=edge if y == max_y else b.bottom,
                )

    def header_cell(ws, addr: str, value: str):
        ws[addr] = value
        ws[addr].fill = PASTEL_FILL
        ws[addr].font = HDR_FONT
        ws[addr].alignment = Alignment(horizontal='center', vertical='center')

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    ws  = wb.create_sheet(title='상단부')
    wsd = wb.create_sheet(title='하단부')
    wsi = wb.create_sheet(title='특수지표')

    ws.sheet_view.showGridLines = False

    ws.merge_cells('A4:B4'); ws.merge_cells('A14:B14'); ws.merge_cells('A20:D20')
    ws.merge_cells('F4:H4'); ws.merge_cells('G5:H5'); ws.merge_cells('J4:K4')
    ws.merge_cells('M4:N4'); ws.merge_cells('M16:P16')

    header_cell(ws, 'A4',  'Location Features')
    header_cell(ws, 'A14', 'Developmental Quality')
    header_cell(ws, 'A20', 'Form Quality')
    header_cell(ws, 'F4',  'Determinants')
    header_cell(ws, 'F5',  'Blends')
    header_cell(ws, 'G5',  'Single')
    header_cell(ws, 'J4',  'Contents')
    header_cell(ws, 'M4',  'approach')
    header_cell(ws, 'M16', 'Special Scores')

    for pos in [
        'A4:B4', 'A5:B12', 'A14:B14', 'A15:B18', 'A20:D20', 'A21:D26',
        'F4:H4', 'F5:F5', 'G5:H5', 'F6:F29', 'G6:H29', 'J4:K4', 'J5:K31',
        'M4:N4', 'M5:N14', 'M16:P16', 'M17:P30'
    ]:
        box_border(ws, pos)

    ws['A5']  = 'Zf'; ws['A6'] = 'Zsum'; ws['A7'] = 'Zest'
    ws['A9']  = 'W';  ws['A10'] = 'D';    ws['A11'] = 'Dd'; ws['A12'] = 'S'

    ws['A15'] = '+'; ws['A16'] = 'o'; ws['A17'] = 'v/+'; ws['A18'] = 'v'

    ws['B21'] = 'FQx'; ws['C21'] = 'MQual'; ws['D21'] = 'W+D'
    ws['A22'] = '+'; ws['A23'] = 'o'; ws['A24'] = 'u'; ws['A25'] = '-'; ws['A26'] = 'none'

    fields = ['M','FM','m',"FC","CF","C","Cn","FC'","C'F","C'",
              'FT','TF','T','FV','VF','V','FY','YF','Y','Fr','rF','FD','F','(2)']
    real_field = ['M','FM','m_l','FC','CF','C','Cn','FCa','CaF','Ca',
                  'FT','TF','T','FV','VF','V','FY','YF','Y','Fr','rF','FD','F','pair']
    s_row = 6
    for name in fields:
        ws.cell(row=s_row, column=7, value=name); s_row += 1

    ws['J5']  = 'H'; ws['J6'] = '(H)'; ws['J7'] = 'Hd'; ws['J8'] = '(Hd)'; ws['J9'] = 'Hx'
    ws['J10'] = 'A'; ws['J11'] = '(A)'; ws['J12'] = 'Ad'; ws['J13'] = '(Ad)'; ws['J14'] = 'An'
    ws['J15'] = 'Art'; ws['J16'] = 'Ay'; ws['J17'] = 'Bl'; ws['J18'] = 'Bt'; ws['J19'] = 'Cg'
    ws['J20'] = 'Cl'; ws['J21'] = 'Ex'; ws['J22'] = 'Fd'; ws['J23'] = 'Fi'; ws['J24'] = 'Ge'
    ws['J25'] = 'Hh'; ws['J26'] = 'Ls'; ws['J27'] = 'Na'; ws['J28'] = 'Sc'; ws['J29'] = 'Sx'
    ws['J30'] = 'Xy'; ws['J31'] = 'Id'

    ws['M5']  = 'I'; ws['M6'] = 'II'; ws['M7'] = 'III'; ws['M8'] = 'IV'; ws['M9']  = 'V'
    ws['M10'] = 'VI'; ws['M11'] = 'VII'; ws['M12'] = 'VIII'; ws['M13'] = 'IX'; ws['M14'] = 'X'

    ws['N17'] = 'Lvl-1'; ws['O17'] = 'Lvl-2'
    for r, name in enumerate(['DV','INC','DR','FAB','ALOG','CON'], start=18):
        ws.cell(row=r, column=13, value=name)
    ws['M24'] = 'Raw Sum6'; ws['M25'] = 'Weighted Sum6'
    ws['M26'] = 'AB'; ws['M27'] = 'AG'; ws['M28'] = 'COP'; ws['M29'] = 'CP'
    ws['O26'] = 'GHR'; ws['O27'] = 'PHR'; ws['O28'] = 'MOR'; ws['O29'] = 'PER'; ws['O30'] = 'PSV'

    ws['B5']  = structural_summary.Zf
    ws['B6']  = structural_summary.Zsum; ws['B6'].number_format = '0.0'
    ws['B7']  = structural_summary.Zest; ws['B7'].number_format = '0.0'
    ws['B9']  = structural_summary.W; ws['B10'] = structural_summary.D
    ws['B11'] = structural_summary.Dd; ws['B12'] = structural_summary.S

    ws['B15'] = structural_summary.dev_plus
    ws['B16'] = structural_summary.dev_o
    ws['B17'] = structural_summary.dev_vplus
    ws['B18'] = structural_summary.dev_v

    ws['B22'] = structural_summary.fqx_plus
    ws['B23'] = structural_summary.fqx_o
    ws['B24'] = structural_summary.fqx_u
    ws['B25'] = structural_summary.fqx_minus
    ws['B26'] = structural_summary.fqx_none

    ws['C22'] = structural_summary.mq_plus
    ws['C23'] = structural_summary.mq_o
    ws['C24'] = structural_summary.mq_u
    ws['C25'] = structural_summary.mq_minus
    ws['C26'] = structural_summary.mq_none

    ws['D22'] = structural_summary.wd_plus
    ws['D23'] = structural_summary.wd_o
    ws['D24'] = structural_summary.wd_u
    ws['D25'] = structural_summary.wd_minus
    ws['D26'] = structural_summary.wd_none

    blends = structural_summary.blends.split(',') if structural_summary.blends else []
    start_row = 6
    for blend in blends:
        ws.cell(row=start_row, column=6, value=blend); start_row += 1

    row = 6
    for field_name in real_field:
        ws.cell(row=row, column=8, value=getattr(structural_summary, field_name)); row += 1

    cont_real_fields = ['H','H_paren','Hd','Hd_paren','Hx','A','A_paren','Ad','Ad_paren','An',
                        'Art','Ay','Bl','Bt','Cg','Cl','Ex','Fd_l','Fi','Ge','Hh','Ls',
                        'Na','Sc','Sx','Xy','Idio']
    row = 5
    for field_name in cont_real_fields:
        ws.cell(row=row, column=11, value=getattr(structural_summary, field_name)); row += 1

    ws['N5']  = structural_summary.app_I
    ws['N6']  = structural_summary.app_II
    ws['N7']  = structural_summary.app_III
    ws['N8']  = structural_summary.app_IV
    ws['N9']  = structural_summary.app_V
    ws['N10'] = structural_summary.app_VI
    ws['N11'] = structural_summary.app_VII
    ws['N12'] = structural_summary.app_VIII
    ws['N13'] = structural_summary.app_IX
    ws['N14'] = structural_summary.app_X

    lv1 = ['sp_dv','sp_inc','sp_dr','sp_fab','sp_alog','sp_con']
    for i, fname in enumerate(lv1, start=18):
        ws.cell(row=i, column=14, value=getattr(structural_summary, fname))
    lv2 = ['sp_dv2','sp_inc2','sp_dr2','sp_fab2']
    for i, fname in enumerate(lv2, start=18):
        ws.cell(row=i, column=15, value=getattr(structural_summary, fname))
    ws['N24'] = structural_summary.sum6
    ws['N25'] = structural_summary.wsum6
    ws['N26'] = structural_summary.sp_ab
    ws['N27'] = structural_summary.sp_ag
    ws['N28'] = structural_summary.sp_cop
    ws['N29'] = structural_summary.sp_cp
    ws['P26'] = structural_summary.sp_ghr
    ws['P27'] = structural_summary.sp_phr
    ws['P28'] = structural_summary.sp_mor
    ws['P29'] = structural_summary.sp_per
    ws['P30'] = structural_summary.sp_psv

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) * 1.1 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length

    wsd.sheet_view.showGridLines = False
    wsd.merge_cells('A3:F3'); wsd.merge_cells('H3:I3'); wsd.merge_cells('K3:N3')
    wsd.merge_cells('A14:D14'); wsd.merge_cells('F14:G14'); wsd.merge_cells('I14:J14'); wsd.merge_cells('L14:M14')

    header_cell(wsd, 'A3',  'Core')
    header_cell(wsd, 'H3',  'Affect')
    header_cell(wsd, 'K3',  'Interpersonal')
    header_cell(wsd, 'A14', 'Ideation')
    header_cell(wsd, 'F14', 'Mediation')
    header_cell(wsd, 'I14', 'Processing')
    header_cell(wsd, 'L14', 'Self')

    for pos in ['A3:F3','A4:F4','A5:F7','A8:F9','H3:I3','H4:I10','K3:N3','K4:N12',
                'A14:D14','A15:D19','F14:G14','F15:G21','I14:J14','I15:J21','L14:M14','L15:M21']:
        box_border(wsd, pos)

    wsd['A4'] = 'R'; wsd['C4'] = 'L'
    wsd['A5'] = 'EB'; wsd['A6'] = 'eb'
    wsd['C5'] = 'EA'; wsd['C6'] = 'es'; wsd['C7'] = 'Adj es'
    wsd['E5'] = 'EBper'; wsd['E6'] = 'D'; wsd['E7'] = 'Adj D'
    wsd['A8'] = 'FM'; wsd['A9'] = 'm'
    wsd["C8"] = "SumC'"; wsd['C9'] = 'SumV'
    wsd['E8'] = 'SumT'; wsd['E9'] = 'SumY'

    wsd['H4'] = 'FC:CF+C'; wsd['H5'] = 'Pure C'; wsd["H6"] = "SumC':WsumC"
    wsd['H7'] = 'Afr'; wsd['H8'] = 'S'; wsd['H9'] = 'Blends:R'; wsd['H10'] = 'CP'

    wsd['K4'] = 'COP'; wsd['M4'] = 'AG'
    wsd['K5'] = 'GHR:PHR'; wsd['K6'] = 'a:p'; wsd['K7'] = 'Food'
    wsd['K8'] = 'SumT'; wsd['K9'] = 'Human Content'; wsd['K10'] = 'Pure H'
    wsd['K11'] = 'PER'; wsd['K12'] = 'Isolation Index'

    wsd['A15'] = 'a:p'; wsd['A16'] = 'Ma:Mp'; wsd['A17'] = 'Intel(2AB+Art+Ay)'; wsd['A18'] = 'MOR'
    wsd['C15'] = 'Sum6'; wsd['C16'] = 'Lvl-2'; wsd['C17'] = 'Wsum6'; wsd['C18'] = 'M-'; wsd['C19'] = 'M none'

    med_fields = ['XA%','WDA%','X-%','S-','P','X+%','Xu%']
    med_real   = ['xa_per','wda_per','x_minus_per','s_minus','popular','x_plus_per','xu_per']
    for i, f in enumerate(med_fields, start=15):
        wsd.cell(row=i, column=6, value=f)

    pro_fields = ['Zf','W:D:Dd','W:M','Zd','PSV','DQ+','DQv']
    pro_real   = ['Zf','W_D_Dd','W_M','Zd','sp_psv','dev_plus','dev_v']
    for i, f in enumerate(pro_fields, start=15):
        wsd.cell(row=i, column=9, value=f)

    self_fields = ['Ego[3r+(2)/R]','Fr+rF','SumV','FD','An+Xy','MOR','H:(H)+Hd+(Hd)']
    self_real   = ['ego','fr_rf','sum_V','fdn','an_xy','sp_mor','h_prop']
    for i, f in enumerate(self_fields, start=15):
        wsd.cell(row=i, column=12, value=f)

    wsd['B4']  = structural_summary.R; wsd['D4']  = structural_summary.L
    wsd['B5']  = structural_summary.ErleBnistypus
    wsd['B6']  = structural_summary.eb
    wsd['D5']  = structural_summary.EA
    wsd['D6']  = structural_summary.es
    wsd['D7']  = structural_summary.adj_es
    wsd['F5']  = 'NA' if structural_summary.EBper == 0 else structural_summary.EBper
    wsd['F6']  = structural_summary.D_score
    wsd['F7']  = structural_summary.adj_D
    wsd['B8']  = structural_summary.sum_FM
    wsd['B9']  = structural_summary.sum_m
    wsd['D8']  = structural_summary.sum_Ca
    wsd['D9']  = structural_summary.sum_V
    wsd['F8']  = structural_summary.sum_T
    wsd['F9']  = structural_summary.sum_Y

    wsd['I4'] = structural_summary.f_c_prop; wsd['I4'].alignment = Alignment(horizontal='right')
    wsd['I5'] = structural_summary.pure_c
    wsd['I6'] = structural_summary.ca_c_prop; wsd['I6'].alignment = Alignment(horizontal='right')
    wsd['I7'] = structural_summary.afr; wsd['I7'].number_format = "0.##;-0.##;0"
    wsd['I8'] = structural_summary.S
    wsd['I9'] = structural_summary.blends_r; wsd['I9'].alignment = Alignment(horizontal='right')
    wsd['I10'] = structural_summary.sp_cp

    wsd['L4']  = structural_summary.sp_cop
    wsd['N4']  = structural_summary.sp_ag
    wsd['M5']  = structural_summary.GHR_PHR; wsd['M5'].alignment = Alignment(horizontal='right')
    wsd['M6']  = structural_summary.a_p;  wsd['M6'].alignment = Alignment(horizontal='right')
    wsd['M7']  = structural_summary.Fd_l
    wsd['M8']  = structural_summary.sum_T
    wsd['M9']  = structural_summary.human_cont
    wsd['M10'] = structural_summary.H
    wsd['M11'] = structural_summary.sp_per
    wsd['M12'] = structural_summary.Isol

    wsd['B15'] = structural_summary.a_p; wsd['B15'].alignment = Alignment(horizontal='right')
    wsd['B16'] = structural_summary.Ma_Mp; wsd['B16'].alignment = Alignment(horizontal='right')
    wsd['B17'] = structural_summary.intel
    wsd['B18'] = structural_summary.sp_mor
    wsd['D15'] = structural_summary.sum6
    wsd['D16'] = structural_summary.Lvl_2
    wsd['D17'] = structural_summary.wsum6
    wsd['D18'] = structural_summary.mq_minus
    wsd['D19'] = structural_summary.mq_none

    for i, fname in enumerate(med_real, start=15):
        val = getattr(structural_summary, fname)
        wsd.cell(row=i, column=7, value=val)
        wsd.cell(row=i, column=7).number_format = "0" if isinstance(val, int) else "0.00"

    for i, fname in enumerate(pro_real, start=15):
        wsd.cell(row=i, column=10, value=getattr(structural_summary, fname))
        wsd.cell(row=i, column=10).alignment = Alignment(horizontal='right')

    for i, fname in enumerate(self_real, start=15):
        wsd.cell(row=i, column=13, value=getattr(structural_summary, fname))

    summary_row = 22
    def cb(text, positive):
        return f"☑ {text}" if positive else text

    pti_pos  = (structural_summary.sumPTI >= 3)
    depi_pos = (structural_summary.sumDEPI >= 5)
    cdi_pos  = (structural_summary.sumCDI  >= 4)
    scon_pos = (structural_summary.sumSCON >= 8)
    hvi_pos  = (structural_summary.sumHVI >= 4) and bool(structural_summary.HVI_premise)

    obs_pos  = bool(structural_summary.OBS_posi)

    obs_score = sum(1 for ch in (structural_summary.OBS or '') if ch == 'o')

    wsd.cell(row=summary_row, column=1,  value=cb(f"PTI={structural_summary.sumPTI}", pti_pos))
    wsd.cell(row=summary_row, column=5,  value=cb(f"HVI={structural_summary.sumHVI}", hvi_pos))
    wsd.cell(row=summary_row, column=6,  value=cb(f"DEPI={structural_summary.sumDEPI}", depi_pos))
    wsd.cell(row=summary_row, column=9,  value=cb(f"OBS={obs_score}", obs_pos))
    wsd.cell(row=summary_row, column=12, value=cb(f"CDI={structural_summary.sumCDI}", cdi_pos))
    wsd.cell(row=summary_row, column=15, value=cb(f"S-CON={structural_summary.sumSCON}", scon_pos))

    for col in (1,5,6,9,12,15):
        c = wsd.cell(row=summary_row, column=col)
        c.alignment = Alignment(horizontal='center')
        c.font = Font(bold=True)

    header_cell(wsi, 'A1',  "PTI")
    header_cell(wsi, 'A9',  "DEPI")
    header_cell(wsi, 'A20', "CDI")
    header_cell(wsi, 'D1',  "S-CON")
    header_cell(wsi, 'D17', "HVI")
    header_cell(wsi, 'A29', "OBS")

    # PTI
    wsi['A2'] = "XA%<.70 AND WDA%<.75"
    wsi['A3'] = "X-%>0.29"
    wsi['A4'] = "LVL2>2 AND FAB2>0"
    wsi['A5'] = "R<17 AND Wsum6>12 OR R>16 AND Wsum6>17*"
    wsi['A6'] = "M- > 1 OR X-% > 0.40"
    wsi['A7'] = "TOTAL"
    row = 2
    for i in range(0, 5):
        value = "✔" if structural_summary.PTI[i] == "o" else ''
        wsi.cell(row=row, column=2, value=value); row += 1
    wsi['B7'] = structural_summary.sumPTI

    # DEPI
    wsi['A9']  = "DEPI"
    wsi['A10'] = "SumV>0 OR FD>2"
    wsi['A11'] = "Col-shd blends>0 OR S>2"
    wsi['A12'] = "ego sup AND Fr+rF=0 OR ego inf"
    wsi['A13'] = "Afr<0.46 OR Blends<4"
    wsi['A14'] = "SumShd>FM+m OR SumC'>2"
    wsi['A15'] = "MOR>2 OR INTELL>3"
    wsi['A16'] = "COP<2 OR ISOL>0.24"
    wsi['A17'] = "TOTAL"
    wsi['A18'] = "POSITIVE?"
    row = 10
    for i in range(0, 7):
        value = "✔" if structural_summary.DEPI[i] == "o" else ''
        wsi.cell(row=row, column=2, value=value); row += 1
    wsi['B17'] = structural_summary.sumDEPI
    wsi['B18'] = structural_summary.sumDEPI >= 5

    # CDI
    wsi['A20'] = "CDI"
    wsi['A21'] = "EA<6 OR Daj<0"
    wsi['A22'] = "COP<2 AND AG<2"
    wsi['A23'] = "WSumC<2.5 OR Afr<0.46"
    wsi['A24'] = "p > a+1 OR pure H<2"
    wsi['A25'] = "SumT>1 OR ISOL>0.24 OR Fd>0"
    wsi['A26'] = "TOTAL"
    wsi['A27'] = "POSITIVE?"
    row = 21
    for i in range(0, 5):
        value = "✔" if structural_summary.CDI[i] == "o" else ''
        wsi.cell(row=row, column=2, value=value); row += 1
    wsi['B26'] = structural_summary.sumCDI
    wsi['B27'] = structural_summary.sumCDI >= 4

    # S-CON
    wsi['D2']  = "SumV+FD>2"
    wsi['D3']  = "col-shd blends>0"
    wsi['D4']  = "ego <0.31 ou >0.44"
    wsi['D5']  = "mor>3"
    wsi['D6']  = "Zd>3.5 ou <-3.5"
    wsi['D7']  = "es>EA"
    wsi['D8']  = "CF+C>FC"
    wsi['D9']  = "X+%<0.70"
    wsi['D10'] = "S>3"
    wsi['D11'] = "P<3 OU P>8"
    wsi['D12'] = "PURE H<2"
    wsi['D13'] = "R<17"
    wsi['D14'] = "TOTAL"
    wsi['D15'] = "POSITIVE?"
    row = 2
    for i in range(0, 12):
        value = "✔" if structural_summary.SCON[i] == "o" else ''
        wsi.cell(row=row, column=5, value=value); row += 1
    wsi['E14'] = structural_summary.sumSCON
    wsi['E15'] = structural_summary.sumSCON >= 8

    # HVI
    wsi['D18'] = 'SumT = 0'
    wsi['D19'] = "Zf>12"
    wsi['D20'] = "Zd>3.5"
    wsi['D21'] = "S>3"
    wsi['D22'] = "H+(H)+Hd+(Hd)>6"
    wsi['D23'] = "(H)+(A)+(Hd)+(Ad)>3"
    wsi['D24'] = "H+A : 4:1"
    wsi['D25'] = "Cg>3"
    wsi['D26'] = 'TOTAL'
    wsi['D27'] = 'POSITIVE?'
    wsi['E18'] = structural_summary.HVI_premise
    row = 19
    for i in range(0, 7):
        value = "✔" if structural_summary.HVI[i] == "o" else ''
        wsi.cell(row=row, column=5, value=value); row += 1
    wsi['E26'] = structural_summary.sumHVI
    wsi['E27'] = (structural_summary.sumHVI >= 4) and bool(structural_summary.HVI_premise)

    # OBS
    wsi['A30'] = 1; wsi['A31'] = 2; wsi['A32'] = 3; wsi['A33'] = 4; wsi['A34'] = 5
    wsi['B30'] = "Dd>3"; wsi['B31'] = "Zf>12"; wsi['B32'] = "Zd>3.0"; wsi['B33'] = "P>7"; wsi['B34'] = "FQ+>1"
    wsi['D30'] = "1-5 are true"
    wsi['D31'] = "FQ+>3 AND 2 items 1-4"
    wsi['D32'] = "X+%>0,89 et 3 items"
    wsi['D33'] = "FQ+>3 et X+%>0,89"
    wsi['D34'] = 'POSITIVE?'
    row = 30
    for i in range(0, 5):
        value = "✔" if structural_summary.OBS[i] == "o" else ''
        wsi.cell(row=row, column=3, value=value); row += 1
    row = 30
    for i in range(5, 9):
        value = "✔" if structural_summary.OBS[i] == "o" else ''
        wsi.cell(row=row, column=5, value=value); row += 1

    obs_score = sum(1 for ch in (structural_summary.OBS or '') if ch == 'o')
    wsi['E29'] = "TOTAL"; wsi['F29'] = obs_score
    wsi['E34'] = structural_summary.OBS_posi

    for column_cells in wsi.columns:
        length = max(len(str(cell.value)) * 1.1 for cell in column_cells)
        wsi.column_dimensions[column_cells[0].column_letter].width = length

    def _card_num(rc):
        try:
            return int(normalize_card_to_num(rc.card))
        except Exception:
            return 999
    def _n(rc):
        return rc.response_num or 0

    response_codes_sorted = sorted(response_codes, key=lambda rc: (_card_num(rc), _n(rc)))
    rows_for_raw = []
    for rc in response_codes_sorted:
        special_s = _normalize_special_tokens(rc.special or "")
        rows_for_raw.append({
            '카드': _card_num(rc),
            'Card': to_roman(rc.card),
            'N': rc.response_num,
            'time': rc.time,
            '반응': rc.response,
            '질문': rc.inquiry,
            'V': rc.rotation,
            'Location': rc.location,
            'Dev Qual': rc.dev_qual,
            'loc_num': rc.loc_num,
            '결정인': rc.determinants,
            'Form Quality': rc.form_qual,
            '내용인': rc.content,
            'P': rc.popular,
            'Z': rc.Z,
            '특수점수': special_s,
        })
    df_raw = pd.DataFrame(rows_for_raw)

    ws_raw = wb.create_sheet(title="반응별 정보")
    if not df_raw.empty:
        # 헤더
        for col_num, column_title in enumerate(df_raw.columns, 1):
            c = ws_raw.cell(row=1, column=col_num, value=column_title)
            c.font = HDR_FONT
            c.fill = PASTEL_FILL
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = Border(top=THIN_EDGE, bottom=THIN_EDGE, left=THIN_EDGE, right=THIN_EDGE)
        # 데이터
        for row_vals in df_raw.values.tolist():
            ws_raw.append(row_vals)
        ws_raw.freeze_panes = "A2"
        ws_raw.auto_filter.ref = f"A1:{get_column_letter(ws_raw.max_column)}1"
        # 폭
        for col in range(1, ws_raw.max_column + 1):
            max_len = 0
            for row in range(1, ws_raw.max_row + 1):
                v = ws_raw.cell(row=row, column=col).value
                max_len = max(max_len, len(str(v)) if v is not None else 0)
            ws_raw.column_dimensions[get_column_letter(col)].width = max(8, min(60, int(max_len * 1.1)))

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = f"{client.name}_{client.testDate:%Y-%m-%d}.xlsx"
    fallback  = f"{slugify(client.name)}_{client.testDate:%Y-%m-%d}.xlsx"

    resp = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = (
        f'attachment; filename="{fallback}"; filename*=UTF-8\'\'{quote(safe_name)}'
    )
    return resp


@group_min_required('intermediate')
def edit_responses(request, client_id):
    response_codes = (
        ResponseCode.objects
        .filter(client_id=client_id)
        .annotate(n_int=Cast(Coalesce('response_num', Value(0)), IntegerField()))
        .order_by('n_int', 'card', 'id')
    )
    current = response_codes.count()
    extra = max(0, min(DEFAULT_EXTRA, TOTAL_CAP - current))
    ResponseCodeFormSet = modelformset_factory(ResponseCode, form=ResponseCodeForm, extra=extra, max_num=TOTAL_CAP)

    if request.method == 'POST':
        formset = ResponseCodeFormSet(request.POST, queryset=response_codes)
        if formset.is_valid():
            fix_count = 0
            for form in formset:
                if form.cleaned_data.get('card') and form.cleaned_data.get('response'):
                    form.instance.client_id = client_id
                    form.instance.card = to_roman(form.cleaned_data.get('card', ''))
                    det_before = (form.instance.determinants or '')
                    form.instance.determinants, notes = _fix_determinant_typos(det_before)
                    if notes:
                        fix_count += 1
                    form.instance.special = _normalize_special_tokens(form.instance.special or '')
                    form.save()
            if fix_count:
                messages.info(request, f"자동 보정 적용 {fix_count}행 (m'p→mp, 특수점수 정규화 등)")
            structural_summary, _ = StructuralSummary.objects.get_or_create(client_id=client_id)
            try:
                structural_summary.calculate_values()
            except Exception:
                structural_summary.save()
            return redirect('scoring:client_list')
        else:
            details = []
            for i, f in enumerate(formset.forms, start=1):
                if f.errors:
                    for k, errs in f.errors.items():
                        details.append(f"{i}행 {k}: {', '.join(errs)}")
            if details:
                shown = " / ".join(details[:10])
                more = f" …외 {len(details)-10}건" if len(details) > 10 else ""
                messages.error(request, "입력 오류: " + shown + more)
    else:
        formset = ResponseCodeFormSet(queryset=response_codes)

    return render(request, 'edit_responses.html', {'formset': formset, 'client_id': client_id})
