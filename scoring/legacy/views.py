import logging
import re

from functools import wraps

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db import transaction
from django.forms import formset_factory, modelformset_factory
from django.http import (
    HttpResponse,
    HttpResponseForbidden,
    HttpResponseNotFound,
    JsonResponse,
)
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.writer.excel import save_virtual_workbook

from .filters import CardImagesFilter, PResponseFilter, SearchReferenceFilter
from .forms import (
    BulkResponseUploadForm,
    ClientForm,
    ResponseCodeForm,
)
from .models import (
    CardImages,
    Client,
    PopularResponse,
    ResponseCode,
    SearchReference,
    StructuralSummary,
)

GROUP_LEVEL = {'beginner': 1, 'intermediate': 2, 'advanced': 3}
GROUP_LABEL = {'beginner': '초급', 'intermediate': '중급', 'advanced': '고급'}

HEADER_MAP = {
    'id': None,  # 무시
    '카드': 'card',
    'n': 'response_num',
    '시간': 'time',
    '반응': 'response',
    '질문': 'inquiry',
    '회전': 'rotation',
    '반응영역': 'location',
    '발달질': 'dev_qual',
    '영역번호': 'loc_num',
    '(2)': 'pair',
    '결정인': 'determinants',
    '형태질': 'form_qual',
    '내용인': 'content',
    'p': 'popular',
    'z': 'Z',
    '특수점수': 'special',
    '코멘트': 'comment',
    # 영문
    'card': 'card',
    'response_num': 'response_num',
    'time': 'time',
    'response': 'response',
    'inquiry': 'inquiry',
    'rotation': 'rotation',
    'location': 'location',
    'dev_qual': 'dev_qual',
    'loc_num': 'loc_num',
    'determinants': 'determinants',
    'form_qual': 'form_qual',
    'pair': 'pair',
    'content': 'content',
    'popular': 'popular',
    'z': 'Z',  # 중복 허용
    'special': 'special',
    'comment': 'comment',
}

REQUIRED_FIELDS = [
    'card','response_num','time','response','inquiry','rotation','location',
    'dev_qual','loc_num','determinants','form_qual','pair','content','popular','Z','special','comment'
]
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.writer.excel import save_virtual_workbook


from django.db.models import Case, When, Value, IntegerField

roman_to_num = {'I':'1','II':'2','III':'3','IV':'4','V':'5','VI':'6','VII':'7','VIII':'8','IX':'9','X':'10'}
num_to_roman = {v:k for k,v in roman_to_num.items()}

def normalize_card_to_num(val: str) -> str:
    s = str(val).strip().upper()
    return roman_to_num.get(s, s)  # 로마→숫자문자열, 숫자면 그대로

def to_roman(val: str) -> str:
    s = normalize_card_to_num(val)  # '1'..'10'
    return num_to_roman.get(s, s)

def _pick_input_sheet(wb):

    for name in ('입력', 'responses'):
        if name in wb.sheetnames:
            return wb[name]

    for name in wb.sheetnames:
        ws_try = wb[name]
        raw = [(c.value or '') for c in ws_try[1]]
        mapped = [normalize_header(h) for h in raw]
        idx_map = {f: i for i, f in enumerate(mapped) if f}
        if all(f in idx_map for f in REQUIRED_FIELDS):
            return ws_try

    # 4) fallback
    return wb.active

def normalize_header(h: str) -> str | None:
    if h is None:
        return None
    s = str(h).strip()
    key = s.lower()
    return HEADER_MAP.get(s, HEADER_MAP.get(key, None))

def group_min_required(min_group_name):
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return login_required(view_func)(request, *args, **kwargs)
            user_level = GROUP_LEVEL.get(getattr(request.user, 'group', None), 0)
            required_level = GROUP_LEVEL[min_group_name]
            if user_level >= required_level:
                return view_func(request, *args, **kwargs)
            return HttpResponseForbidden(f"{GROUP_LABEL[min_group_name]} 이상 이수자만 접속 가능한 페이지입니다.")
        return _wrapped_view
    return decorator

@group_min_required('intermediate')
def download_response_template(request):
    wb = Workbook()

    ws = wb.active
    ws.title = "입력"

    headers = [
        'ID','카드','N','시간','반응','질문','회전','반응영역','발달질','영역번호',
        '결정인','형태질','(2)','내용인','P','Z','특수점수','코멘트'
    ]
    ws.append(headers)

    ws.append([
        '', 1, 1, "12s", "박쥐", "윗부분이 날개 같아요", '', 'W', '+', '',
        'M.F', '+', '', 'H', 'P', 'ZW', '', ''
    ])

    head_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
    bold = Font(bold=True)
    for col in range(1, len(headers)+1):
        c = ws.cell(row=1, column=col)
        c.font = bold
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.auto_filter.ref = f"A1:{chr(64+len(headers))}1"
    ws.freeze_panes = "A2"

    widths = [6, 6, 6, 8, 28, 28, 8, 10, 8, 10, 12, 10, 6, 10, 6, 6, 10, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w

    help_ws = wb.create_sheet("도움말")
    help_ws.append(["열 이름", "설명(업로드 시 매핑되는 내부 필드)"])
    mapping = [
        ("ID", "서버에서 사용하지 않음(무시됨)"),
        ("카드", "card"),
        ("N", "response_num(반응 번호; 정수)"),
        ("시간", "time"),
        ("반응", "response"),
        ("질문", "inquiry"),
        ("회전", "rotation"),
        ("반응영역", "location"),
        ("발달질", "dev_qual"),
        ("영역번호", "loc_num(정수)"),
        ("결정인", "determinants"),
        ("형태질", "form_qual"),
        ("(2)", "pair"),
        ("내용인", "content"),
        ("P", "popular"),
        ("Z", "Z"),
        ("특수점수", "special"),
        ("코멘트", "comment"),
    ]
    for r in mapping:
        help_ws.append(list(r))
    for col in ("A","B"):
        help_ws.column_dimensions[col].width = 36
    for col in range(1,3):
        help_ws.cell(row=1, column=col).font = bold
        help_ws.cell(row=1, column=col).fill = head_fill
        help_ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

    # 응답 전송
    xlsx = save_virtual_workbook(wb)
    resp = HttpResponse(
        xlsx,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="response_template_ko.xlsx"'
    return resp

@group_min_required('intermediate')
def search(request, client_id=None):
    client_id = client_id or request.GET.get('client_id')
    if not client_id:
        return HttpResponseNotFound("client_id가 필요합니다.")

    client = get_object_or_404(Client, id=client_id)

    # 소유자 확인
    if client.tester != request.user:
        return HttpResponse("액세스 거부: 작성 권한이 없습니다.", status=403)

    extra_forms = 40
    max_total = 70
    ResponseCodeFormSet = formset_factory(ResponseCodeForm, extra=extra_forms, max_num=max_total)

    if request.method == 'POST' and request.POST.get('mode') == 'upload_preview':
        xfile = request.FILES.get('xlsx_file')
        if not xfile:
            messages.error(request, "엑셀 파일을 선택해 주세요.")
            formset = ResponseCodeFormSet()
        else:
            try:
                wb = load_workbook(filename=xfile, data_only=True)
                ws = wb.active
            except Exception:
                messages.error(request, "엑셀 파일을 열 수 없습니다. (.xlsx 형식 확인)")
                formset = ResponseCodeFormSet()
            else:
                raw_headers = [(c.value or '') for c in ws[1]]
                mapped = [normalize_header(h) for h in raw_headers]
                index_by_field = {f: idx for idx, f in enumerate(mapped) if f}

                missing = [f for f in REQUIRED_FIELDS if f not in index_by_field]
                if missing:
                    messages.error(request, "필수 열 누락: " + ", ".join(missing) + "  (샘플 템플릿을 사용하세요)")
                    formset = ResponseCodeFormSet()
                else:
                    rows = list(ws.iter_rows(min_row=2, values_only=True))
                    initial = []
                    for row in rows:
                        data = {}
                        for f in REQUIRED_FIELDS:
                            idx = index_by_field[f]
                            v = row[idx] if idx < len(row) else ''
                            data[f] = '' if v is None else v
                        # 타입 보정
                        for int_key in ('response_num', 'loc_num'):
                            if str(data[int_key]).strip():
                                try:
                                    data[int_key] = int(data[int_key])
                                except Exception:
                                    pass
                        initial.append(data)

                    if len(initial) > max_total:
                        initial = initial[:max_total]
                        messages.warning(request, f"최대 {max_total}행까지만 미리 채웁니다.")

                    formset = ResponseCodeFormSet(initial=initial)
                    messages.success(request, f"엑셀에서 {len(initial)}건을 불러왔습니다. 확인 후 저장하세요.")

        reference_list = SearchReference.objects.all()
        card_image_list = CardImages.objects.all()
        Presponse_list = PopularResponse.objects.all()
        search_filter = SearchReferenceFilter(request.GET, queryset=reference_list)
        card_image_filter = CardImagesFilter(request.GET, queryset=card_image_list)
        p_response_filter = PResponseFilter(request.GET, queryset=Presponse_list)

        return render(
            request,
            'intermediate.html',
            {
                'client': client,
                'formset': formset,
                'filter': search_filter,
                'image_filter': card_image_filter,
                'p_response_filter': p_response_filter,
            },
        )

    if request.method == 'POST':
        if request.POST.get('additems') == 'true':
            formset_dictionary_copy = request.POST.copy()
            formset_dictionary_copy['form-TOTAL_FORMS'] = int(formset_dictionary_copy['form-TOTAL_FORMS']) + 1
            formset = ResponseCodeFormSet(formset_dictionary_copy)
        else:
            formset = ResponseCodeFormSet(request.POST)
            if formset.is_valid():
                warnings = []  # 권고 위반 행 번호 모으기
                created = 0
                for idx, form in enumerate(formset.forms, start=1):
                    cd = getattr(form, 'cleaned_data', {}) or {}
                    # 비어있는 행은 무시
                    if not (cd.get('card') and cd.get('response')):
                        continue

                    # 권고 규칙(저장은 허용, 경고만)
                    card = str(cd.get('card', '')).strip().upper()
                    loc  = cd.get('location', '') or ''
                    dq   = cd.get('dev_qual', '') or ''
                    zval = cd.get('Z', '') or ''
                    if card in {'1','4','5','I','II','III'} and 'W' in loc and '+' in dq and zval == 'ZW':
                        warnings.append(idx)

                    form.instance.client = client
                    form.save()
                    created += 1

                if warnings:
                    messages.warning(
                        request,
                        f"Z점수 재검토 권고 {len(warnings)}건(행: {', '.join(map(str, warnings))}). "
                        "Z가 결과(예: Zf/Zsum/Zd)에 영향을 줄 수 있으니 검토하세요."
                    )
                messages.success(request, f"{created}건 저장되었습니다.")
                return redirect('scoring:client_detail', client_id=client.id)
            else:
                details = []
                for i, f in enumerate(formset.forms, start=1):
                    if f.errors:
                        for k, errs in f.errors.items():
                            details.append(f"{i}행 {k}: {', '.join(errs)}")
                if details:
                    messages.error(request, "입력 오류: " + " / ".join(details))
    else:
        formset = ResponseCodeFormSet()


    reference_list = SearchReference.objects.all()
    card_image_list = CardImages.objects.all()
    Presponse_list = PopularResponse.objects.all()

    search_filter = SearchReferenceFilter(request.GET, queryset=reference_list)
    card_image_filter = CardImagesFilter(request.GET, queryset=card_image_list)
    p_response_filter = PResponseFilter(request.GET, queryset=Presponse_list)

    if request.GET.get('Card'):
        card_number = request.GET['Card']
        card_image_filter = CardImagesFilter(request.GET, queryset=CardImages.objects.filter(card_number=card_number))
        p_response_filter = PResponseFilter(request.GET, queryset=PopularResponse.objects.filter(card_number=card_number))

    return render(
        request,
        'intermediate.html',
        {
            'client': client,
            'formset': formset,
            'filter': search_filter,
            'image_filter': card_image_filter,
            'p_response_filter': p_response_filter,
        },
    )

@group_min_required('intermediate')
def update_response_codes(request, client_id):
    client = get_object_or_404(Client, id=client_id)
    if client.tester != request.user:
        return HttpResponse("액세스 거부: 작성 권한이 없습니다.", status=403)

    response_codes = ResponseCode.objects.filter(client=client)
    extra_forms = 40
    ResponseCodeFormSet = modelformset_factory(ResponseCode, form=ResponseCodeForm, extra=extra_forms, max_num=70)
    if request.method == 'POST':
        if request.POST.get('additems') == 'true':
            formset_dictionary_copy = request.POST.copy()
            formset_dictionary_copy['form-TOTAL_FORMS'] = int(formset_dictionary_copy['form-TOTAL_FORMS']) + 1
            formset = ResponseCodeFormSet(formset_dictionary_copy)
        else:
            formset = ResponseCodeFormSet(request.POST, queryset=response_codes)
            if formset.is_valid():
                for form in formset:
                    if form.cleaned_data.get('card') and form.cleaned_data.get('response'):
                        form.instance.client = client
                        form.save()
                return redirect('client_list')
    else:
        formset = ResponseCodeFormSet(queryset=response_codes)

    reference_list = SearchReference.objects.all()
    card_image_list = CardImages.objects.all()
    Presponse_list = PopularResponse.objects.all()

    search_filter = SearchReferenceFilter(request.GET, queryset=reference_list)
    card_image_filter = CardImagesFilter(request.GET, queryset=card_image_list)
    p_response_filter = PResponseFilter(request.GET, queryset=Presponse_list)

    if request.GET.get('Card'):
        card_number = request.GET['Card']
        card_image_filter = CardImagesFilter(request.GET, queryset=CardImages.objects.filter(card_number=card_number))
        p_response_filter = PResponseFilter(request.GET, queryset=PopularResponse.objects.filter(card_number=card_number))

    return render(
        request,
        'update_response_codes.html',
        {
            'client': client,
            'formset': formset,
            'filter': search_filter,
            'image_filter': card_image_filter,
            'p_response_filter': p_response_filter,
        },
    )

@group_min_required('advanced')
def advanced_upload(request, client_id):
    client = get_object_or_404(Client, id=client_id)

    if client.tester != request.user:
        return HttpResponseForbidden("본인 수검자에게만 업로드할 수 있습니다.")

    existing_count = ResponseCode.objects.filter(client=client).count()
    has_existing = existing_count > 0

    form = BulkResponseUploadForm(request.user, request.POST or None, request.FILES or None)
    form.fields['client'].queryset = Client.objects.filter(id=client.id, tester=request.user)
    if request.method == 'GET' and not form.is_bound:
        form.initial['client'] = client

    if request.method == 'POST' and form.is_valid():
        posted_client = form.cleaned_data['client']
        if posted_client.id != client.id:
            messages.error(request, "요청 경로의 수검자와 폼의 수검자가 일치하지 않습니다.")
            return render(request, 'advanced_upload.html', {
                'client': client, 'form': form,
                'has_existing': has_existing, 'existing_count': existing_count,
            })

        xfile = form.cleaned_data['file']
        replace = form.cleaned_data['replace_existing']

        try:
            wb = load_workbook(filename=xfile, data_only=True)
        except Exception:
            messages.error(request, "엑셀 파일을 열 수 없습니다. (.xlsx 형식 확인)")
            return render(request, 'advanced_upload.html', {
                'client': client, 'form': form,
                'has_existing': has_existing, 'existing_count': existing_count,
            })

        ws = wb.active
        raw_headers = [(c.value or '') for c in ws[1]]
        mapped = [normalize_header(h) for h in raw_headers]
        index_by_field = {f: idx for idx, f in enumerate(mapped) if f}

        missing = [f for f in REQUIRED_FIELDS if f not in index_by_field]
        if missing:
            messages.error(
                request,
                "필수 열이 누락되었습니다: " + ", ".join(missing) + "  (샘플 템플릿을 내려받아 그대로 사용하세요.)"
            )
            return render(request, 'advanced_upload.html', {
                'client': client, 'form': form,
                'has_existing': has_existing, 'existing_count': existing_count,
            })

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            messages.warning(request, "데이터 행이 없습니다.")
            return render(request, 'advanced_upload.html', {
                'client': client, 'form': form,
                'has_existing': has_existing, 'existing_count': existing_count,
            })

        created, errors = 0, []
        with transaction.atomic():
            if replace:
                ResponseCode.objects.filter(client=client).delete()
            for idx, row in enumerate(rows, start=2):
                data = {}
                for f in REQUIRED_FIELDS:
                    # 셀 값 읽기
                    v = row[index_by_field[f]] if index_by_field[f] < len(row) else None

                    # 숫자 선택 필드 정규화
                    if f in ('loc_num',):
                        data[f] = None if v in (None, '', ' ') else v
                    else:
                        # 그 외는 기존대로
                        data[f] = '' if v is None else v

                # 필수 정수 필드 캐스팅
                for int_key in ('response_num',):
                    if str(data[int_key]).strip():
                        try:
                            data[int_key] = int(data[int_key])
                        except Exception:
                            errors.append(f"{idx}행: {int_key} 정수 변환 실패")
                            continue

                # loc_num은 선택이므로 빈칸이면 None 그대로 두면 됩니다(캐스팅 불필요)

                form_row = ResponseCodeForm(data)
                if not form_row.is_valid():
                    errs = '; '.join([f"{fld}: {','.join(e)}" for fld, e in form_row.errors.items()])
                    errors.append(f"{idx}행 유효성 오류 → {errs}")
                    continue

                obj = form_row.save(commit=False)
                obj.client = client
                obj.save()
                created += 1

        if errors:
            messages.warning(request, f"성공 {created}건, 오류 {len(errors)}건")
        else:
            messages.success(request, f"성공 {created}건 업로드 완료")

        return redirect('scoring:client_detail', client_id=client.id)
    
    return render(request, 'advanced_upload.html', {
        'client': client, 'form': form,
        'has_existing': has_existing, 'existing_count': existing_count,
    })

def search_results(request):
    reference_list = SearchReference.objects.all()
    card_image_list = CardImages.objects.all()
    Presponse_list = PopularResponse.objects.all()

    search_filter = SearchReferenceFilter(request.GET, queryset=reference_list)
    card_image_filter = CardImagesFilter(request.GET, queryset=card_image_list)
    p_response_filter = PResponseFilter(request.GET, queryset=Presponse_list)

    if request.GET.get('Card'):
        card_number = request.GET['Card']
        card_image_filter = CardImagesFilter(request.GET, queryset=CardImages.objects.filter(card_number=card_number))
        p_response_filter = PResponseFilter(request.GET, queryset=PopularResponse.objects.filter(card_number=card_number))

    return render(
        request,
        'search_results.html',
        {'filter': search_filter, 'image_filter': card_image_filter, 'p_response_filter': p_response_filter},
    )


@group_min_required('intermediate')
def add_client(request):
    next_step = request.POST.get('next') or request.GET.get('next', 'intermediate')

    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            client_instance = form.save(commit=False)
            client_instance.tester = request.user
            client_instance.save()

            # next에 따라 분기
            if next_step == 'advanced':
                return redirect('scoring:advanced_upload', client_id=client_instance.id)
            
            return redirect('scoring:search', client_id=client_instance.id)
    else:
        form = ClientForm(initial={
            'name': request.GET.get('name', ''),
            'gender': request.GET.get('gender', ''),
            'birthdate': request.GET.get('birthdate', ''),
            'testDate': request.GET.get('testDate', ''),
            'notes': request.GET.get('notes', ''),
            'consent': request.GET.get('consent', ''),
        })
    return render(request, 'add_client.html', {'form': form})

@group_min_required('intermediate')
def client_list(request):
    user = request.user
    clients = Client.objects.filter(tester=user)
    return render(request, 'client_list.html', {'clients': clients})


@group_min_required('intermediate')
def client_detail(request, client_id):
    user = request.user
    client_obj = get_object_or_404(Client, id=client_id, tester=user)
    response_codes = ResponseCode.objects.filter(client=client_obj)
    return render(request, 'client_detail.html', {'client': client_obj, 'response_codes': response_codes})


# 구조적 요약 엑셀 내보내기
@group_min_required('intermediate')
def export_structural_summary_xlsx(request, client_id):
    try:
        client = Client.objects.get(id=client_id)
        if client.tester != request.user:
            return HttpResponse("액세스 거부: 해당 정보를 볼 수 있는 권한이 없습니다.", status=403)

        response_codes = ResponseCode.objects.filter(client_id=client_id)
        ##
        roman_to_num = {'I':'1','II':'2','III':'3','IV':'4','V':'5','VI':'6','VII':'7','VIII':'8','IX':'9','X':'10'}
        num_to_roman = {v:k for k,v in roman_to_num.items()}

        def normalize_card(val):
            s = str(val).strip().upper()
            return roman_to_num.get(s, s)  # 로마면 숫자 문자열로, 아니면 그대로(숫자도 허용)

        numbers_found = { normalize_card(rc.card) for rc in response_codes if rc.card }

        required = {str(i) for i in range(1, 11)}  # '1'..'10'
        missing = sorted(required - numbers_found, key=int)

        if missing:
            missing_roman = [num_to_roman.get(n, n) for n in missing]
            return HttpResponse("다음 카드의 반응이 없습니다: " + ", ".join(missing_roman))
        ## 
        numbers_found = set()
        roman_dict = {
            'I': '1', 'II': '2', 'III': '3', 'IV': '4', 'V': '5',
            'VI': '6', 'VII': '7', 'VIII': '8', 'IX': '9', 'X': '10'
        }
        for response_code in response_codes:
            card_value = response_code.card
            card_val = roman_dict.get(card_value, card_value)
            numbers_found.add(card_val)

        if len(numbers_found) < 10:
            return HttpResponse("한 카드에 적어도 하나의 반응을 입력해야 합니다.")

        structural_summary, created = StructuralSummary.objects.get_or_create(client_id=client_id)
        if not created:
            structural_summary.calculate_values()
    except Client.DoesNotExist:
        logging.error("해당 ID의 클라이언트를 찾을 수 없음")
        return HttpResponseNotFound("클라이언트 정보를 찾을 수 없습니다.")
    except Exception as e:
        logging.error(f"예기치 못한 오류 발생: {e}")
        error_message = f"예기치 못한 오류 발생: {type(e).__name__}, {str(e)}"
        return JsonResponse({'error': error_message}, status=500)

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    ws = wb.create_sheet(title='상단부')
    wsd = wb.create_sheet(title='하단부')
    wsp = wb.create_sheet(title='특수지표')
    ws.merge_cells('A4:B4')
    ws.merge_cells('A14:B14')
    ws.merge_cells('A20:D20')
    ws.merge_cells('F4:H4')
    ws.merge_cells('G5:H5')
    ws.merge_cells('J4:K4')
    ws.merge_cells('M4:N4')
    ws.merge_cells('M16:P16')
    wsd.merge_cells('A3:F3')
    wsd.merge_cells('H3:I3')
    wsd.merge_cells('K3:N3')
    wsd.merge_cells('K5:L5')
    wsd.merge_cells('K6:L6')
    wsd.merge_cells('K7:L7')
    wsd.merge_cells('K8:L8')
    wsd.merge_cells('K9:L9')
    wsd.merge_cells('K10:L10')
    wsd.merge_cells('K11:L11')
    wsd.merge_cells('K12:L12')
    wsd.merge_cells('A14:D14')
    wsd.merge_cells('F14:G14')
    wsd.merge_cells('I14:J14')
    wsd.merge_cells('L14:M14')

    bckground_cells = ['A4', 'A14', 'A20', 'F4', 'F5', 'G5', 'J4', 'M4', 'M16']
    for cell in bckground_cells:
        ws[cell].fill = PatternFill(start_color="7B68EE", end_color="7B68EE", fill_type='solid')
    bckground_cells2 = ['A3', 'H3', 'K3', 'A14', 'F14', 'I14', 'L14']
    for cell in bckground_cells2:
        wsd[cell].fill = PatternFill(start_color="7B68EE", end_color="7B68EE", fill_type='solid')

    ws.sheet_view.showGridLines = False
    wsd.sheet_view.showGridLines = False

    BORDER_LIST = ['A5:B12', 'A4:B4', 'A14:B14', 'A15:B18', 'A20:D20', 'A21:D21', 'A22:D26', 'F4:H4', 'F5:F5', 'G5:H5',
                   'F6:F29', 'G6:H29', 'J4:K4', 'J5:K31', 'M4:N4', 'M5:N14', 'M16:P16', 'M17:P17', 'M18:P25', 'M26:P30']
    BORDER_LIST2 = ['A3:F3', 'A4:F4', 'A5:F7', 'A8:F9', 'H3:I3', 'H4:I10', 'K3:N3', 'K4:N12', 'A14:D14', 'A15:D19',
                    'F14:G14', 'F15:G21', 'I14:J14', 'I15:J21', 'L14:M14', 'L15:M21']

    def set_border(worksheet, cell_range):
        rows = worksheet[cell_range]
        side = Side(border_style='thin', color="FF000000")

        rows = list(rows)
        max_y = len(rows) - 1
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1
            for pos_x, c in enumerate(cells):
                border = Border(
                    left=c.border.left,
                    right=c.border.right,
                    top=c.border.top,
                    bottom=c.border.bottom
                )
                if pos_x == 0:
                    border.left = side
                if pos_x == max_x:
                    border.right = side
                if pos_y == 0:
                    border.top = side
                if pos_y == max_y:
                    border.bottom = side
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    c.border = border

    for pos in BORDER_LIST:
        set_border(ws, pos)
    for pos in BORDER_LIST2:
        set_border(wsd, pos)

    
    ws['A4'] = 'Location Features'
    ws['A5'] = 'Zf'
    ws['A6'] = 'Zsum'
    ws['A7'] = 'Zest'
    ws['A9'] = 'W'
    ws['A10'] = 'D'
    ws['A11'] = 'Dd'
    ws['A12'] = 'S'

    ws['A14'] = 'Developmental Quality'
    ws['A15'] = '+'
    ws['A16'] = 'o'
    ws['A17'] = 'v/+'
    ws['A18'] = 'v'

    ws['A20'] = 'Form Quality'
    ws['B21'] = 'FQx'
    ws['C21'] = 'MQual'
    ws['D21'] = 'W+D'
    ws['A22'] = '+'
    ws['A23'] = 'o'
    ws['A24'] = 'u'
    ws['A25'] = '-'
    ws['A26'] = 'none'

    ws['F4'] = 'Determinants'
    ws['F5'] = 'Blends'
    ws['G5'] = 'Single'
    fields = [
        'M', 'FM', 'm', 'FC', 'CF', 'C', 'Cn', "FC'", "C'F", "C'",
        'FT', 'TF', 'T', 'FV', 'VF', 'V', 'FY', 'YF', 'Y', 'Fr', 'rF', 'FD', 'F', '(2)'
    ]
    real_field = [
        'M', 'FM', 'm_l', 'FC', 'CF', 'C', 'Cn', 'FCa', 'CaF', 'Ca',
        'FT', 'TF', 'T', 'FV', 'VF', 'V', 'FY', 'YF', 'Y', 'Fr', 'rF', 'FD', 'F', 'pair'
    ]
    s_row = 6
    for field_name in fields:
        ws.cell(row=s_row, column=7, value=field_name)
        s_row += 1

    ws['J4'] = 'Contents'
    cont_fields = [
        'H', '(H)', 'Hd', '(Hd)', 'Hx', 'A', '(A)', 'Ad', '(Ad)', 'An',
        'Art', 'Ay', 'Bl', 'Bt', 'Cg', 'Cl', 'Ex', 'Fd', 'Fi', 'Ge', 'Hh', 'Ls',
        'Na', 'Sc', 'Sx', 'Xy', 'Id'
    ]
    cont_real_fields = [
        'H', 'H_paren', 'Hd', 'Hd_paren', 'Hx', 'A', 'A_paren', 'Ad', 'Ad_paren', 'An',
        'Art', 'Ay', 'Bl', 'Bt', 'Cg', 'Cl', 'Ex', 'Fd_l', 'Fi', 'Ge', 'Hh', 'Ls',
        'Na', 'Sc', 'Sx', 'Xy', 'Idio'
    ]
    s_row = 5
    for field_name in cont_fields:
        ws.cell(row=s_row, column=10, value=field_name)
        s_row += 1

    ws['M4'] = "approach"
    for i, label in enumerate(['I','II','III','IV','V','VI','VII','VIII','IX','X'], start=5):
        ws.cell(row=i, column=13 if i == 5 else 13, value=label)

    ws['M16'] = 'Special Scores'
    ws['N17'] = 'Lvl-1'
    ws['O17'] = 'Lvl-2'
    sp_fields = ['DV', 'INC', 'DR', 'FAB', 'ALOG', 'CON']
    sp_real_fields = ['sp_dv', 'sp_inc', 'sp_dr', 'sp_fab', 'sp_alog', 'sp_con']
    sp_real_fields2 = ['sp_dv2', 'sp_inc2', 'sp_dr2', 'sp_fab2']
    s_row = 18
    for field_name in sp_fields:
        ws.cell(row=s_row, column=13, value=field_name)
        s_row += 1
    ws['M24'] = 'Raw Sum6'
    ws['M25'] = 'Weighted Sum6'
    ws['M26'] = 'AB'
    ws['M27'] = 'AG'
    ws['M28'] = 'COP'
    ws['M29'] = 'CP'
    ws['O26'] = 'GHR'
    ws['O27'] = 'PHR'
    ws['O28'] = 'MOR'
    ws['O29'] = 'PER'
    ws['O30'] = 'PSV'

    wsd['A3'] = 'Core'
    wsd['A4'] = 'R'
    wsd['C4'] = 'L'
    wsd['A5'] = 'EB'
    wsd['A6'] = 'eb'
    wsd['C5'] = 'EA'
    wsd['C6'] = 'es'
    wsd['C7'] = 'Adj es'
    wsd['E5'] = 'EBper'
    wsd['E6'] = 'D'
    wsd['E7'] = 'Adj D'
    wsd['A8'] = 'FM'
    wsd['A9'] = 'm'
    wsd['C8'] = "SumC'"
    wsd['C9'] = 'SumV'
    wsd['E8'] = "SumT"
    wsd['E9'] = 'SumY'

    wsd['H3'] = 'Affect'
    wsd['H4'] = 'FC:CF+C'
    wsd['H5'] = "Pure C"
    wsd['H6'] = "SumC':WsumC"
    wsd['H7'] = 'Afr'
    wsd['H8'] = "S"
    wsd['H9'] = "Blends:R"
    wsd['H10'] = "CP"

    wsd['K3'] = 'Interpersonal'
    wsd['K4'] = 'COP'
    wsd['M4'] = 'AG'
    wsd['K5'] = 'GHR:PHR'
    wsd['K6'] = 'a:p'
    wsd['K7'] = 'Food'
    wsd['K8'] = 'SumT'
    wsd['K9'] = 'Human Content'
    wsd['K10'] = 'Pure H'
    wsd['K11'] = 'PER'
    wsd['K12'] = 'Isolation Index'

    wsd['A14'] = 'Ideation'
    wsd['A15'] = 'a:p'
    wsd['A16'] = 'Ma:Mp'
    wsd['A17'] = 'Intel(2AB+Art+Ay)'
    wsd['A18'] = 'MOR'
    wsd['C15'] = 'Sum6'
    wsd['C16'] = 'Lvl-2'
    wsd['C17'] = 'Wsum6'
    wsd['C18'] = 'M-'
    wsd['C19'] = 'M none'

    wsd['F14'] = 'Mediation'
    med_fields = ['XA%', 'WDA%', 'X-%', 'S-', 'P', 'X+%', 'Xu%']
    med_real_fields = ['xa_per', 'wda_per', 'x_minus_per', 's_minus', 'popular', 'x_plus_per', 'xu_per']
    s_row = 15
    for field_name in med_fields:
        wsd.cell(row=s_row, column=6, value=field_name)
        s_row += 1

    wsd['I14'] = 'Processing'
    pro_fields = ['Zf', 'W:D:Dd', 'W:M', 'Zd', 'PSV', 'DQ+', 'DQv']
    pro_real_fields = ['Zf', 'W_D_Dd', 'W_M', 'Zd', 'sp_psv', 'dev_plus', 'dev_v']
    s_row = 15
    for field_name in pro_fields:
        wsd.cell(row=s_row, column=9, value=field_name)
        s_row += 1

    wsd['L14'] = 'Self'
    self_fields = ['Ego[3r+(2)/R]', 'Fr+rF', 'SumV', 'FD', 'An+Xy', 'MOR', 'H:(H)+Hd+(Hd)']
    self_real_fields = ['ego', 'fr_rf', 'sum_V', 'fdn', 'an_xy', 'sp_mor', 'h_prop']
    s_row = 15
    for field_name in self_fields:
        wsd.cell(row=s_row, column=12, value=field_name)
        s_row += 1

    ws['B5'] = structural_summary.Zf
    ws['B6'] = structural_summary.Zsum
    ws['B6'].number_format = '0.0'
    ws['B7'] = structural_summary.Zest
    ws['B7'].number_format = '0.0'
    ws['B9'] = structural_summary.W
    ws['B10'] = structural_summary.D
    ws['B11'] = structural_summary.Dd
    ws['B12'] = structural_summary.S

    ws['B15'] = structural_summary.dev_plus
    ws['B16'] = structural_summary.dev_o
    ws['B17'] = structural_summary.dev_vplus
    ws['B18'] = structural_summary.dev_v

    ws['B22'] = structural_summary.fqx_plus
    ws['B23'] = structural_summary.fqx_o
    ws['B24'] = structural_summary.fqx_u
    ws['B25'] = structural_summary.fqx_minus
    ws['B26'] = structural_summary.fqx_none
    ws['C22'] = structural_summary.mq_plus
    ws['C23'] = structural_summary.mq_o
    ws['C24'] = structural_summary.mq_u
    ws['C25'] = structural_summary.mq_minus
    ws['C26'] = structural_summary.mq_none
    ws['D22'] = structural_summary.wd_plus
    ws['D23'] = structural_summary.wd_o
    ws['D24'] = structural_summary.wd_u
    ws['D25'] = structural_summary.wd_minus
    ws['D26'] = structural_summary.wd_none

    blends = structural_summary.blends.split(',')
    start_row = 6
    start_column = 6
    for blend in blends:
        ws.cell(row=start_row, column=start_column, value=blend)
        start_row += 1

    row = 6
    for field_name in real_field:
        field_value = getattr(structural_summary, field_name)
        ws.cell(row=row, column=8, value=field_value)
        ws.cell(row=row, column=8, value=field_value).number_format = "#"
        row += 1

    row = 5
    for field_name in cont_real_fields:
        field_value = getattr(structural_summary, field_name)
        ws.cell(row=row, column=11, value=field_value)
        ws.cell(row=row, column=11, value=field_value).number_format = "#"
        row += 1

    ws['N5'] = structural_summary.app_I
    ws['N6'] = structural_summary.app_II
    ws['N7'] = structural_summary.app_III
    ws['N8'] = structural_summary.app_IV
    ws['N9'] = structural_summary.app_V
    ws['N10'] = structural_summary.app_VI
    ws['N11'] = structural_summary.app_VII
    ws['N12'] = structural_summary.app_VIII
    ws['N13'] = structural_summary.app_IX
    ws['N14'] = structural_summary.app_X

    lv1_row = 18
    for field_name in sp_real_fields:
        field_value = getattr(structural_summary, field_name)
        ws.cell(row=lv1_row, column=14, value=field_value)
        lv1_row += 1
    lv2_row = 18
    for field_name in sp_real_fields2:
        field_value = getattr(structural_summary, field_name)
        ws.cell(row=lv2_row, column=15, value=field_value)
        lv2_row += 1
    ws['N24'] = structural_summary.sum6
    ws['N25'] = structural_summary.wsum6
    ws['N26'] = structural_summary.sp_ab
    ws['N27'] = structural_summary.sp_ag
    ws['N28'] = structural_summary.sp_cop
    ws['N29'] = structural_summary.sp_cp
    ws['P26'] = structural_summary.sp_ghr
    ws['P27'] = structural_summary.sp_phr
    ws['P28'] = structural_summary.sp_mor
    ws['P29'] = structural_summary.sp_per
    ws['P30'] = structural_summary.sp_psv

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) * 1.1 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length

    wsd['B4'] = structural_summary.R
    wsd['D4'] = structural_summary.L
    wsd['B5'] = structural_summary.ErleBnistypus
    wsd['B6'] = structural_summary.eb
    wsd['D5'] = structural_summary.EA
    wsd['D6'] = structural_summary.es
    wsd['D7'] = structural_summary.adj_es
    wsd['F5'] = 'NA' if structural_summary.EBper == 0 else structural_summary.EBper
    wsd['F6'] = structural_summary.D_score
    wsd['F7'] = structural_summary.adj_D
    wsd['B8'] = structural_summary.sum_FM
    wsd['B9'] = structural_summary.sum_m
    wsd['D8'] = structural_summary.sum_Ca
    wsd['D9'] = structural_summary.sum_V
    wsd['F8'] = structural_summary.sum_T
    wsd['F9'] = structural_summary.sum_Y

    wsd['I4'] = structural_summary.f_c_prop
    wsd['I4'].alignment = Alignment(horizontal='right')
    wsd['I5'] = structural_summary.pure_c
    wsd['I6'] = structural_summary.ca_c_prop
    wsd['I6'].alignment = Alignment(horizontal='right')
    wsd['I7'] = structural_summary.afr
    wsd['I7'].number_format = "0.##;-0.##;0"
    wsd['I8'] = structural_summary.S
    wsd['I9'] = structural_summary.blends_r
    wsd['I9'].alignment = Alignment(horizontal='right')
    wsd['I10'] = structural_summary.sp_cp

    wsd['L4'] = structural_summary.sp_cop
    wsd['N4'] = structural_summary.sp_ag
    wsd['M5'] = structural_summary.GHR_PHR
    wsd['M5'].alignment = Alignment(horizontal='right')
    wsd['M6'] = structural_summary.a_p
    wsd['M6'].alignment = Alignment(horizontal='right')
    wsd['M7'] = structural_summary.Fd_l
    wsd['M8'] = structural_summary.sum_T
    wsd['M9'] = structural_summary.human_cont
    wsd['M10'] = structural_summary.H
    wsd['M11'] = structural_summary.sp_per
    wsd['M12'] = structural_summary.Isol

    wsd['B15'] = structural_summary.a_p
    wsd['B15'].alignment = Alignment(horizontal='right')
    wsd['B16'] = structural_summary.Ma_Mp
    wsd['B16'].alignment = Alignment(horizontal='right')
    wsd['B17'] = structural_summary.intel
    wsd['B18'] = structural_summary.sp_mor
    wsd['D15'] = structural_summary.sum6
    wsd['D16'] = structural_summary.Lvl_2
    wsd['D17'] = structural_summary.wsum6
    wsd['D18'] = structural_summary.mq_minus
    wsd['D19'] = structural_summary.mq_none

    row = 15
    for field_name in med_real_fields:
        field_value = getattr(structural_summary, field_name)
        wsd.cell(row=row, column=7, value=field_value)
        if isinstance(field_value, int):
            number_format = "0"
        else:
            number_format = "0.00"
        wsd.cell(row=row, column=7).number_format = number_format
        row += 1

    row = 15
    for field_name in pro_real_fields:
        field_value = getattr(structural_summary, field_name)
        wsd.cell(row=row, column=10, value=field_value)
        wsd.cell(row=row, column=10).alignment = Alignment(horizontal='right')
        row += 1

    row = 15
    for field_name in self_real_fields:
        field_value = getattr(structural_summary, field_name)
        wsd.cell(row=row, column=13, value=field_value)
        row += 1

    wsd['A23'] = f"PTI={structural_summary.sumPTI}"
    wsd['B23'] = "☑" if structural_summary.sumDEPI >= 5 else "☐"
    wsd['B23'].alignment = Alignment(horizontal='right')
    wsd['C23'] = f"DEPI={structural_summary.sumDEPI}"
    wsd['D23'] = "☑" if structural_summary.sumCDI >= 4 else "☐"
    wsd['D23'].alignment = Alignment(horizontal='right')
    wsd['E23'] = f"CDI={structural_summary.sumCDI}"
    wsd['F23'] = "☑" if structural_summary.sumSCON >= 8 else "☐"
    wsd['F23'].alignment = Alignment(horizontal='right')
    wsd['G23'] = f"S-CON={structural_summary.sumSCON}"
    wsd['H23'] = "☑ HVI" if structural_summary.HVI_premise is True and structural_summary.sumHVI >= 4 else "☐ HVI"
    wsd['H23'].alignment = Alignment(horizontal='right')
    wsd['J23'] = "☑ OBS" if structural_summary.OBS_posi else "☐ OBS"
    wsd['J23'].alignment = Alignment(horizontal='right')

    for column_cells in wsd.columns:
        length = max(len(str(cell.value)) * 1.1 for cell in column_cells)
        wsd.column_dimensions[column_cells[0].column_letter].width = length

    import pandas as pd

    # 특수지표 시트
    wsp['A1'] = "PTI"
    wsp['A2'] = "XA%<.70 AND WDA%<.75"
    wsp['A3'] = "X-%>0.29"
    wsp['A4'] = "LVL2>2 AND FAB2>0"
    wsp['A5'] = "R<17 AND Wsum6>12 OR R>16 AND Wsum6>17*"
    wsp['A6'] = "M- > 1 OR X-% > 0.40"
    wsp['A7'] = "TOTAL"
    row = 2
    for i in range(0, 5):
        value = "✔" if structural_summary.PTI[i] == "o" else ''
        wsp.cell(row=row, column=2, value=value)
        row += 1
    wsp['B7'] = structural_summary.sumPTI

    wsp['A9'] = "DEPI"
    wsp['A10'] = "SumV>0 OR FD>2"
    wsp['A11'] = "Col-shd blends>0 OR S>2"
    wsp['A12'] = "ego sup AND Fr+rF=0 OR ego inf"
    wsp['A13'] = "Afr<0.46 OR Blends<4"
    wsp['A14'] = "SumShd>FM+m OR SumC'>2"
    wsp['A15'] = "MOR>2 OR INTELL>3"
    wsp['A16'] = "COP<2 OR ISOL>0.24"
    wsp['A17'] = "TOTAL"
    wsp['A18'] = "POSITIVE?"
    row = 10
    for i in range(0, 7):
        value = "✔" if structural_summary.DEPI[i] == "o" else ''
        wsp.cell(row=row, column=2, value=value)
        row += 1
    wsp['B17'] = structural_summary.sumDEPI
    wsp['B18'] = structural_summary.sumDEPI >= 5

    wsp['A20'] = "CDI"
    wsp['A21'] = "EA<6 OR Daj<0"
    wsp['A22'] = "COP<2 AND AG<2"
    wsp['A23'] = "WSumC<2.5 OR Afr<0.46"
    wsp['A24'] = "p > a+1 OR pure H<2"
    wsp['A25'] = "SumT>1 OR ISOL>0.24 OR Fd>0"
    wsp['A26'] = "TOTAL"
    wsp['A27'] = "POSITIVE?"
    row = 21
    for i in range(0, 5):
        value = "✔" if structural_summary.CDI[i] == "o" else ''
        wsp.cell(row=row, column=2, value=value)
        row += 1
    wsp['B26'] = structural_summary.sumCDI
    wsp['B27'] = structural_summary.sumCDI >= 4

    wsp['D1'] = "S-CON"
    wsp['D2'] = "SumV+FD>2"
    wsp['D3'] = "col-shd blends>0"
    wsp['D4'] = "ego <0.31 ou >0.44"
    wsp['D5'] = "mor>3"
    wsp['D6'] = "Zd>3.5 ou <-3.5"
    wsp['D7'] = "es>EA"
    wsp['D8'] = "CF+C>FC"
    wsp['D9'] = "X+%<0.70"
    wsp['D10'] = "S>3"
    wsp['D11'] = "P<3 OU P>8"
    wsp['D12'] = "PURE H<2"
    wsp['D13'] = "R<17"
    wsp['D14'] = 'TOTAL'
    wsp['D15'] = 'POSITIVE?'
    row = 2
    for i in range(0, 12):
        value = "✔" if structural_summary.SCON[i] == "o" else ''
        wsp.cell(row=row, column=5, value=value)
        row += 1
    wsp['E14'] = structural_summary.sumSCON
    wsp['E15'] = structural_summary.sumSCON >= 8

    wsp['D17'] = 'HVI'
    wsp['D18'] = 'SumT = 0'
    wsp['D19'] = "Zf>12"
    wsp['D20'] = "Zd>3.5"
    wsp['D21'] = "S>3"
    wsp['D22'] = "H+(H)+Hd+(Hd)>6"
    wsp['D23'] = "(H)+(A)+(Hd)+(Ad)>3"
    wsp['D24'] = "H+A : 4:1"
    wsp['D25'] = "Cg>3"
    wsp['D26'] = 'TOTAL'
    wsp['D26'] = 'TOTAL'
    wsp['D27'] = 'POSITIVE?'
    wsp['E18'] = structural_summary.HVI_premise
    row = 19
    for i in range(0, 7):
        value = "✔" if structural_summary.HVI[i] == "o" else ''
        wsp.cell(row=row, column=5, value=value)
        row += 1
    wsp['E26'] = structural_summary.sumHVI
    wsp['E27'] = structural_summary.sumHVI >= 4 and structural_summary.HVI_premise

    wsp['A29'] = 'OBS'
    wsp['A30'] = 1
    wsp['A31'] = 2
    wsp['A32'] = 3
    wsp['A33'] = 4
    wsp['A34'] = 5
    wsp['B30'] = "Dd>3"
    wsp['B31'] = "Zf>12"
    wsp['B32'] = "Zd>3.0"
    wsp['B33'] = "P>7"
    wsp['B34'] = "FQ+>1"
    wsp['D30'] = "1-5 are true"
    wsp['D31'] = "FQ+>3 AND 2 items 1-4"
    wsp['D32'] = "X+%>0,89 et 3 items"
    wsp['D33'] = "FQ+>3 et X+%>0,89"
    wsp['D34'] = 'POSITIVE?'
    row = 30
    for i in range(0, 5):
        value = "✔" if structural_summary.OBS[i] == "o" else ''
        wsp.cell(row=row, column=3, value=value)
        row += 1
    row = 30
    for i in range(5, 9):
        value = "✔" if structural_summary.OBS[i] == "o" else ''
        wsp.cell(row=row, column=5, value=value)
        row += 1
    wsp['E34'] = structural_summary.OBS_posi

    # 열 너비 조정
    for column_cells in wsp.columns:
        length = max(len(str(cell.value)) * 1.1 for cell in column_cells)
        wsp.column_dimensions[column_cells[0].column_letter].width = length

    response_codes = ResponseCode.objects.filter(client=client_id)
    response_code_data = []
    for rc in response_codes:
        response_code_data.append({
            'Card': rc.card,
            'N': rc.response_num,
            'time': rc.time,
            'response': rc.response,
            'V': rc.rotation,
            'inquiry': rc.inquiry,
            'Location': rc.location,
            'loc_num': rc.loc_num,
            'Dev Qual': rc.dev_qual,
            'determinants': rc.determinants,
            'Form Quality': rc.form_qual,
            '(2)': rc.pair,
            'Content': rc.content,
            'P': rc.popular,
            'Z': rc.Z,
            'special': rc.special,
            'comment': rc.comment
        })

    import pandas as pd
    response_code_df = pd.DataFrame(response_code_data)
    ws2 = wb.create_sheet(title="raw data")
    data_values = response_code_df.values.tolist()

    for col_num, column_title in enumerate(response_code_df.columns, 1):
        cell = ws2.cell(row=1, column=col_num, value=column_title)
        cell.font = Font(bold=True)

    for row_data in data_values:
        ws2.append(row_data)

    response = HttpResponse(
        save_virtual_workbook(wb),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=structural_summary.xlsx'
    return response


def edit_responses(request, client_id):
    response_codes = ResponseCode.objects.filter(client_id=client_id)
    ResponseCodeFormSet = modelformset_factory(ResponseCode, form=ResponseCodeForm, extra=40, max_num=70)
    if request.method == 'POST':
        formset = ResponseCodeFormSet(request.POST, queryset=response_codes)
        if formset.is_valid():
            for form in formset:
                if form.cleaned_data.get('card') and form.cleaned_data.get('response'):
                    form.instance.client_id = client_id
                    form.save()
            structural_summary, _ = StructuralSummary.objects.get_or_create(client_id=client_id)
            structural_summary.save()
            return redirect('client_list')
    else:
        formset = ResponseCodeFormSet(queryset=response_codes)

    return render(request, 'edit_responses.html', {'formset': formset, 'client_id': client_id})

@group_min_required('advanced')
def export_structural_summary_xlsx_advanced(request, client_id):

    try:
        client = Client.objects.get(id=client_id)
        if client.tester != request.user:
            return HttpResponse("액세스 거부: 해당 정보를 볼 수 있는 권한이 없습니다.", status=403)

        response_codes = ResponseCode.objects.filter(client_id=client_id)

        roman_to_num = {'I':'1','II':'2','III':'3','IV':'4','V':'5','VI':'6','VII':'7','VIII':'8','IX':'9','X':'10'}
        num_to_roman = {v:k for k,v in roman_to_num.items()}
        def normalize_card(val):
            s = str(val).strip().upper()
            return roman_to_num.get(s, s)

        numbers_found = { normalize_card(rc.card) for rc in response_codes if rc.card }
        required = {str(i) for i in range(1, 11)}
        missing = sorted(required - numbers_found, key=int)
        if missing:
            missing_roman = [num_to_roman.get(n, n) for n in missing]
            return HttpResponse("다음 카드의 반응이 없습니다: " + ", ".join(missing_roman))

        structural_summary, created = StructuralSummary.objects.get_or_create(client_id=client_id)
        if not created:
            structural_summary.calculate_values()

    except Client.DoesNotExist:
        logging.error("해당 ID의 클라이언트를 찾을 수 없음")
        return HttpResponseNotFound("클라이언트 정보를 찾을 수 없습니다.")
    except Exception as e:
        logging.error(f"예기치 못한 오류 발생: {e}")
        error_message = f"예기치 못한 오류 발생: {type(e).__name__}, {str(e)}"
        return JsonResponse({'error': error_message}, status=500)

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    ws = wb.create_sheet(title='상단부')
    wsd = wb.create_sheet(title='하단부')
    wsp = wb.create_sheet(title='특수지표')
    ws.merge_cells('A4:B4')
    ws.merge_cells('A14:B14')
    ws.merge_cells('A20:D20')
    ws.merge_cells('F4:H4')
    ws.merge_cells('G5:H5')
    ws.merge_cells('J4:K4')
    ws.merge_cells('M4:N4')
    ws.merge_cells('M16:P16')
    wsd.merge_cells('A3:F3')
    wsd.merge_cells('H3:I3')
    wsd.merge_cells('K3:N3')
    wsd.merge_cells('K5:L5')
    wsd.merge_cells('K6:L6')
    wsd.merge_cells('K7:L7')
    wsd.merge_cells('K8:L8')
    wsd.merge_cells('K9:L9')
    wsd.merge_cells('K10:L10')
    wsd.merge_cells('K11:L11')
    wsd.merge_cells('K12:L12')
    wsd.merge_cells('A14:D14')
    wsd.merge_cells('F14:G14')
    wsd.merge_cells('I14:J14')
    wsd.merge_cells('L14:M14')

    bckground_cells = ['A4', 'A14', 'A20', 'F4', 'F5', 'G5', 'J4', 'M4', 'M16']
    for cell in bckground_cells:
        ws[cell].fill = PatternFill(start_color="7B68EE", end_color="7B68EE", fill_type='solid')
    bckground_cells2 = ['A3', 'H3', 'K3', 'A14', 'F14', 'I14', 'L14']
    for cell in bckground_cells2:
        wsd[cell].fill = PatternFill(start_color="7B68EE", end_color="7B68EE", fill_type='solid')

    ws.sheet_view.showGridLines = False
    wsd.sheet_view.showGridLines = False

    BORDER_LIST = ['A5:B12', 'A4:B4', 'A14:B14', 'A15:B18', 'A20:D20', 'A21:D21', 'A22:D26', 'F4:H4', 'F5:F5', 'G5:H5',
                   'F6:F29', 'G6:H29', 'J4:K4', 'J5:K31', 'M4:N4', 'M5:N14', 'M16:P16', 'M17:P17', 'M18:P25', 'M26:P30']
    BORDER_LIST2 = ['A3:F3', 'A4:F4', 'A5:F7', 'A8:F9', 'H3:I3', 'H4:I10', 'K3:N3', 'K4:N12', 'A14:D14', 'A15:D19',
                    'F14:G14', 'F15:G21', 'I14:J14', 'I15:J21', 'L14:M14', 'L15:M21']

    def set_border(worksheet, cell_range):
        rows = worksheet[cell_range]
        side = Side(border_style='thin', color="FF000000")

        rows = list(rows)
        max_y = len(rows) - 1
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1
            for pos_x, c in enumerate(cells):
                border = Border(
                    left=c.border.left,
                    right=c.border.right,
                    top=c.border.top,
                    bottom=c.border.bottom
                )
                if pos_x == 0:
                    border.left = side
                if pos_x == max_x:
                    border.right = side
                if pos_y == 0:
                    border.top = side
                if pos_y == max_y:
                    border.bottom = side
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    c.border = border

    for pos in BORDER_LIST:
        set_border(ws, pos)
    for pos in BORDER_LIST2:
        set_border(wsd, pos)

    
    ws['A4'] = 'Location Features'
    ws['A5'] = 'Zf'
    ws['A6'] = 'Zsum'
    ws['A7'] = 'Zest'
    ws['A9'] = 'W'
    ws['A10'] = 'D'
    ws['A11'] = 'Dd'
    ws['A12'] = 'S'

    ws['A14'] = 'Developmental Quality'
    ws['A15'] = '+'
    ws['A16'] = 'o'
    ws['A17'] = 'v/+'
    ws['A18'] = 'v'

    ws['A20'] = 'Form Quality'
    ws['B21'] = 'FQx'
    ws['C21'] = 'MQual'
    ws['D21'] = 'W+D'
    ws['A22'] = '+'
    ws['A23'] = 'o'
    ws['A24'] = 'u'
    ws['A25'] = '-'
    ws['A26'] = 'none'

    ws['F4'] = 'Determinants'
    ws['F5'] = 'Blends'
    ws['G5'] = 'Single'
    fields = [
        'M', 'FM', 'm', 'FC', 'CF', 'C', 'Cn', "FC'", "C'F", "C'",
        'FT', 'TF', 'T', 'FV', 'VF', 'V', 'FY', 'YF', 'Y', 'Fr', 'rF', 'FD', 'F', '(2)'
    ]
    real_field = [
        'M', 'FM', 'm_l', 'FC', 'CF', 'C', 'Cn', 'FCa', 'CaF', 'Ca',
        'FT', 'TF', 'T', 'FV', 'VF', 'V', 'FY', 'YF', 'Y', 'Fr', 'rF', 'FD', 'F', 'pair'
    ]
    s_row = 6
    for field_name in fields:
        ws.cell(row=s_row, column=7, value=field_name)
        s_row += 1

    ws['J4'] = 'Contents'
    cont_fields = [
        'H', '(H)', 'Hd', '(Hd)', 'Hx', 'A', '(A)', 'Ad', '(Ad)', 'An',
        'Art', 'Ay', 'Bl', 'Bt', 'Cg', 'Cl', 'Ex', 'Fd', 'Fi', 'Ge', 'Hh', 'Ls',
        'Na', 'Sc', 'Sx', 'Xy', 'Id'
    ]
    cont_real_fields = [
        'H', 'H_paren', 'Hd', 'Hd_paren', 'Hx', 'A', 'A_paren', 'Ad', 'Ad_paren', 'An',
        'Art', 'Ay', 'Bl', 'Bt', 'Cg', 'Cl', 'Ex', 'Fd_l', 'Fi', 'Ge', 'Hh', 'Ls',
        'Na', 'Sc', 'Sx', 'Xy', 'Idio'
    ]
    s_row = 5
    for field_name in cont_fields:
        ws.cell(row=s_row, column=10, value=field_name)
        s_row += 1

    ws['M4'] = "approach"
    for i, label in enumerate(['I','II','III','IV','V','VI','VII','VIII','IX','X'], start=5):
        ws.cell(row=i, column=13 if i == 5 else 13, value=label)

    ws['M16'] = 'Special Scores'
    ws['N17'] = 'Lvl-1'
    ws['O17'] = 'Lvl-2'
    sp_fields = ['DV', 'INC', 'DR', 'FAB', 'ALOG', 'CON']
    sp_real_fields = ['sp_dv', 'sp_inc', 'sp_dr', 'sp_fab', 'sp_alog', 'sp_con']
    sp_real_fields2 = ['sp_dv2', 'sp_inc2', 'sp_dr2', 'sp_fab2']
    s_row = 18
    for field_name in sp_fields:
        ws.cell(row=s_row, column=13, value=field_name)
        s_row += 1
    ws['M24'] = 'Raw Sum6'
    ws['M25'] = 'Weighted Sum6'
    ws['M26'] = 'AB'
    ws['M27'] = 'AG'
    ws['M28'] = 'COP'
    ws['M29'] = 'CP'
    ws['O26'] = 'GHR'
    ws['O27'] = 'PHR'
    ws['O28'] = 'MOR'
    ws['O29'] = 'PER'
    ws['O30'] = 'PSV'

    wsd['A3'] = 'Core'
    wsd['A4'] = 'R'
    wsd['C4'] = 'L'
    wsd['A5'] = 'EB'
    wsd['A6'] = 'eb'
    wsd['C5'] = 'EA'
    wsd['C6'] = 'es'
    wsd['C7'] = 'Adj es'
    wsd['E5'] = 'EBper'
    wsd['E6'] = 'D'
    wsd['E7'] = 'Adj D'
    wsd['A8'] = 'FM'
    wsd['A9'] = 'm'
    wsd['C8'] = "SumC'"
    wsd['C9'] = 'SumV'
    wsd['E8'] = "SumT"
    wsd['E9'] = 'SumY'

    wsd['H3'] = 'Affect'
    wsd['H4'] = 'FC:CF+C'
    wsd['H5'] = "Pure C"
    wsd['H6'] = "SumC':WsumC"
    wsd['H7'] = 'Afr'
    wsd['H8'] = "S"
    wsd['H9'] = "Blends:R"
    wsd['H10'] = "CP"

    wsd['K3'] = 'Interpersonal'
    wsd['K4'] = 'COP'
    wsd['M4'] = 'AG'
    wsd['K5'] = 'GHR:PHR'
    wsd['K6'] = 'a:p'
    wsd['K7'] = 'Food'
    wsd['K8'] = 'SumT'
    wsd['K9'] = 'Human Content'
    wsd['K10'] = 'Pure H'
    wsd['K11'] = 'PER'
    wsd['K12'] = 'Isolation Index'

    wsd['A14'] = 'Ideation'
    wsd['A15'] = 'a:p'
    wsd['A16'] = 'Ma:Mp'
    wsd['A17'] = 'Intel(2AB+Art+Ay)'
    wsd['A18'] = 'MOR'
    wsd['C15'] = 'Sum6'
    wsd['C16'] = 'Lvl-2'
    wsd['C17'] = 'Wsum6'
    wsd['C18'] = 'M-'
    wsd['C19'] = 'M none'

    wsd['F14'] = 'Mediation'
    med_fields = ['XA%', 'WDA%', 'X-%', 'S-', 'P', 'X+%', 'Xu%']
    med_real_fields = ['xa_per', 'wda_per', 'x_minus_per', 's_minus', 'popular', 'x_plus_per', 'xu_per']
    s_row = 15
    for field_name in med_fields:
        wsd.cell(row=s_row, column=6, value=field_name)
        s_row += 1

    wsd['I14'] = 'Processing'
    pro_fields = ['Zf', 'W:D:Dd', 'W:M', 'Zd', 'PSV', 'DQ+', 'DQv']
    pro_real_fields = ['Zf', 'W_D_Dd', 'W_M', 'Zd', 'sp_psv', 'dev_plus', 'dev_v']
    s_row = 15
    for field_name in pro_fields:
        wsd.cell(row=s_row, column=9, value=field_name)
        s_row += 1

    wsd['L14'] = 'Self'
    self_fields = ['Ego[3r+(2)/R]', 'Fr+rF', 'SumV', 'FD', 'An+Xy', 'MOR', 'H:(H)+Hd+(Hd)']
    self_real_fields = ['ego', 'fr_rf', 'sum_V', 'fdn', 'an_xy', 'sp_mor', 'h_prop']
    s_row = 15
    for field_name in self_fields:
        wsd.cell(row=s_row, column=12, value=field_name)
        s_row += 1

    ws['B5'] = structural_summary.Zf
    ws['B6'] = structural_summary.Zsum
    ws['B6'].number_format = '0.0'
    ws['B7'] = structural_summary.Zest
    ws['B7'].number_format = '0.0'
    ws['B9'] = structural_summary.W
    ws['B10'] = structural_summary.D
    ws['B11'] = structural_summary.Dd
    ws['B12'] = structural_summary.S

    ws['B15'] = structural_summary.dev_plus
    ws['B16'] = structural_summary.dev_o
    ws['B17'] = structural_summary.dev_vplus
    ws['B18'] = structural_summary.dev_v

    ws['B22'] = structural_summary.fqx_plus
    ws['B23'] = structural_summary.fqx_o
    ws['B24'] = structural_summary.fqx_u
    ws['B25'] = structural_summary.fqx_minus
    ws['B26'] = structural_summary.fqx_none
    ws['C22'] = structural_summary.mq_plus
    ws['C23'] = structural_summary.mq_o
    ws['C24'] = structural_summary.mq_u
    ws['C25'] = structural_summary.mq_minus
    ws['C26'] = structural_summary.mq_none
    ws['D22'] = structural_summary.wd_plus
    ws['D23'] = structural_summary.wd_o
    ws['D24'] = structural_summary.wd_u
    ws['D25'] = structural_summary.wd_minus
    ws['D26'] = structural_summary.wd_none

    blends = structural_summary.blends.split(',')
    start_row = 6
    start_column = 6
    for blend in blends:
        ws.cell(row=start_row, column=start_column, value=blend)
        start_row += 1

    row = 6
    for field_name in real_field:
        field_value = getattr(structural_summary, field_name)
        ws.cell(row=row, column=8, value=field_value)
        ws.cell(row=row, column=8, value=field_value).number_format = "#"
        row += 1

    row = 5
    for field_name in cont_real_fields:
        field_value = getattr(structural_summary, field_name)
        ws.cell(row=row, column=11, value=field_value)
        ws.cell(row=row, column=11, value=field_value).number_format = "#"
        row += 1

    ws['N5'] = structural_summary.app_I
    ws['N6'] = structural_summary.app_II
    ws['N7'] = structural_summary.app_III
    ws['N8'] = structural_summary.app_IV
    ws['N9'] = structural_summary.app_V
    ws['N10'] = structural_summary.app_VI
    ws['N11'] = structural_summary.app_VII
    ws['N12'] = structural_summary.app_VIII
    ws['N13'] = structural_summary.app_IX
    ws['N14'] = structural_summary.app_X

    lv1_row = 18
    for field_name in sp_real_fields:
        field_value = getattr(structural_summary, field_name)
        ws.cell(row=lv1_row, column=14, value=field_value)
        lv1_row += 1
    lv2_row = 18
    for field_name in sp_real_fields2:
        field_value = getattr(structural_summary, field_name)
        ws.cell(row=lv2_row, column=15, value=field_value)
        lv2_row += 1
    ws['N24'] = structural_summary.sum6
    ws['N25'] = structural_summary.wsum6
    ws['N26'] = structural_summary.sp_ab
    ws['N27'] = structural_summary.sp_ag
    ws['N28'] = structural_summary.sp_cop
    ws['N29'] = structural_summary.sp_cp
    ws['P26'] = structural_summary.sp_ghr
    ws['P27'] = structural_summary.sp_phr
    ws['P28'] = structural_summary.sp_mor
    ws['P29'] = structural_summary.sp_per
    ws['P30'] = structural_summary.sp_psv

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) * 1.1 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length

    wsd['B4'] = structural_summary.R
    wsd['D4'] = structural_summary.L
    wsd['B5'] = structural_summary.ErleBnistypus
    wsd['B6'] = structural_summary.eb
    wsd['D5'] = structural_summary.EA
    wsd['D6'] = structural_summary.es
    wsd['D7'] = structural_summary.adj_es
    wsd['F5'] = 'NA' if structural_summary.EBper == 0 else structural_summary.EBper
    wsd['F6'] = structural_summary.D_score
    wsd['F7'] = structural_summary.adj_D
    wsd['B8'] = structural_summary.sum_FM
    wsd['B9'] = structural_summary.sum_m
    wsd['D8'] = structural_summary.sum_Ca
    wsd['D9'] = structural_summary.sum_V
    wsd['F8'] = structural_summary.sum_T
    wsd['F9'] = structural_summary.sum_Y

    wsd['I4'] = structural_summary.f_c_prop
    wsd['I4'].alignment = Alignment(horizontal='right')
    wsd['I5'] = structural_summary.pure_c
    wsd['I6'] = structural_summary.ca_c_prop
    wsd['I6'].alignment = Alignment(horizontal='right')
    wsd['I7'] = structural_summary.afr
    wsd['I7'].number_format = "0.##;-0.##;0"
    wsd['I8'] = structural_summary.S
    wsd['I9'] = structural_summary.blends_r
    wsd['I9'].alignment = Alignment(horizontal='right')
    wsd['I10'] = structural_summary.sp_cp

    wsd['L4'] = structural_summary.sp_cop
    wsd['N4'] = structural_summary.sp_ag
    wsd['M5'] = structural_summary.GHR_PHR
    wsd['M5'].alignment = Alignment(horizontal='right')
    wsd['M6'] = structural_summary.a_p
    wsd['M6'].alignment = Alignment(horizontal='right')
    wsd['M7'] = structural_summary.Fd_l
    wsd['M8'] = structural_summary.sum_T
    wsd['M9'] = structural_summary.human_cont
    wsd['M10'] = structural_summary.H
    wsd['M11'] = structural_summary.sp_per
    wsd['M12'] = structural_summary.Isol

    wsd['B15'] = structural_summary.a_p
    wsd['B15'].alignment = Alignment(horizontal='right')
    wsd['B16'] = structural_summary.Ma_Mp
    wsd['B16'].alignment = Alignment(horizontal='right')
    wsd['B17'] = structural_summary.intel
    wsd['B18'] = structural_summary.sp_mor
    wsd['D15'] = structural_summary.sum6
    wsd['D16'] = structural_summary.Lvl_2
    wsd['D17'] = structural_summary.wsum6
    wsd['D18'] = structural_summary.mq_minus
    wsd['D19'] = structural_summary.mq_none

    row = 15
    for field_name in med_real_fields:
        field_value = getattr(structural_summary, field_name)
        wsd.cell(row=row, column=7, value=field_value)
        if isinstance(field_value, int):
            number_format = "0"
        else:
            number_format = "0.00"
        wsd.cell(row=row, column=7).number_format = number_format
        row += 1

    row = 15
    for field_name in pro_real_fields:
        field_value = getattr(structural_summary, field_name)
        wsd.cell(row=row, column=10, value=field_value)
        wsd.cell(row=row, column=10).alignment = Alignment(horizontal='right')
        row += 1

    row = 15
    for field_name in self_real_fields:
        field_value = getattr(structural_summary, field_name)
        wsd.cell(row=row, column=13, value=field_value)
        row += 1

    wsd['A23'] = f"PTI={structural_summary.sumPTI}"
    wsd['B23'] = "☑" if structural_summary.sumDEPI >= 5 else "☐"
    wsd['B23'].alignment = Alignment(horizontal='right')
    wsd['C23'] = f"DEPI={structural_summary.sumDEPI}"
    wsd['D23'] = "☑" if structural_summary.sumCDI >= 4 else "☐"
    wsd['D23'].alignment = Alignment(horizontal='right')
    wsd['E23'] = f"CDI={structural_summary.sumCDI}"
    wsd['F23'] = "☑" if structural_summary.sumSCON >= 8 else "☐"
    wsd['F23'].alignment = Alignment(horizontal='right')
    wsd['G23'] = f"S-CON={structural_summary.sumSCON}"
    wsd['H23'] = "☑ HVI" if structural_summary.HVI_premise is True and structural_summary.sumHVI >= 4 else "☐ HVI"
    wsd['H23'].alignment = Alignment(horizontal='right')
    wsd['J23'] = "☑ OBS" if structural_summary.OBS_posi else "☐ OBS"
    wsd['J23'].alignment = Alignment(horizontal='right')

    for column_cells in wsd.columns:
        length = max(len(str(cell.value)) * 1.1 for cell in column_cells)
        wsd.column_dimensions[column_cells[0].column_letter].width = length

    import pandas as pd

    # 특수지표 시트
    wsp['A1'] = "PTI"
    wsp['A2'] = "XA%<.70 AND WDA%<.75"
    wsp['A3'] = "X-%>0.29"
    wsp['A4'] = "LVL2>2 AND FAB2>0"
    wsp['A5'] = "R<17 AND Wsum6>12 OR R>16 AND Wsum6>17*"
    wsp['A6'] = "M- > 1 OR X-% > 0.40"
    wsp['A7'] = "TOTAL"
    row = 2
    for i in range(0, 5):
        value = "✔" if structural_summary.PTI[i] == "o" else ''
        wsp.cell(row=row, column=2, value=value)
        row += 1
    wsp['B7'] = structural_summary.sumPTI

    wsp['A9'] = "DEPI"
    wsp['A10'] = "SumV>0 OR FD>2"
    wsp['A11'] = "Col-shd blends>0 OR S>2"
    wsp['A12'] = "ego sup AND Fr+rF=0 OR ego inf"
    wsp['A13'] = "Afr<0.46 OR Blends<4"
    wsp['A14'] = "SumShd>FM+m OR SumC'>2"
    wsp['A15'] = "MOR>2 OR INTELL>3"
    wsp['A16'] = "COP<2 OR ISOL>0.24"
    wsp['A17'] = "TOTAL"
    wsp['A18'] = "POSITIVE?"
    row = 10
    for i in range(0, 7):
        value = "✔" if structural_summary.DEPI[i] == "o" else ''
        wsp.cell(row=row, column=2, value=value)
        row += 1
    wsp['B17'] = structural_summary.sumDEPI
    wsp['B18'] = structural_summary.sumDEPI >= 5

    wsp['A20'] = "CDI"
    wsp['A21'] = "EA<6 OR Daj<0"
    wsp['A22'] = "COP<2 AND AG<2"
    wsp['A23'] = "WSumC<2.5 OR Afr<0.46"
    wsp['A24'] = "p > a+1 OR pure H<2"
    wsp['A25'] = "SumT>1 OR ISOL>0.24 OR Fd>0"
    wsp['A26'] = "TOTAL"
    wsp['A27'] = "POSITIVE?"
    row = 21
    for i in range(0, 5):
        value = "✔" if structural_summary.CDI[i] == "o" else ''
        wsp.cell(row=row, column=2, value=value)
        row += 1
    wsp['B26'] = structural_summary.sumCDI
    wsp['B27'] = structural_summary.sumCDI >= 4

    wsp['D1'] = "S-CON"
    wsp['D2'] = "SumV+FD>2"
    wsp['D3'] = "col-shd blends>0"
    wsp['D4'] = "ego <0.31 ou >0.44"
    wsp['D5'] = "mor>3"
    wsp['D6'] = "Zd>3.5 ou <-3.5"
    wsp['D7'] = "es>EA"
    wsp['D8'] = "CF+C>FC"
    wsp['D9'] = "X+%<0.70"
    wsp['D10'] = "S>3"
    wsp['D11'] = "P<3 OU P>8"
    wsp['D12'] = "PURE H<2"
    wsp['D13'] = "R<17"
    wsp['D14'] = 'TOTAL'
    wsp['D15'] = 'POSITIVE?'
    row = 2
    for i in range(0, 12):
        value = "✔" if structural_summary.SCON[i] == "o" else ''
        wsp.cell(row=row, column=5, value=value)
        row += 1
    wsp['E14'] = structural_summary.sumSCON
    wsp['E15'] = structural_summary.sumSCON >= 8

    wsp['D17'] = 'HVI'
    wsp['D18'] = 'SumT = 0'
    wsp['D19'] = "Zf>12"
    wsp['D20'] = "Zd>3.5"
    wsp['D21'] = "S>3"
    wsp['D22'] = "H+(H)+Hd+(Hd)>6"
    wsp['D23'] = "(H)+(A)+(Hd)+(Ad)>3"
    wsp['D24'] = "H+A : 4:1"
    wsp['D25'] = "Cg>3"
    wsp['D26'] = 'TOTAL'
    wsp['D26'] = 'TOTAL'
    wsp['D27'] = 'POSITIVE?'
    wsp['E18'] = structural_summary.HVI_premise
    row = 19
    for i in range(0, 7):
        value = "✔" if structural_summary.HVI[i] == "o" else ''
        wsp.cell(row=row, column=5, value=value)
        row += 1
    wsp['E26'] = structural_summary.sumHVI
    wsp['E27'] = structural_summary.sumHVI >= 4 and structural_summary.HVI_premise

    wsp['A29'] = 'OBS'
    wsp['A30'] = 1
    wsp['A31'] = 2
    wsp['A32'] = 3
    wsp['A33'] = 4
    wsp['A34'] = 5
    wsp['B30'] = "Dd>3"
    wsp['B31'] = "Zf>12"
    wsp['B32'] = "Zd>3.0"
    wsp['B33'] = "P>7"
    wsp['B34'] = "FQ+>1"
    wsp['D30'] = "1-5 are true"
    wsp['D31'] = "FQ+>3 AND 2 items 1-4"
    wsp['D32'] = "X+%>0,89 et 3 items"
    wsp['D33'] = "FQ+>3 et X+%>0,89"
    wsp['D34'] = 'POSITIVE?'
    row = 30
    for i in range(0, 5):
        value = "✔" if structural_summary.OBS[i] == "o" else ''
        wsp.cell(row=row, column=3, value=value)
        row += 1
    row = 30
    for i in range(5, 9):
        value = "✔" if structural_summary.OBS[i] == "o" else ''
        wsp.cell(row=row, column=5, value=value)
        row += 1
    wsp['E34'] = structural_summary.OBS_posi

    # 열 너비 조정
    for column_cells in wsp.columns:
        length = max(len(str(cell.value)) * 1.1 for cell in column_cells)
        wsp.column_dimensions[column_cells[0].column_letter].width = length

    response_codes = ResponseCode.objects.filter(client=client_id)
    response_code_data = []
    for rc in response_codes:
        response_code_data.append({
            'Card': rc.card,
            'N': rc.response_num,
            'time': rc.time,
            'response': rc.response,
            'V': rc.rotation,
            'inquiry': rc.inquiry,
            'Location': rc.location,
            'loc_num': rc.loc_num,
            'Dev Qual': rc.dev_qual,
            'determinants': rc.determinants,
            'Form Quality': rc.form_qual,
            '(2)': rc.pair,
            'Content': rc.content,
            'P': rc.popular,
            'Z': rc.Z,
            'special': rc.special,
            'comment': rc.comment
        })

    import pandas as pd
    response_code_df = pd.DataFrame(response_code_data)
    ws2 = wb.create_sheet(title="raw data")
    data_values = response_code_df.values.tolist()

    for col_num, column_title in enumerate(response_code_df.columns, 1):
        cell = ws2.cell(row=1, column=col_num, value=column_title)
        cell.font = Font(bold=True)

    for row_data in data_values:
        ws2.append(row_data)
    response = HttpResponse(
        save_virtual_workbook(wb),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=structural_summary_advanced.xlsx'
    return response

@group_min_required('intermediate')
def export_structural_summary_xlsx_auto(request, client_id):
    if getattr(request.user, 'group', '') == 'advanced':
        # 고급 함수가 없는 경우
        try:
            return export_structural_summary_xlsx_advanced(request, client_id)
        except NameError:
            # 고급 함수가 아직 준비되지 않았다면 기본으로
            return export_structural_summary_xlsx(request, client_id)
    else:
        # 중급
        return export_structural_summary_xlsx(request, client_id)