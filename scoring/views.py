from .models import SearchReference, CardImages, PopularResponse, StructuralSummary
from .filters import SearchReferenceFilter, CardImagesFilter, PResponseFilter
from .forms import ClientForm
from django.http import HttpResponseNotFound
import logging
from django.http import JsonResponse
from django.forms import formset_factory
from django.shortcuts import redirect
from .forms import ResponseCodeForm

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
from .models import Client, ResponseCode

from django.shortcuts import render
from functools import wraps
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden

from django.forms import modelformset_factory


def group_required(group_name):
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return login_required(view_func)(request, *args, **kwargs)

            if request.user.group == group_name:
                return view_func(request, *args, **kwargs)
            else:
                return HttpResponseForbidden("중급 이상 이수자만 접속 가능한 페이지입니다.")

        return _wrapped_view

    return decorator


@group_required('intermediate')
def search(request):
    client_id = request.GET.get('client_id')
    # 클라이언트 ID를 받아오지 못했을 경우 정보 입력창으로 이동
    if not client_id:
        return redirect('client_info')
    try:
        client = Client.objects.get(id=client_id)
        # 클라이언트의 tester가 현재 로그인한 사용자가 아닌 경우 액세스 거부
        if client.tester != request.user:
            return HttpResponse("액세스 거부: 작성 권한이 없습니다.")
    except Client.DoesNotExist:
        # 클라이언트에 대한 정보가 없을 경우 처리
        logging.error("해당 ID의 클라이언트를 찾을 수 없음")
        return HttpResponseNotFound("클라이언트 정보를 찾을 수 없습니다.")
    
    extra_forms = 40
    ResponseCodeFormSet = formset_factory(ResponseCodeForm, extra=extra_forms, max_num=70)
    if request.method == 'POST':
        if 'additems' in request.POST and request.POST['additems'] == 'true':
            formset_dictionary_copy = request.POST.copy()
            formset_dictionary_copy['form-TOTAL_FORMS'] = int(formset_dictionary_copy['form-TOTAL_FORMS']) + 1
            formset = ResponseCodeFormSet(formset_dictionary_copy)
        else:
            formset = ResponseCodeFormSet(request.POST)
            if formset.is_valid():
                for form in formset:
                    if form.cleaned_data.get('card') and form.cleaned_data.get('response'):
                        form.instance.client_id = client_id
                        form.save()
                return redirect('client_list')
            else:
                for form in formset:
                    for field in form:
                        print(field.errors)
    else:
        formset = ResponseCodeFormSet()
    card_image_list = CardImages.objects.all()
    reference_list = SearchReference.objects.all()
    search_filter = SearchReferenceFilter(request.GET, queryset=reference_list)
    card_image_filter = CardImagesFilter(request.GET, queryset=card_image_list)
    Presponse_list = PopularResponse.objects.all()
    Presponse_filter = PResponseFilter(request.GET, queryset=Presponse_list)

    if request.GET.get('Card'):
        card_number = request.GET['Card']
        card_image_filter = CardImagesFilter(request.GET, queryset=CardImages.objects.filter(card_number=card_number))
        Presponse_filter = PResponseFilter(request.GET,
                                           queryset=PopularResponse.objects.filter(card_number=card_number))

    return render(request, 'base.html', {'formset': formset, 'filter': search_filter, 'image_filter': card_image_filter,
                                         'p_response_filter': Presponse_filter})


@group_required('intermediate')
def update_response_codes(request, client_id):
    try:
        client = Client.objects.get(id=client_id)
        # 클라이언트의 tester가 현재 로그인한 사용자가 아닌 경우 액세스 거부
        if client.tester != request.user:
            return HttpResponse("액세스 거부: 작성 권한이 없습니다.")
    except Client.DoesNotExist:
        # 클라이언트에 대한 정보가 없을 경우 처리
        logging.error("해당 ID의 클라이언트를 찾을 수 없음")
        return HttpResponseNotFound("클라이언트 정보를 찾을 수 없습니다.")
    response_codes = ResponseCode.objects.filter(client_id=client_id)
    extra_forms = 40
    ResponseCodeFormSet = modelformset_factory(ResponseCode, form=ResponseCodeForm, extra=extra_forms, max_num=70)
    formset = ResponseCodeFormSet(queryset=response_codes)

    if request.method == 'POST':
        if 'additems' in request.POST and request.POST['additems'] == 'true':
            formset_dictionary_copy = request.POST.copy()
            formset_dictionary_copy['form-TOTAL_FORMS'] = int(formset_dictionary_copy['form-TOTAL_FORMS']) + 1
            formset = ResponseCodeFormSet(formset_dictionary_copy)
        else:
            formset = ResponseCodeFormSet(request.POST, queryset=response_codes)
            if formset.is_valid():
                for form in formset:
                    if form.cleaned_data.get('card') and form.cleaned_data.get('response'):
                        form.instance.client_id = client_id
                        form.save()
                return redirect('client_list')
            else:
                for form in formset:
                    for field in form:
                        print(field.errors)

    card_image_list = CardImages.objects.all()
    reference_list = SearchReference.objects.all()
    search_filter = SearchReferenceFilter(request.GET, queryset=reference_list)
    card_image_filter = CardImagesFilter(request.GET, queryset=card_image_list)
    Presponse_list = PopularResponse.objects.all()
    Presponse_filter = PResponseFilter(request.GET, queryset=Presponse_list)

    if request.GET.get('Card'):
        card_number = request.GET['Card']
        card_image_filter = CardImagesFilter(request.GET, queryset=CardImages.objects.filter(card_number=card_number))
        Presponse_filter = PResponseFilter(request.GET,
                                           queryset=PopularResponse.objects.filter(card_number=card_number))

    return render(request, 'update_response_codes.html',
                  {'formset': formset, 'filter': search_filter, 'image_filter': card_image_filter,
                   'p_response_filter': Presponse_filter})


def search_results(request):
    card_image_list = CardImages.objects.all()
    reference_list = SearchReference.objects.all()
    search_filter = SearchReferenceFilter(request.GET, queryset=reference_list)
    card_image_filter = CardImagesFilter(request.GET, queryset=card_image_list)
    Presponse_list = PopularResponse.objects.all()
    Presponse_filter = PResponseFilter(request.GET, queryset=Presponse_list)

    if request.GET.get('Card'):
        card_number = request.GET['Card']
        card_image_filter = CardImagesFilter(request.GET, queryset=CardImages.objects.filter(card_number=card_number))
        Presponse_filter = PResponseFilter(request.GET,
                                           queryset=PopularResponse.objects.filter(card_number=card_number))

    return render(request, 'search_results.html',
                  {'filter': search_filter, 'image_filter': card_image_filter, 'p_response_filter': Presponse_filter})


@group_required('intermediate')
def add_client(request):
    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            client_instance = form.save(commit=False)
            client_instance.tester = request.user  # 현재 로그인된 유저를 user 필드에 할당
            client_instance.save()  # 클라이언트 정보 저장
            return redirect('/search/?client_id=' + str(client_instance.id))
        else:
            print(form.errors)
    else:
        form = ClientForm(initial={
            'name': request.GET.get('name', ''),
            'gender': request.GET.get('gender', ''),
            'birthdate': request.GET.get('birthdate', ''),
            'testDate': request.GET.get('testDate', ''),
            'notes': request.GET.get('notes', ''),
        })
    return render(request, 'add_client.html', {'form': form})


@group_required('intermediate')
def client_list(request):
    user = request.user
    clients = Client.objects.filter(tester=user)
    return render(request, 'client_list.html', {'clients': clients})


@group_required('intermediate')
def client_detail(request, client_id):
    user = request.user
    try:
        client_obj = Client.objects.get(id=client_id, tester=user)
    except Client.DoesNotExist:
        return HttpResponse("액세스 거부: 해당 정보를 볼 수 있는 권한이 없습니다.")

    response_codes = ResponseCode.objects.filter(client=client_obj)
    return render(request, 'client_detail.html', {'client': client_obj, 'response_codes': response_codes})


@group_required('intermediate')
def export_structural_summary_xlsx(request, client_id):
    try:
        client = Client.objects.get(id=client_id)
        # 클라이언트의 tester가 현재 로그인한 사용자가 아닌 경우 액세스 거부
        if client.tester != request.user:
            return HttpResponse("액세스 거부: 해당 정보를 볼 수 있는 권한이 없습니다.")

        # ResponseCode 모델에서 해당 client_id와 관련된 데이터 가져오기
        response_codes = ResponseCode.objects.filter(client_id=client_id)

        # 1부터 10까지의 숫자가 적어도 한 번씩 나왔는지 검사
        numbers_found = set()
        roman_dict = {
            'I': '1', 'II': '2', 'III': '3', 'IV': '4', 'V': '5',
            'VI': '6', 'VII': '7', 'VIII': '8', 'IX': '9', 'X': '10'
        }
        for response_code in response_codes:
            card_value = response_code.card
            card_val = roman_dict.get(card_value, card_value)
            numbers_found.add(card_val)

        # 1부터 10까지의 숫자가 모두 나왔는지 확인
        if len(numbers_found) < 10:
            return HttpResponse("한 카드에 적어도 하나의 반응을 입력해야 합니다.")

        structural_summary, created = StructuralSummary.objects.get_or_create(client_id=client_id)
        if not created:
            structural_summary.calculate_values()  # Structural summary가 이미 있는 경우 값 업데이트
    except Client.DoesNotExist:
        # 클라이언트에 대한 정보가 없을 경우 처리
        logging.error("해당 ID의 클라이언트를 찾을 수 없음")
        return HttpResponseNotFound("클라이언트 정보를 찾을 수 없습니다.")
    except Exception as e:
        # 기타 예외 처리
        logging.error(f"예기치 못한 오류 발생: {e}")
        error_message = f"예기치 못한 오류 발생: {type(e).__name__}, {str(e)}"
        return JsonResponse({'error': error_message}, status=500)
        #return HttpResponse("예기치 못한 오류가 발생했습니다.")

    # 새로운 워크북 및 worksheet 생성
    wb = Workbook()
    # 빈 시트 제거
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    ws = wb.create_sheet(title='상단부')
    wsd = wb.create_sheet(title='하단부')
    wsp = wb.create_sheet(title='특수지표')
    ws.merge_cells('A4:B4')
    ws.merge_cells('A14:B14')
    ws.merge_cells('A20:D20')
    ws.merge_cells('F4:H4')
    ws.merge_cells('G5:H5')
    ws.merge_cells('J4:K4')
    ws.merge_cells('M4:N4')
    ws.merge_cells('M16:P16')
    wsd.merge_cells('A3:F3')
    wsd.merge_cells('H3:I3')
    wsd.merge_cells('K3:N3')
    wsd.merge_cells('K5:L5')
    wsd.merge_cells('K6:L6')
    wsd.merge_cells('K7:L7')
    wsd.merge_cells('K8:L8')
    wsd.merge_cells('K9:L9')
    wsd.merge_cells('K10:L10')
    wsd.merge_cells('K11:L11')
    wsd.merge_cells('K12:L12')
    wsd.merge_cells('A14:D14')
    wsd.merge_cells('F14:G14')
    wsd.merge_cells('I14:J14')
    wsd.merge_cells('L14:M14')

    # 배경색 넣기
    bckground_cells = ['A4', 'A14', 'A20', 'F4', 'F5', 'G5', 'J4', 'M4', 'M16']
    for cell in bckground_cells:
        ws[cell].fill = PatternFill(start_color="7B68EE", end_color="7B68EE", fill_type='solid')
    bckground_cells2 = ['A3', 'H3', 'K3', 'A14', 'F14', 'I14', 'L14']
    for cell in bckground_cells2:
        wsd[cell].fill = PatternFill(start_color="7B68EE", end_color="7B68EE", fill_type='solid')

    # 눈금 없애기
    ws.sheet_view.showGridLines = False
    wsd.sheet_view.showGridLines = False

    # 테두리 설정
    BORDER_LIST = ['A5:B12', 'A4:B4', 'A14:B14', 'A15:B18', 'A20:D20', 'A21:D21', 'A22:D26', 'F4:H4', 'F5:F5', 'G5:H5',
                   'F6:F29', 'G6:H29', 'J4:K4', 'J5:K31', 'M4:N4', 'M5:N14', 'M16:P16', 'M17:P17', 'M18:P25', 'M26:P30']
    BORDER_LIST2 = ['A3:F3', 'A4:F4', 'A5:F7', 'A8:F9', 'H3:I3', 'H4:I10', 'K3:N3', 'K4:N12', 'A14:D14', 'A15:D19',
                    'F14:G14', 'F15:G21', 'I14:J14', 'I15:J21', 'L14:M14', 'L15:M21']

    def set_border(worksheet, cell_range):
        rows = worksheet[cell_range]
        side = Side(border_style='thin', color="FF000000")

        rows = list(rows)
        max_y = len(rows) - 1  # index of the last row
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1  # index of the last cell
            for pos_x, c in enumerate(cells):
                border = Border(
                    left=c.border.left,
                    right=c.border.right,
                    top=c.border.top,
                    bottom=c.border.bottom
                )
                if pos_x == 0:
                    border.left = side
                if pos_x == max_x:
                    border.right = side
                if pos_y == 0:
                    border.top = side
                if pos_y == max_y:
                    border.bottom = side

                # set new border only if it's one of the edge cells
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    c.border = border

    # border
    for pos in BORDER_LIST:
        set_border(ws, pos)
    for pos in BORDER_LIST2:
        set_border(wsd, pos)

    # 엑셀 열 제목 추가
    # 1. location features
    ws['A4'] = 'Location Features'
    ws['A5'] = 'Zf'
    ws['A6'] = 'Zsum'
    ws['A7'] = 'Zest'
    ws['A9'] = 'W'
    ws['A10'] = 'D'
    ws['A11'] = 'Dd'
    ws['A12'] = 'S'

    # 2. Developmental Quality
    ws['A14'] = 'Developmental Quality'
    ws['A15'] = '+'
    ws['A16'] = 'o'
    ws['A17'] = 'v/+'
    ws['A18'] = 'v'

    # 3. Form Quality
    ws['A20'] = 'Form Quality'
    ws['B21'] = 'FQx'
    ws['C21'] = 'MQual'
    ws['D21'] = 'W+D'
    ws['A22'] = '+'
    ws['A23'] = 'o'
    ws['A24'] = 'u'
    ws['A25'] = '-'
    ws['A26'] = 'none'

    # 4. Determinants
    ws['F4'] = 'Determinants'
    ws['F5'] = 'Blends'
    ws['G5'] = 'Single'
    fields = [
        'M', 'FM', 'm', 'FC', 'CF', 'C', 'Cn', "FC'", "C'F", "C'",
        'FT', 'TF', 'T', 'FV', 'VF', 'V', 'FY', 'YF', 'Y', 'Fr', 'rF', 'FD', 'F', '(2)'
    ]
    real_field = [
        'M', 'FM', 'm_l', 'FC', 'CF', 'C', 'Cn', 'FCa', 'CaF', 'Ca',
        'FT', 'TF', 'T', 'FV', 'VF', 'V', 'FY', 'YF', 'Y', 'Fr', 'rF', 'FD', 'F', 'pair'
    ]
    s_row = 6
    for field_name in fields:
        ws.cell(row=s_row, column=7, value=field_name)
        s_row += 1

    # 5. Contents
    ws['J4'] = 'Contents'
    cont_fields = [
        'H', '(H)', 'Hd', '(Hd)', 'Hx', 'A', '(A)', 'Ad', '(Ad)', 'An',
        'Art', 'Ay', 'Bl', 'Bt', 'Cg', 'Cl', 'Ex', 'Fd', 'Fi', 'Ge', 'Hh', 'Ls',
        'Na', 'Sc', 'Sx', 'Xy', 'Id'
    ]
    cont_real_fields = [
        'H', 'H_paren', 'Hd', 'Hd_paren', 'Hx', 'A', 'A_paren', 'Ad', 'Ad_paren', 'An',
        'Art', 'Ay', 'Bl', 'Bt', 'Cg', 'Cl', 'Ex', 'Fd_l', 'Fi', 'Ge', 'Hh', 'Ls',
        'Na', 'Sc', 'Sx', 'Xy', 'Idio'
    ]
    s_row = 5
    for field_name in cont_fields:
        ws.cell(row=s_row, column=10, value=field_name)
        s_row += 1

    # 6. approche
    ws['M4'] = "approach"
    ws['M5'] = 'I'
    ws['M6'] = "II"
    ws['M7'] = 'III'
    ws['M8'] = "IV"
    ws['M9'] = 'V'
    ws['M10'] = "VI"
    ws['M11'] = 'VII'
    ws['M12'] = "VIII"
    ws['M13'] = 'IX'
    ws['M14'] = "X"

    # 7. Special Scores
    ws['M16'] = 'Special Scores'
    ws['N17'] = 'Lvl-1'
    ws['O17'] = 'Lvl-2'
    sp_fields = [
        'DV', 'INC', 'DR', 'FAB', 'ALOG', 'CON'
    ]
    sp_real_fields = [
        'sp_dv', 'sp_inc', 'sp_dr', 'sp_fab', 'sp_alog', 'sp_con'
    ]
    sp_real_fields2 = [
        'sp_dv2', 'sp_inc2', 'sp_dr2', 'sp_fab2'
    ]
    s_row = 18
    for field_name in sp_fields:
        ws.cell(row=s_row, column=13, value=field_name)
        s_row += 1
    ws['M24'] = 'Raw Sum6'
    ws['M25'] = 'Weighted Sum6'
    ws['M26'] = 'AB'
    ws['M27'] = 'AG'
    ws['M28'] = 'COP'
    ws['M29'] = 'CP'
    ws['O26'] = 'GHR'
    ws['O27'] = 'PHR'
    ws['O28'] = 'MOR'
    ws['O29'] = 'PER'
    ws['O30'] = 'PSV'

    # 하단부
    # 8. core
    wsd['A3'] = 'Core'
    wsd['A4'] = 'R'
    wsd['C4'] = 'L'
    wsd['A5'] = 'EB'
    wsd['A6'] = 'eb'
    wsd['C5'] = 'EA'
    wsd['C6'] = 'es'
    wsd['C7'] = 'Adj es'
    wsd['E5'] = 'EBper'
    wsd['E6'] = 'D'
    wsd['E7'] = 'Adj D'
    wsd['A8'] = 'FM'
    wsd['A9'] = 'm'
    wsd['C8'] = "SumC'"
    wsd['C9'] = 'SumV'
    wsd['E8'] = "SumT"
    wsd['E9'] = 'SumY'

    # 9. Affect
    wsd['H3'] = 'Affect'
    wsd['H4'] = 'FC:CF+C'
    wsd['H5'] = "Pure C"
    wsd['H6'] = "SumC':WsumC"
    wsd['H7'] = 'Afr'
    wsd['H8'] = "S"
    wsd['H9'] = "Blends:R"
    wsd['H10'] = "CP"

    # 10. Interpersonal
    wsd['K3'] = 'Interpersonal'
    wsd['K4'] = 'COP'
    wsd['M4'] = 'AG'
    wsd['K5'] = 'GHR:PHR'
    wsd['K6'] = 'a:p'
    wsd['K7'] = 'Food'
    wsd['K8'] = 'SumT'
    wsd['K9'] = 'Human Content'
    wsd['K10'] = 'Pure H'
    wsd['K11'] = 'PER'
    wsd['K12'] = 'Isolation Index'

    # 11. Ideation
    wsd['A14'] = 'Ideation'
    wsd['A15'] = 'a:p'
    wsd['A16'] = 'Ma:Mp'
    wsd['A17'] = 'Intel(2AB+Art+Ay)'
    wsd['A18'] = 'MOR'
    wsd['C15'] = 'Sum6'
    wsd['C16'] = 'Lvl-2'
    wsd['C17'] = 'Wsum6'
    wsd['C18'] = 'M-'
    wsd['C19'] = 'M none'

    # 12. Mediation
    wsd['F14'] = 'Mediation'
    med_fields = [
        'XA%', 'WDA%', 'X-%', 'S-', 'P', 'X+%', 'Xu%'
    ]
    med_real_fields = [
        'xa_per', 'wda_per', 'x_minus_per', 's_minus', 'popular',
        'x_plus_per', 'xu_per'
    ]
    s_row = 15
    for field_name in med_fields:
        wsd.cell(row=s_row, column=6, value=field_name)
        s_row += 1

    # 13. Processing
    wsd['I14'] = 'Processing'
    pro_fields = [
        'Zf', 'W:D:Dd', 'W:M', 'Zd', 'PSV', 'DQ+', 'DQv'
    ]
    pro_real_fields = [
        'Zf', 'W_D_Dd', 'W_M', 'Zd', 'sp_psv', 'dev_plus', 'dev_v'
    ]
    s_row = 15
    for field_name in pro_fields:
        wsd.cell(row=s_row, column=9, value=field_name)
        s_row += 1

    # 14. Self-Perception
    wsd['L14'] = 'Self'
    self_fields = [
        'Ego[3r+(2)/R]', 'Fr+rF', 'SumV', 'FD', 'An+Xy', 'MOR', 'H:(H)+Hd+(Hd)'
    ]
    self_real_fields = [
        'ego', 'fr_rf', 'sum_V', 'fdn', 'an_xy', 'sp_mor', 'h_prop'
    ]
    s_row = 15
    for field_name in self_fields:
        wsd.cell(row=s_row, column=12, value=field_name)
        s_row += 1

    # 데이터 추가
    # 1. location features
    ws['B5'] = structural_summary.Zf
    ws['B6'] = structural_summary.Zsum
    ws['B6'].number_format = '0.0'
    ws['B7'] = structural_summary.Zest
    ws['B7'].number_format = '0.0'
    ws['B9'] = structural_summary.W
    ws['B10'] = structural_summary.D
    ws['B11'] = structural_summary.Dd
    ws['B12'] = structural_summary.S

    # 2. Developmental Quality
    ws['B15'] = structural_summary.dev_plus
    ws['B16'] = structural_summary.dev_o
    ws['B17'] = structural_summary.dev_vplus
    ws['B18'] = structural_summary.dev_v

    # 3. Form Quality
    ws['B22'] = structural_summary.fqx_plus
    ws['B23'] = structural_summary.fqx_o
    ws['B24'] = structural_summary.fqx_u
    ws['B25'] = structural_summary.fqx_minus
    ws['B26'] = structural_summary.fqx_none
    ws['C22'] = structural_summary.mq_plus
    ws['C23'] = structural_summary.mq_o
    ws['C24'] = structural_summary.mq_u
    ws['C25'] = structural_summary.mq_minus
    ws['C26'] = structural_summary.mq_none
    ws['D22'] = structural_summary.wd_plus
    ws['D23'] = structural_summary.wd_o
    ws['D24'] = structural_summary.wd_u
    ws['D25'] = structural_summary.wd_minus
    ws['D26'] = structural_summary.wd_none

    # 4. Determinants - blends
    blends = structural_summary.blends.split(',')
    start_row = 6
    start_column = 6  # F열
    for blend in blends:
        ws.cell(row=start_row, column=start_column, value=blend)
        start_row += 1

    # 4. Determinants - single
    row = 6  # 시작 행
    for field_name in real_field:
        field_value = getattr(structural_summary, field_name)
        ws.cell(row=row, column=8, value=field_value)
        ws.cell(row=row, column=8, value=field_value).number_format = ",,0,"
        row += 1

    # 5. Contents
    row = 5
    for field_name in cont_real_fields:
        field_value = getattr(structural_summary, field_name)
        ws.cell(row=row, column=11, value=field_value)
        ws.cell(row=row, column=11, value=field_value).number_format = ",,0,"
        row += 1

    # 6. approach
    ws['N5'] = structural_summary.app_I
    ws['N6'] = structural_summary.app_II
    ws['N7'] = structural_summary.app_III
    ws['N8'] = structural_summary.app_IV
    ws['N9'] = structural_summary.app_V
    ws['N10'] = structural_summary.app_VI
    ws['N11'] = structural_summary.app_VII
    ws['N12'] = structural_summary.app_VIII
    ws['N13'] = structural_summary.app_IX
    ws['N14'] = structural_summary.app_X

    # 7. special scores
    lv1_row = 18
    for field_name in sp_real_fields:
        field_value = getattr(structural_summary, field_name)
        ws.cell(row=lv1_row, column=14, value=field_value)
        lv1_row += 1
    lv2_row = 18
    for field_name in sp_real_fields2:
        field_value = getattr(structural_summary, field_name)
        ws.cell(row=lv2_row, column=15, value=field_value)
        lv2_row += 1
    ws['N24'] = structural_summary.sum6
    ws['N25'] = structural_summary.wsum6
    ws['N26'] = structural_summary.sp_ab
    ws['N27'] = structural_summary.sp_ag
    ws['N28'] = structural_summary.sp_cop
    ws['N29'] = structural_summary.sp_cp
    ws['P26'] = structural_summary.sp_ghr
    ws['P27'] = structural_summary.sp_phr
    ws['P28'] = structural_summary.sp_mor
    ws['P29'] = structural_summary.sp_per
    ws['P30'] = structural_summary.sp_psv

    # 열 너비 조정
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) * 1.1 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length

    # 8. core
    wsd['B4'] = structural_summary.R
    wsd['D4'] = structural_summary.L
    wsd['B5'] = structural_summary.ErleBnistypus
    wsd['B6'] = structural_summary.eb
    wsd['D5'] = structural_summary.EA
    wsd['D6'] = structural_summary.es
    wsd['D7'] = structural_summary.adj_es
    wsd['F5'] = 'NA' if structural_summary.EBper == 0 else structural_summary.EBper
    wsd['F6'] = structural_summary.D_score
    wsd['F7'] = structural_summary.adj_D
    wsd['B8'] = structural_summary.sum_FM
    wsd['B9'] = structural_summary.sum_m
    wsd['D8'] = structural_summary.sum_Ca
    wsd['D9'] = structural_summary.sum_V
    wsd['F8'] = structural_summary.sum_T
    wsd['F9'] = structural_summary.sum_Y

    # 9. Affect
    wsd['I4'] = structural_summary.f_c_prop
    wsd['I4'].alignment = Alignment(horizontal='right')
    wsd['I5'] = structural_summary.pure_c
    wsd['I6'] = structural_summary.ca_c_prop
    wsd['I6'].alignment = Alignment(horizontal='right')
    wsd['I7'] = structural_summary.afr
    wsd['I7'].number_format = "0.##;-0.##;0"
    wsd['I8'] = structural_summary.S
    wsd['I9'] = structural_summary.blends_r
    wsd['I9'].alignment = Alignment(horizontal='right')
    wsd['I10'] = structural_summary.sp_cp

    # 10. Interpersonal
    wsd['L4'] = structural_summary.sp_cop
    wsd['N4'] = structural_summary.sp_ag
    wsd['M5'] = structural_summary.GHR_PHR
    wsd['M5'].alignment = Alignment(horizontal='right')
    wsd['M6'] = structural_summary.a_p
    wsd['M6'].alignment = Alignment(horizontal='right')
    wsd['M7'] = structural_summary.Fd_l
    wsd['M8'] = structural_summary.sum_T
    wsd['M9'] = structural_summary.human_cont
    wsd['M10'] = structural_summary.H
    wsd['M11'] = structural_summary.sp_per
    wsd['M12'] = structural_summary.Isol

    # 11. Ideation
    wsd['B15'] = structural_summary.a_p
    wsd['B15'].alignment = Alignment(horizontal='right')
    wsd['B16'] = structural_summary.Ma_Mp
    wsd['B16'].alignment = Alignment(horizontal='right')
    wsd['B17'] = structural_summary.intel
    wsd['B18'] = structural_summary.sp_mor
    wsd['D15'] = structural_summary.sum6
    wsd['D16'] = structural_summary.Lvl_2
    wsd['D17'] = structural_summary.wsum6
    wsd['D18'] = structural_summary.mq_minus
    wsd['D19'] = structural_summary.mq_none

    # 12. Mediation
    row = 15
    for field_name in med_real_fields:
        field_value = getattr(structural_summary, field_name)
        wsd.cell(row=row, column=7, value=field_value)
        wsd.cell(row=row, column=7).number_format = "0.##;-0.##;#;"
        row += 1

    # 13. Processing
    row = 15
    for field_name in pro_real_fields:
        field_value = getattr(structural_summary, field_name)
        wsd.cell(row=row, column=10, value=field_value)
        wsd.cell(row=row, column=10).alignment = Alignment(horizontal='right')
        row += 1

    # 14. Self-Perception
    row = 15
    for field_name in self_real_fields:
        field_value = getattr(structural_summary, field_name)
        wsd.cell(row=row, column=13, value=field_value)
        row += 1

    # 15. Special Indices
    wsd['A23'] = f"PTI={structural_summary.sumPTI}"
    wsd['B23'] = "☑" if structural_summary.sumDEPI >= 5 else "☐"
    wsd['B23'].alignment = Alignment(horizontal='right')
    wsd['C23'] = f"DEPI={structural_summary.sumDEPI}"
    wsd['D23'] = "☑" if structural_summary.sumCDI >= 4 else "☐"
    wsd['D23'].alignment = Alignment(horizontal='right')
    wsd['E23'] = f"CDI={structural_summary.sumCDI}"
    wsd['F23'] = "☑" if structural_summary.sumSCON >= 8 else "☐"
    wsd['F23'].alignment = Alignment(horizontal='right')
    wsd['G23'] = f"S-CON={structural_summary.sumSCON}"
    wsd['H23'] = "☑ HVI" if structural_summary.HVI_premise is True and structural_summary.sumHVI >= 4 else "☐ HVI"
    wsd['H23'].alignment = Alignment(horizontal='right')
    wsd['J23'] = "☑ OBS" if structural_summary.OBS_posi else "☐ OBS"
    wsd['J23'].alignment = Alignment(horizontal='right')

    # 열 너비 조정
    for column_cells in wsd.columns:
        length = max(len(str(cell.value)) * 1.1 for cell in column_cells)
        wsd.column_dimensions[column_cells[0].column_letter].width = length

    import pandas as pd

    # 특수지표 시트
    wsp['A1'] = "PTI"
    wsp['A2'] = "XA%<.70 AND WDA%<.75"
    wsp['A3'] = "X-%>0.29"
    wsp['A4'] = "LVL2>2 AND FAB2>0"
    wsp['A5'] = "R<17 AND Wsum6>12 OR R>16 AND Wsum6>17*"
    wsp['A6'] = "M- > 1 OR X-% > 0.40"
    wsp['A7'] = "TOTAL"
    row = 2
    for i in range(0, 5):
        value = "✔" if structural_summary.PTI[i] == "o" else ''
        wsp.cell(row=row, column=2, value=value)
        row += 1
    wsp['B7'] = structural_summary.sumPTI

    wsp['A9'] = "DEPI"
    wsp['A10'] = "SumV>0 OR FD>2"
    wsp['A11'] = "Col-shd blends>0 OR S>2"
    wsp['A12'] = "ego sup AND Fr+rF=0 OR ego inf"
    wsp['A13'] = "Afr<0.46 OR Blends<4"
    wsp['A14'] = "SumShd>FM+m OR SumC'>2"
    wsp['A15'] = "MOR>2 OR INTELL>3"
    wsp['A16'] = "COP<2 OR ISOL>0.24"
    wsp['A17'] = "TOTAL"
    wsp['A18'] = "POSITIVE?"
    row = 10
    for i in range(0, 7):
        value = "✔" if structural_summary.DEPI[i] == "o" else ''
        wsp.cell(row=row, column=2, value=value)
        row += 1
    wsp['B17'] = structural_summary.sumDEPI
    wsp['B18'] = structural_summary.sumDEPI >= 5

    wsp['A20'] = "CDI"
    wsp['A21'] = "EA<6 OR Daj<0"
    wsp['A22'] = "COP<2 AND AG<2"
    wsp['A23'] = "WSumC<2.5 OR Afr<0.46"
    wsp['A24'] = "p > a+1 OR pure H<2"
    wsp['A25'] = "SumT>1 OR ISOL>0.24 OR Fd>0"
    wsp['A26'] = "TOTAL"
    wsp['A27'] = "POSITIVE?"
    row = 21
    for i in range(0, 5):
        value = "✔" if structural_summary.CDI[i] == "o" else ''
        wsp.cell(row=row, column=2, value=value)
        row += 1
    wsp['B26'] = structural_summary.sumCDI
    wsp['B27'] = structural_summary.sumCDI >= 4

    wsp['D1'] = "S-CON"
    wsp['D2'] = "SumV+FD>2"
    wsp['D3'] = "col-shd blends>0"
    wsp['D4'] = "ego <0.31 ou >0.44"
    wsp['D5'] = "mor>3"
    wsp['D6'] = "Zd>3.5 ou <-3.5"
    wsp['D7'] = "es>EA"
    wsp['D8'] = "CF+C>FC"
    wsp['D9'] = "X+%<0.70"
    wsp['D10'] = "S>3"
    wsp['D11'] = "P<3 OU P>8"
    wsp['D12'] = "PURE H<2"
    wsp['D13'] = "R<17"
    wsp['D14'] = 'TOTAL'
    wsp['D15'] = 'POSITIVE?'
    row = 2
    for i in range(0, 12):
        value = "✔" if structural_summary.SCON[i] == "o" else ''
        wsp.cell(row=row, column=5, value=value)
        row += 1
    wsp['E14'] = structural_summary.sumSCON
    wsp['E15'] = structural_summary.sumSCON >= 8

    wsp['D17'] = 'HVI'
    wsp['D18'] = 'SumT = 0'
    wsp['D19'] = "Zf>12"
    wsp['D20'] = "Zd>3.5"
    wsp['D21'] = "S>3"
    wsp['D22'] = "H+(H)+Hd+(Hd)>6"
    wsp['D23'] = "(H)+(A)+(Hd)+(Ad)>3"
    wsp['D24'] = "H+A : 4:1"
    wsp['D25'] = "Cg>3"
    wsp['D26'] = 'TOTAL'
    wsp['D26'] = 'TOTAL'
    wsp['D27'] = 'POSITIVE?'
    wsp['E18'] = structural_summary.HVI_premise
    row = 19
    for i in range(0, 7):
        value = "✔" if structural_summary.HVI[i] == "o" else ''
        wsp.cell(row=row, column=5, value=value)
        row += 1
    wsp['E26'] = structural_summary.sumHVI
    wsp['E27'] = structural_summary.sumHVI >= 4 and structural_summary.HVI_premise

    wsp['A29'] = 'OBS'
    wsp['A30'] = 1
    wsp['A31'] = 2
    wsp['A32'] = 3
    wsp['A33'] = 4
    wsp['A34'] = 5
    wsp['B30'] = "Dd>3"
    wsp['B31'] = "Zf>12"
    wsp['B32'] = "Zd>3.0"
    wsp['B33'] = "P>7"
    wsp['B34'] = "FQ+>1"
    wsp['D30'] = "1-5 are true"
    wsp['D31'] = "FQ+>3 AND 2 items 1-4"
    wsp['D32'] = "X+%>0,89 et 3 items"
    wsp['D33'] = "FQ+>3 et X+%>0,89"
    wsp['D34'] = 'POSITIVE?'
    row = 30
    for i in range(0, 5):
        value = "✔" if structural_summary.OBS[i] == "o" else ''
        wsp.cell(row=row, column=3, value=value)
        row += 1
    row = 30
    for i in range(5, 9):
        value = "✔" if structural_summary.OBS[i] == "o" else ''
        wsp.cell(row=row, column=5, value=value)
        row += 1
    wsp['E34'] = structural_summary.OBS_posi

    # 열 너비 조정
    for column_cells in wsp.columns:
        length = max(len(str(cell.value)) * 1.1 for cell in column_cells)
        wsp.column_dimensions[column_cells[0].column_letter].width = length

    # Create a DataFrame from ResponseCode data
    response_codes = ResponseCode.objects.filter(client=client_id)
    response_code_data = []

    for response_code in response_codes:
        response_code_data.append({
            'Card': response_code.card,
            'N': response_code.response_num,
            'time': response_code.time,
            'response': response_code.response,
            'V': response_code.rotation,
            'inquiry': response_code.inquiry,
            'Location': response_code.location,
            'loc_num': response_code.loc_num,
            'Dev Qual': response_code.dev_qual,
            'determinants': response_code.determinants,
            'Form Quality': response_code.form_qual,
            '(2)': response_code.pair,
            'Content': response_code.content,
            'P': response_code.popular,
            'Z': response_code.Z,
            'special': response_code.special,
            'comment': response_code.comment
        })
    response_code_df = pd.DataFrame(response_code_data)

    # Create a new worksheet in the Excel workbook
    ws2 = wb.create_sheet(title="raw data")

    data_values = response_code_df.values.tolist()

    # Add column headers to the new worksheet
    for col_num, column_title in enumerate(response_code_df.columns, 1):
        cell = ws2.cell(row=1, column=col_num, value=column_title)
        cell.font = Font(bold=True)

    # Add data to the new worksheet
    for row_data in data_values:
        ws2.append(row_data)

    # 엑셀 파일 저장
    response = HttpResponse(save_virtual_workbook(wb),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=structural_summary.xlsx'

    return response


def edit_responses(request, client_id):
    # Get existing ResponseCode objects for the client_id
    response_codes = ResponseCode.objects.filter(client_id=client_id)

    # Create a formset with the existing ResponseCode objects
    ResponseCodeFormSet = modelformset_factory(ResponseCode, form=ResponseCodeForm, extra=40, max_num=70)
    formset = ResponseCodeFormSet(queryset=response_codes)

    if request.method == 'POST':
        formset = ResponseCodeFormSet(request.POST, queryset=response_codes)
        if formset.is_valid():
            # Save the modified forms
            for form in formset:
                if form.cleaned_data.get('card') and form.cleaned_data.get('response'):
                    form.instance.client_id = client_id
                    form.save()
            structural_summary, created = StructuralSummary.objects.get_or_create(client_id=client_id)
            structural_summary.save()
            return redirect('client_list')
        else:
            # Handle formset errors
            for form in formset:
                for field in form:
                    print(field.errors)

    return render(request, 'edit_responses.html', {'formset': formset, 'client_id': client_id})
